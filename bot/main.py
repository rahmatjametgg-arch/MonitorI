import httpx
from bs4 import BeautifulSoup
import re
from datetime import datetime, timedelta, timezone
import time
import zipfile
import threading
import json
import os
import hashlib
import socket
import phonenumbers
from phonenumbers import geocoder
import requests
import signal
# ── Database layer (PostgreSQL persistence) ──────────────────────────────────
try:
    from db import db_init, db_save, db_save_async, db_load
    _DB_ENABLED = True
except ImportError:
    def db_init(): pass
    def db_save(k, v): pass
    def db_save_async(k, v): pass
    def db_load(k, default=None): return default
    _DB_ENABLED = False
import sys
from langdetect import detect, LangDetectException, DetectorFactory
from colorama import init, Fore, Style
import qrcode
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed

# Real-time console output — nonaktifkan buffering stdout/stderr
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

# ================= CONSOLE LOGGER =================
def _log(tag, msg, color=Fore.CYAN):
    print(color + f"  [{tag:<10}] {msg}" + Style.RESET_ALL, flush=True)

def _print_banner(bot_username=""):
    name = f"@{bot_username}" if bot_username else "SPIDERMAT OTP BOT"
    line = "─" * 44
    print(Fore.CYAN + Style.BRIGHT + f"\n  {line}")
    print(f"   SPIDERMAT OTP BOT  {name}")
    print(f"  {line}" + Style.RESET_ALL)

# ================= MULTI WORKER PROXY (auto-failover saat kena rate limit) =================
# Urutan = urutan prioritas. Worker#1 dipakai terus SELAMA tidak kena limit.
# Hanya pindah ke worker berikutnya saat worker aktif terdeteksi rate-limited (Cloudflare
# error 1027/1015/429/dst) — sesuai permintaan: "wajib hanya ke ubah kalau kena limit aja".
WORKER_POOL = [
    "https://plain-butterfly-d9e9.kicenivas.workers.dev",
    "https://ivasmunchen.serverprivate1.web.id",
    "https://ivasmsbykicenv2.kikixrakaofficial.biz.id",
    "https://ivasbykiven.alwayskixyzshop.web.id",
]

_worker_lock          = threading.Lock()
_active_worker_idx    = 0
_worker_limited_until = {}     # worker_url -> timestamp sampai kapan dianggap masih limited
WORKER_LIMIT_COOLDOWN = 900    # 15 menit sebelum worker yang limit dicoba lagi

def get_active_worker():
    with _worker_lock:
        return WORKER_POOL[_active_worker_idx % len(WORKER_POOL)]

def _common_headers():
    origin = _IVAS_ORIGIN
    return {
        "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Origin":           origin,
        "Referer":          f"{origin}/",
        "X-Forwarded-Host": _IVAS_HOST,
    }

def _apply_worker_globals(new_base):
    """
    Update SEMUA konstanta URL global ke worker baru.
    Sengaja pakai variabel global biasa (bukan fungsi) supaya seluruh kode lama
    yang sudah memakai BASE/LOGIN_URL/dst tetap otomatis ikut berubah tanpa
    perlu diedit satu-satu — karena Python me-lookup global by name saat dieksekusi.
    """
    global _IVAS_ORIGIN, _IVAS_HOST, _COMMON_HDR, BASE
    global LOGIN_URL, RECV_URL, GET_RANGE_URL, GET_NUMBER_URL, GET_SMS_URL
    global RETURN_NUMBER_URL, RETURN_ALL_URL, EXPORT_URL, _RECV_POST_HEADERS
    _IVAS_ORIGIN = new_base
    _IVAS_HOST   = new_base.split("://", 1)[-1]
    BASE         = new_base
    LOGIN_URL         = f"{BASE}/login"
    RECV_URL          = f"{BASE}/portal/sms/received"
    GET_RANGE_URL     = f"{BASE}/portal/sms/received/getsms"
    GET_NUMBER_URL    = f"{BASE}/portal/sms/received/getsms/number"
    GET_SMS_URL       = f"{BASE}/portal/sms/received/getsms/number/sms"
    RETURN_NUMBER_URL = f"{BASE}/portal/numbers/return/number"
    RETURN_ALL_URL    = f"{BASE}/portal/numbers/return/allnumber/bluck"
    EXPORT_URL        = f"{BASE}/portal/numbers/export"
    _COMMON_HDR = _common_headers()
    _RECV_POST_HEADERS = {
        "Accept":           "text/html,*/*;q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
        "Referer":          RECV_URL,
        "Origin":           BASE,
    }

def _refresh_all_session_headers():
    """Update header Origin/Referer/X-Forwarded-Host di semua session HTTP yang sudah aktif."""
    fresh_hdr = _COMMON_HDR
    try:
        with accounts_lock:
            accs = list(accounts) + list(_premium_acc_cache.values())
    except Exception:
        accs = list(accounts) if 'accounts' in globals() else []
    for acc in accs:
        sess = acc.get("session") if isinstance(acc, dict) else None
        if sess is not None:
            try:
                sess.headers.update(fresh_hdr)
            except Exception:
                pass

def mark_worker_limited(worker_url):
    """Dipanggil ketika worker aktif terdeteksi kena rate-limit — pindah otomatis ke worker lain."""
    global _active_worker_idx
    now = time.time()
    with _worker_lock:
        _worker_limited_until[worker_url] = now + WORKER_LIMIT_COOLDOWN
        n = len(WORKER_POOL)
        chosen = None
        for i in range(1, n + 1):
            idx = (_active_worker_idx + i) % n
            candidate = WORKER_POOL[idx]
            if _worker_limited_until.get(candidate, 0) < now:
                chosen = idx
                break
        if chosen is None:
            chosen = (_active_worker_idx + 1) % n
        _active_worker_idx = chosen
        new_base = WORKER_POOL[chosen]
    _apply_worker_globals(new_base)
    _refresh_all_session_headers()
    _log("WORKER", f"⚠️  {worker_url} rate-limited → pindah ke {new_base}", Fore.YELLOW)
    return new_base

def maybe_recover_primary_worker():
    """Kalau worker#1 (utama) sudah lewat cooldown, coba balik lagi ke situ."""
    global _active_worker_idx
    now = time.time()
    with _worker_lock:
        if _active_worker_idx == 0:
            return
        primary = WORKER_POOL[0]
        if _worker_limited_until.get(primary, 0) >= now:
            return
        _active_worker_idx = 0
        new_base = primary
    _apply_worker_globals(new_base)
    _refresh_all_session_headers()
    _log("WORKER", f"✅ worker utama {new_base} pulih — pindah balik", Fore.GREEN)

_RATE_LIMIT_MARKERS = (
    "temporarily rate limited", "error 1027", "please check back later",
    "has been rate limited", "error 1015", "you have been blocked",
    "attention required", "error 1020", "sorry, you have been blocked",
    "checking your browser", "just a moment",
)

def is_worker_blocked(resp=None, exc=None) -> bool:
    """Deteksi tanda-tanda worker/proxy kena limit atau block dari Cloudflare."""
    if exc is not None:
        return True
    if resp is None:
        return False
    try:
        if resp.status_code in (429):
            return True
        sample = resp.text[:2000].lower()
        if any(marker in sample for marker in _RATE_LIMIT_MARKERS):
            return True
    except Exception:
        pass
    return False

_IVAS_ORIGIN  = WORKER_POOL[0]
_IVAS_HOST    = _IVAS_ORIGIN.split("://", 1)[-1]
_COMMON_HDR   = _common_headers()

def make_httpx_client(timeout=30):
    return httpx.Client(
        follow_redirects=True,
        timeout=timeout,
        headers=_COMMON_HDR,
        # Limits dilebarkan supaya banyak request get_sms paralel per-akun
        # (banyak nomor/OTP sekaligus) tidak saling nunggu slot koneksi.
        limits=httpx.Limits(max_connections=100, max_keepalive_connections=40),
    )

def make_requests_session():
    s = requests.Session()
    s.headers.update(_COMMON_HDR)
    return s

# ================= FILES =================
ACCOUNTS_FILE = "accounts.json"
COOKIES_FILE = "cookie.json"
CACHE_FILE = "file/sent_cache.json"
MAX_CACHE_SIZE = 2000
LANG_CODE_MAP = {  
    "id": "#Indonesia", "en": "#English", "fr": "#French", "es": "#Spanish",  
    "pt": "#Portuguese", "ar": "#Arabic", "ru": "#Russian", "tr": "#Turkish",  
    "hi": "#Hindi", "th": "#Thai", "vi": "#Vietnamese", "ms": "#Malay",  
    "tl": "#Filipino", "ja": "#Japanese", "ko": "#Korean", "zh-cn": "#Chinese",  
    "nl": "#Dutch", "sv": "#Swedish", "pl": "#Polish", "uk": "#Ukrainian",  
    "cs": "#Czech", "ro": "#Romanian", "el": "#Greek", "he": "#Hebrew", "fa": "#Persian"  
}  
    
# ================= CONFIG =================
# ⚙️  Semua nilai sensitif dibaca dari environment variable (via .env atau secret)
# Salin file .env.example → .env lalu isi nilainya sebelum menjalankan bot.

OWNER_ID   = int(os.getenv("OWNER_ID", "0"))           # Telegram user-id owner (integer)
BOT_TOKEN  = os.getenv("BOT_TOKEN", "")                 # Token dari @BotFather

# BASE, LOGIN_URL, RECV_URL, dst sudah didefinisikan dinamis di blok MULTI WORKER PROXY di atas
# (mengikuti worker aktif dari WORKER_POOL) — jangan didefinisikan ulang statis di sini.
_apply_worker_globals(WORKER_POOL[_active_worker_idx])

GROUPS_FILE = "groups.json"
ADDNUM_API_URL = "https://ws.websocket.web.id/admin/addnumber"
ADDNUM_API_KEY = os.getenv("ADDNUM_API_KEY", "112231")
USERS_FILE = "users.json"
PREMIUM_FILE = "premium.json"
AMBIL_FILE = "file/ambil_nomor.json"
PREMIUM_COOKIE_FILE = "premium-cookie.json"
LINK_OWNER   = os.getenv("LINK_OWNER",   "t.me/owner_anda")
LINK_CHANNEL = os.getenv("LINK_CHANNEL", "https://t.me/channel_anda")

# ================= LOG & FORCE JOIN =================
LOG_CHANNEL_ID = int(os.getenv("LOG_CHANNEL_ID", "0"))  # ID channel log (angka negatif untuk grup/channel)
BOT_USERNAME   = ""   # diisi otomatis saat startup via getMe

# Format: "username1,username2" → pisah koma, tanpa @
_FORCE_JOIN_RAW = os.getenv("FORCE_JOIN_CHANNELS", "")
FORCE_JOIN_CHANNELS = []
for _ch in [c.strip() for c in _FORCE_JOIN_RAW.split(",") if c.strip()]:
    FORCE_JOIN_CHANNELS.append({"username": _ch, "url": f"https://t.me/{_ch}", "label": f"📢 {_ch}"})

# ================= PAKASIR PAYMENT =================
PAKASIR_PROJECT = os.getenv("PAKASIR_PROJECT", "")
PAKASIR_API_KEY  = os.getenv("PAKASIR_API_KEY", "")
PAKASIR_BASE     = "https://app.pakasir.com"

# Harga paket: {tier: {durasi_hari: harga_rupiah}}
PACKAGE_PRICES = {
    "starter": {7: 25_000,  15: 50_000,  30: 100_000},
    "pro":     {7: 75_000,  15: 140_000, 30: 250_000},
    "elite":   {7: 150_000, 15: 280_000, 30: 500_000},
    "ultra":   {7: 250_000, 15: 480_000, 30: 900_000},
}
DURATION_INFO = {
    7:  {"label": "7 Hari",  "emoji": "📅"},
    15: {"label": "15 Hari", "emoji": "📆"},
    30: {"label": "30 Hari", "emoji": "🗓️"},
}
QRIS_PATH = "./qris.jpg"  # Foto QRIS statis untuk pembayaran manual

SERVICE_SHORT = {
    "WHATSAPP": "#WS", "TELEGRAM": "#TG", "GOOGLE": "#G", "FACEBOOK": "#FB",
    "INSTAGRAM": "#IG", "SHOPEE": "#SP", "TOKOPEDIA": "#TP", "GRAB": "#GR",
    "GOJEK": "#GJ", "TIKTOK": "#TT"
}
sms_stats = {
    "total_sms": 0,
    "total_otp": 0,
    "total_number": set()
}
last_update_id = 0
MAX_EMAIL = 20 # Setting Max Email User/Owner
TOKEN_TIERS = {
    "free":    {"label": "FREE",    "emoji": "👤", "tokens_day": 3,     "max_email": 1},
    "starter": {"label": "PREMIUM", "emoji": "⭐", "tokens_day": 100,   "max_email": 5},
    "pro":     {"label": "PRO",     "emoji": "💎", "tokens_day": 300,   "max_email": 10},
    "elite":   {"label": "ELITE",   "emoji": "🔥", "tokens_day": 500,   "max_email": 15},
    "ultra":   {"label": "ULTRA",   "emoji": "👑", "tokens_day": 99999, "max_email": 20},
}
DetectorFactory.seed = 0
init(autoreset=True)
accounts_lock = threading.Lock()
LOGIN_COOLDOWN = 300  # 5 menit
SESSION_RETRY_INTERVAL = 600  # retry setiap 10 menit kalau session gagal

# ================= TELEGRAM SESSION (persistent + retry) =================
_TG_SESSION = requests.Session()
_TG_ADAPTER = requests.adapters.HTTPAdapter(
    pool_connections=4, pool_maxsize=10, max_retries=0
)
_TG_SESSION.mount("https://", _TG_ADAPTER)

def _tg_request(method, data=None, json_data=None, files=None, timeout=12):
    """
    Kirim request ke Telegram API via persistent session (connection pooling).
    Retry max 3x dengan exponential backoff + auto-handle 429 Flood Wait.
    """
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    for attempt in range(3):
        try:
            r = _TG_SESSION.post(url, data=data, json=json_data, files=files, timeout=timeout)
            if r.status_code == 429:
                retry_after = r.json().get("parameters", {}).get("retry_after", 5)
                _log("TG-429", f"{method} — tunggu {retry_after}s", Fore.YELLOW)
                time.sleep(retry_after + 1)
                continue
            return r
        except Exception as e:
            if attempt == 2:
                _log("TG-ERR", f"{method} — {e}", Fore.RED)
            else:
                time.sleep(1.5 ** (attempt + 1))
    return None

pending_setcookie = {}   # user_id -> {"email": str, "msg_id": int}
pending_addcookie = {}   # user_id -> {"email": str, "msg_id": int}
pending_addnum    = {}   # user_id -> {"email": str, "msg_id": int}
pending_payments  = {}   # user_id -> {order_id, tier, days, amount, chat_id, qr_msg_id}

# ── Auto-delete tracker untuk /cekivas /cekrange /toprcv /rangeterbaru ──
_last_cek_msgs  : dict = {}   # chat_id -> msg_id pesan bot sebelumnya
_last_cek_lock  = threading.Lock()

# ================= SESSION TRACKER =================
_session_fail_time   = {}   # email -> timestamp pertama kali gagal
_session_notified    = {}   # email -> bool sudah notif atau belum
_session_retry_time  = {}   # email -> timestamp terakhir retry
_session_recovered   = {}   # email -> bool sudah notif recover

# ================= AUTO COOKIE REFRESHER =================
COOKIE_KEEPALIVE_INTERVAL = 600   # keepalive tiap 10 menit (sebelum session sempat expire)
COOKIE_NOTIF_COOLDOWN     = 3600  # notif ulang maks 1x per jam per akun
_last_cookie_refresh      = {}    # email -> timestamp terakhir keepalive
_last_cookie_notif        = {}    # email -> timestamp terakhir notif dikirim
_keepalive_warn_count     = {}    # email -> jumlah gagal keepalive berturut-turut

# ================= RANGES CACHE (kurangi beban IVAS server) =================
_ranges_cache    = {}   # email -> (timestamp, ranges_list)
RANGES_CACHE_TTL = 300  # 5 menit — ranges jarang berubah

# ================= RECV CSRF CACHE =================
# iVAS pakai per-page CSRF — /portal/sms/received punya token berbeda dari /portal
# Semua POST ke getsms, getsms/number, getsms/number/sms WAJIB pakai recv_csrf ini
_recv_csrf_cache = {}   # email -> {"csrf": str, "ts": float}
RECV_CSRF_TTL    = 900  # 15 menit — refresh sebelum expired

# ================= AUTO BACKUP =================
# Direktori & pola yang TIDAK perlu dibackup (sistem/cache/package)
BACKUP_SKIP_DIRS = {
    ".git", ".pythonlibs", ".local", ".cache",
    ".agents", ".upm", "nix", "__pycache__",
}
BACKUP_SKIP_EXTS = {".pyc", ".pyo", ".zip"}
BACKUP_SKIP_FILES = {".replit", "replit.nix"}

# ================= EXPIRY NOTIFIER =================
_notif_expiry_sent = {}   # str(uid) -> set {"24h", "3h", "1h"}

# ================= THREAD-SAFE CACHE & SHARED STATE =================
_sent_cache_lock  = threading.Lock()
_cookie_file_lock = threading.Lock()   # lindungi read-modify-write cookie files dari banyak thread
_cache_dirty      = False
_last_cache_save  = 0.0
# Dibaca worker threads — diupdate oleh run_bot() manager thread
_bot_state = {"email_to_uid": {}, "total_accounts": 0}
# Flag untuk memaksa run_bot sync segera (set True setelah addcookie/setcookie berhasil)
_force_bot_sync   = False

# ================= ACCOUNT MANAGEMENT =================
def load_accounts():
    # Coba dari file dulu, fallback ke DB
    if os.path.exists(ACCOUNTS_FILE):
        try:
            with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                saved = data.get("accounts", [])
                if saved:
                    return saved
        except:
            pass
    # Fallback ke database
    saved = db_load("accounts", [])
    if saved:
        try:
            with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
                json.dump({"accounts": saved}, f, indent=2)
        except:
            pass
    return saved

def save_accounts():
    data_to_save = []
    for acc in accounts:
        data_to_save.append({
            "email": acc.get("email"),
            "password": acc.get("password"),
            "cookies": acc.get("cookies", {})
        })
    try:
        with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
            json.dump({"accounts": data_to_save}, f, indent=2)
    except Exception as e:
        print(f"[SAVE] accounts file error: {e}")
    db_save_async("accounts", data_to_save)

def load_cookies():
    if os.path.exists(COOKIES_FILE):
        try:
            with open(COOKIES_FILE, "r", encoding="utf-8") as f:
                d = json.load(f)
                if d:
                    return d
        except:
            pass
    saved = db_load("cookies", {})
    if saved:
        try:
            with open(COOKIES_FILE, "w", encoding="utf-8") as f:
                json.dump(saved, f, indent=2)
        except:
            pass
    return saved

def load_premium():
    if os.path.exists(PREMIUM_FILE):
        try:
            with open(PREMIUM_FILE, "r") as f:
                d = json.load(f)
                if d:
                    return d
        except:
            pass
    saved = db_load("premium", {})
    if saved:
        try:
            with open(PREMIUM_FILE, "w") as f:
                json.dump(saved, f, indent=2)
        except:
            pass
    return saved

def save_premium(data):
    try:
        with open(PREMIUM_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print(f"[SAVE] premium file error: {e}")
    db_save_async("premium", data)

premium_users = load_premium()

def get_user_tier(user_id):
    if user_id == OWNER_ID:
        return "ultra"
    user = premium_users.get(str(user_id))
    if not user:
        return "free"
    if time.time() > user.get("expired", 0):
        return "free"
    return user.get("tier", "free")

def get_tier_tokens_day(user_id):
    return TOKEN_TIERS.get(get_user_tier(user_id), TOKEN_TIERS["free"])["tokens_day"]

def get_tier_email_limit(user_id):
    if user_id == OWNER_ID:
        return MAX_EMAIL
    return TOKEN_TIERS.get(get_user_tier(user_id), TOKEN_TIERS["free"])["max_email"]

def is_premium(user_id):
    return get_user_tier(user_id) != "free"
    
def save_cookies(cookies_dict):
    try:
        with open(COOKIES_FILE, "w") as f:
            json.dump(cookies_dict, f, indent=2)
    except Exception as e:
        print(f"[SAVE] cookies file error: {e}")
    db_save_async("cookies", cookies_dict)

def extract_session_cookies(session):
    """Ekstrak semua cookies dari httpx session sebagai dict (fresh cookies)."""
    try:
        return dict(session.cookies)
    except:
        return {}

def save_fresh_cookies_auto(email, fresh_cookies):
    """Simpan fresh cookies ke file yang sesuai (owner/premium) berdasarkan email."""
    if not fresh_cookies:
        return
    with _cookie_file_lock:
        all_cookies = load_cookies()
        if email in all_cookies:
            all_cookies[email] = fresh_cookies
            save_cookies(all_cookies)
            return
        prem = load_premium_cookies()
        if email in prem:
            prem[email] = fresh_cookies
            save_premium_cookies(prem)

def load_groups():
    if os.path.exists(GROUPS_FILE):
        try:
            with open(GROUPS_FILE, "r") as f:
                data = json.load(f)
            if not (isinstance(data, list) or "groups" in data) and data:
                return data
        except:
            pass
    saved = db_load("groups", {})
    if saved:
        try:
            with open(GROUPS_FILE, "w") as f:
                json.dump(saved, f, indent=2)
        except:
            pass
    return saved

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                d = json.load(f)
                if d:
                    return d
        except:
            pass
    saved = db_load("users", {})
    if saved:
        try:
            with open(USERS_FILE, "w") as f:
                json.dump(saved, f, indent=2)
        except:
            pass
    return saved
        
def load_premium_cookies():
    if os.path.exists(PREMIUM_COOKIE_FILE):
        try:
            with open(PREMIUM_COOKIE_FILE, "r") as f:
                d = json.load(f)
                if d:
                    return d
        except:
            pass
    saved = db_load("premium_cookies", {})
    if saved:
        try:
            with open(PREMIUM_COOKIE_FILE, "w") as f:
                json.dump(saved, f, indent=2)
        except:
            pass
    return saved

def save_premium_cookies(data):
    try:
        with open(PREMIUM_COOKIE_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print(f"[SAVE] premium_cookies file error: {e}")
    db_save_async("premium_cookies", data)

premium_cookies = load_premium_cookies()        

def save_users(data):
    try:
        with open(USERS_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        print(f"[SAVE] users file error: {e}")
    db_save_async("users", data)

def save_groups():
    try:
        with open(GROUPS_FILE, "w") as f:
            json.dump(user_groups, f, indent=2)
    except Exception as e:
        print(f"[SAVE] groups file error: {e}")
    db_save_async("groups", user_groups)

user_groups = load_groups()

# ===== PER-USER GROUP HELPERS =====
def get_user_groups(user_id):
    return user_groups.get(str(user_id), [])

def add_user_group(user_id, gid):
    uid = str(user_id)
    if uid not in user_groups:
        user_groups[uid] = []
    if gid not in user_groups[uid]:
        user_groups[uid].append(gid)
    save_groups()

def remove_user_group(user_id, gid):
    uid = str(user_id)
    if uid in user_groups and gid in user_groups[uid]:
        user_groups[uid].remove(gid)
        save_groups()
        return True
    return False

# ===== USER KEY (IDENTITY) =====
def generate_user_key(user_id):
    raw = f"LESEH-{user_id}-IVAS-SECRET"
    h = hashlib.md5(raw.encode()).hexdigest()[:8].upper()
    return f"KX-{h[:4]}-{h[4:]}"

def get_or_create_user_key(user_id):
    if user_id == OWNER_ID:
        return generate_user_key(user_id)
    users = load_users()
    uid = str(user_id)
    changed = False
    if uid not in users:
        users[uid] = {"emails": [], "key": generate_user_key(user_id)}
        changed = True
    elif "key" not in users.get(uid, {}):
        users[uid]["key"] = generate_user_key(user_id)
        changed = True
    if changed:
        save_users(users)
    return users[uid]["key"]

# ================= TOKEN SYSTEM =================
TOKEN_MAX = 3  # token default free user per hari (reset jam 00:00 WIB)

def get_wib_date():
    """Return tanggal hari ini dalam WIB (UTC+7)."""
    from datetime import timezone, timedelta
    tz_wib = timezone(timedelta(hours=7))
    return datetime.now(tz_wib).strftime("%Y-%m-%d")

def get_user_tokens(user_id):
    """Ambil sisa token user hari ini. Auto-reset jika hari baru (WIB)."""
    if user_id == OWNER_ID:
        return 99999
    uid = str(user_id)
    users = load_users()
    today = get_wib_date()
    daily_limit = get_tier_tokens_day(user_id)
    if uid not in users:
        users[uid] = {"emails": [], "tokens": daily_limit, "last_token_reset": today}
        save_users(users)
        return daily_limit
    u = users[uid]
    if u.get("last_token_reset") != today:
        u["tokens"] = daily_limit
        u["last_token_reset"] = today
        save_users(users)
    return u.get("tokens", daily_limit)

def use_token(user_id):
    """Kurangi 1 token. Return True berhasil, False kalau habis. Owner unlimited."""
    if user_id == OWNER_ID:
        return True
    uid = str(user_id)
    users = load_users()
    today = get_wib_date()
    daily_limit = get_tier_tokens_day(user_id)
    if uid not in users:
        users[uid] = {"emails": [], "tokens": daily_limit, "last_token_reset": today}
    u = users[uid]
    if u.get("last_token_reset") != today:
        u["tokens"] = daily_limit
        u["last_token_reset"] = today
    if u.get("tokens", 0) <= 0:
        save_users(users)
        return False
    u["tokens"] = u.get("tokens", daily_limit) - 1
    save_users(users)
    return True

def token_status_str(user_id):
    """Return string singkat sisa token user untuk ditampilkan."""
    if user_id == OWNER_ID:
        return "♾️ Unlimited"
    t = get_user_tokens(user_id)
    daily = get_tier_tokens_day(user_id)
    return f"🎫 {t}/{daily}"

def no_token_msg(chat_id):
    send_msg(chat_id,
        "❌ <b>Token habis!</b>\n\n"
        "<blockquote>Token kamu sudah habis hari ini.\n"
        "Reset otomatis jam <b>00:00 WIB</b>.\n\n"
        "💡 Upgrade ke PREMIUM untuk token lebih banyak!\n"
        "Ketik /beli untuk lihat paket.</blockquote>"
    )

# ===== PREMIUM ACCOUNT SESSION CACHE =====
_premium_acc_cache = {}  # email -> acc dict (persistent session untuk premium user)

def get_acc_by_email(email):
    """Cari akun berdasarkan email — cek accounts (owner) dan _premium_acc_cache (user)."""
    with accounts_lock:
        for a in accounts:
            if a.get("email") == email:
                return a
    acc = _premium_acc_cache.get(email)
    if acc:
        return acc
    prem_cookies = load_premium_cookies()
    if email in prem_cookies:
        new_acc = {
            "email": email, "password": None,
            "cookies": prem_cookies[email],
            "session": make_httpx_client(),
            "last_login": 0,
            "csrf_token": "",
        }
        new_acc["session"].cookies.update(prem_cookies[email])
        _premium_acc_cache[email] = new_acc
        return new_acc
    return None

# ===== USERNAME CACHE (in-memory, diisi saat user kirim pesan) =====
_username_cache = {}  # user_id (int) -> "@username" atau "uid:xxx"

_MAX_USERNAME_CACHE = 500

def store_username(user_id, from_obj):
    """Simpan username dari objek 'from' Telegram ke cache (max 500 entry)."""
    if len(_username_cache) >= _MAX_USERNAME_CACHE:
        # Hapus entry paling lama (FIFO)
        try:
            _username_cache.pop(next(iter(_username_cache)))
        except Exception:
            pass
    uname = from_obj.get("username")
    if uname:
        _username_cache[user_id] = f"@{uname}"
    else:
        first = from_obj.get("first_name", "")
        _username_cache[user_id] = first if first else f"uid:{user_id}"

def get_user_display(user_id):
    """Ambil label display user untuk console log."""
    return _username_cache.get(user_id, f"uid:{user_id}")

def load_sent_cache():
    os.makedirs("file", exist_ok=True)
    if not os.path.exists(CACHE_FILE):
        return set()
    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data) if isinstance(data, list) else set()
    except:
        return set()

def save_sent_cache():
    try:
        os.makedirs("file", exist_ok=True)
        cache_list = list(sent_cache)
        if len(cache_list) > MAX_CACHE_SIZE:
            cache_list = cache_list[-MAX_CACHE_SIZE:]
            sent_cache.clear()
            sent_cache.update(cache_list)
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache_list, f)
        db_save_async("sent_cache", cache_list)
    except Exception as e:
        print(f"Error save cache: {e}")

def save_sent_cache_debounced():
    """Tandai cache perlu disimpan; flush max sekali per 5 detik."""
    global _cache_dirty, _last_cache_save
    _cache_dirty = True
    if time.time() - _last_cache_save >= 5:
        try:
            with _sent_cache_lock:
                save_sent_cache()
            _last_cache_save = time.time()
            _cache_dirty = False
        except Exception as e:
            _log("WARN", f"save cache: {e}", Fore.YELLOW)

# ================= LOAD DATA =================
accounts = load_accounts()
cookies_data = load_cookies()
sent_cache = load_sent_cache()

for acc in accounts:
    acc["session"] = make_httpx_client()
    acc["last_login"] = 0
    acc["csrf_token"] = "" 

    email = acc["email"]
    if email in cookies_data:
        acc["cookies"] = cookies_data[email]
        acc["session"].cookies.update(cookies_data[email])

# ================= ACCOUNT COMMANDS =================
def add_account(text, chat_id=None, user_id=None):
    try:
        parts = text.split()
        if len(parts) < 3:
            msg = "❌ Format:\n<code>/addakun email password</code>"
            if chat_id: send_msg(chat_id, msg)
            else: tg_active("  Format: /addakun email password")
            return

        email, password = parts[1].strip().lower(), parts[2].strip()

        with accounts_lock:
            for acc in accounts:
                if acc["email"] == email:
                    msg = f"❌ Akun sudah ada: <code>{email}</code>"
                    if chat_id: send_msg(chat_id, msg)
                    else: tg_active(f"  Akun sudah ada: {email}")
                    return

            acc = {
                "email": email,
                "password": password,
                "cookies": {},
                "session": make_httpx_client(),
                "last_login": 0,
                "csrf_token": ""
            }
            accounts.append(acc)
            save_accounts()

        if chat_id:
            send_msg(chat_id, f"⏳ Mencoba login ke <code>{email}</code>...")

        if login(acc):
            acc["last_login"] = time.time()
            msg = (
                f"✅ <b>Akun aktif &amp; login:</b>\n<code>{email}</code>"
            )
            if chat_id: send_msg(chat_id, msg)
            else: tg_active(f"  Akun aktif & login: {email}")
            # Langsung cek ulang SEMUA akun lain — jangan sampai ada yang
            # diam-diam ikut ke-invalidate tanpa ketahuan.
            _recheck_other_accounts_async(email, chat_id)
        else:
            if chat_id and user_id:
                guide_msg_id = send_msg(chat_id,
                    f"⚠️ <b>Login otomatis gagal</b> untuk <code>{email}</code>\n\n"
                    f"<blockquote>Password mungkin salah atau akun butuh verifikasi.\n"
                    f"Pasang cookie manual di bawah agar akun langsung aktif 👇</blockquote>\n\n"
                    + _cookie_guide_text("SET COOKIE — OWNER", email)
                )
                pending_setcookie[user_id] = {"email": email, "msg_id": guide_msg_id}
            else:
                tg_active(f"   Akun masuk tapi login gagal: {email}")

    except Exception as e:
        msg = f"❌ Error add akun: <code>{e}</code>"
        if chat_id: send_msg(chat_id, msg)
        else: tg_active(f"  Error add akun: {e}")

def del_account(text):
    """Legacy text-based — dipertahankan untuk kompatibilitas console."""
    try:
        _, email = text.split()
        global accounts
        accounts = [a for a in accounts if a["email"] != email]
        save_accounts()
        with _cookie_file_lock:
            cookies_d = load_cookies()
            if email in cookies_d:
                del cookies_d[email]
                save_cookies(cookies_d)
        tg_active(f"  Akun dihapus: {email}")
    except:
        tg_active("  Format salah /delakun email")

# ================= DELAKUN OWNER FLOW (interactive) =================
def command_delakun(chat_id, user_id):
    if not accounts:
        return send_msg(chat_id, "❌ Belum ada akun yang terdaftar.")
    with accounts_lock:
        acc_list = list(accounts)
    buttons = [{"text": f"🗑️ {acc['email']}", "callback_data": f"dok:{acc['email']}"} for acc in acc_list]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:dok"})
    send_inline_keyboard(chat_id,
        "🗑️ <b>HAPUS AKUN — OWNER</b>\n\n"
        "<blockquote>⚠️ Pilih akun yang ingin dihapus.\n"
        "Session, cookie, dan data akun akan langsung dihentikan.\n\n"
        "Aksi ini <b>tidak bisa dibatalkan</b>!</blockquote>\n\n"
        "👇 Pilih akun:",
        buttons)

def handle_delakun_select_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⚠️ Konfirmasi dulu!")
    with accounts_lock:
        found = any(a["email"] == email for a in accounts)
    if not found:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>HAPUS AKUN</b>\n\n❌ Akun tidak ditemukan.")
        return
    buttons = [
        {"text": "✅ Ya, Hapus",  "callback_data": f"dokc:{email}"},
        {"text": "❌ Batalkan",   "callback_data": "cancel:dok"},
    ]
    delete_and_send_inline(chat_id, msg_id,
        f"🗑️ <b>HAPUS AKUN — KONFIRMASI</b>\n\n"
        f"<blockquote>"
        f"📧 Akun: <code>{email}</code>\n\n"
        f"❗ Cookie, session, dan semua data akun ini akan dihapus permanen.\n"
        f"Apakah kamu yakin?</blockquote>",
        buttons)

def handle_delakun_confirm_cb(chat_id, user_id, email, cb_id, msg_id):
    global accounts
    answer_callback_query(cb_id, "⏳ Menghapus akun...")
    with accounts_lock:
        found = any(a["email"] == email for a in accounts)
    if not found:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>HAPUS AKUN</b>\n\n❌ Akun tidak ditemukan.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"🗑️ <b>HAPUS AKUN — OWNER</b>\n\n"
        f"<blockquote>"
        f"📧 Akun: <code>{email}</code>\n\n"
        f"⏳ Menghentikan session &amp; menghapus data..."
        f"</blockquote>")

    try:
        # 1. Hapus dari accounts list
        with accounts_lock:
            accounts = [a for a in accounts if a["email"] != email]
        save_accounts()

        # 2. Hapus dari cookie.json (owner cookie) & premium-cookie.json — dikunci
        #    supaya tidak nabrak/menimpa cookie akun lain yang sedang ditulis bersamaan
        with _cookie_file_lock:
            cookies_d = load_cookies()
            if email in cookies_d:
                del cookies_d[email]
                save_cookies(cookies_d)

            prem_cookies = load_premium_cookies()
            if email in prem_cookies:
                del prem_cookies[email]
                save_premium_cookies(prem_cookies)

        # 4. Bersihkan semua cache terkait
        _premium_acc_cache.pop(email, None)
        _ranges_cache.pop(email, None)
        _recv_csrf_cache.pop(email, None)
        _last_cookie_refresh.pop(email, None)
        _last_cookie_notif.pop(email, None)
        _keepalive_warn_count.pop(email, None)
        _session_fail_time.pop(email, None)
        _session_notified.pop(email, None)
        _session_retry_time.pop(email, None)
        _session_recovered.pop(email, None)

        # 5. Bersihkan pending state
        pending_addnum.pop(user_id, None)
        pending_addcookie.pop(user_id, None)
        pending_setcookie.pop(user_id, None)

        delete_and_send(chat_id, proc_id,
            f"🗑️ <b>HAPUS AKUN BERHASIL</b>\n\n"
            f"<blockquote>"
            f"📧 Akun: <code>{email}</code>\n"
            f"✅ Akun berhasil dihapus\n"
            f"✅ Cookie &amp; session dihentikan\n"
            f"✅ Semua data akun dibersihkan"
            f"</blockquote>")

    except Exception as e:
        _log("DELAKUN", f"error: {e}", Fore.RED)
        try:
            delete_and_send(chat_id, proc_id,
                f"🗑️ <b>HAPUS AKUN ERROR</b>\n\n"
                f"<blockquote>"
                f"📧 Akun: <code>{email}</code>\n"
                f"❌ Error: <code>{str(e)[:200]}</code>"
                f"</blockquote>")
        except Exception:
            pass
        
def save_number(number):
    try:
        with open(AMBIL_FILE, "r") as f:
            data = json.load(f)
    except:
        data = {"numbers": []}

    if number not in data["numbers"]:
        data["numbers"].append(number)

    with open(AMBIL_FILE, "w") as f:
        json.dump(data, f, indent=2)

def ambilnomor_to_txt():
    try:
        with open(AMBIL_FILE, "r") as f:
            data = json.load(f)
            numbers = data.get("numbers", [])
    except:
        numbers = []

    if not numbers:
        return None

    filename = "file/nomor.txt"
    with open(filename, "w") as f:
        for n in numbers:
            f.write(f"{n}\n") 

    return filename

def export_numbers_ivas(chat_id, acc, status_msg_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    email = acc["email"] if isinstance(acc, dict) else acc

    def _status(text):
        if status_msg_id:
            delete_and_send(chat_id, status_msg_id, text)
        else:
            send_msg(chat_id, text)

    # Gunakan session yang sudah terautentikasi dari acc
    if isinstance(acc, dict) and acc.get("session"):
        session = acc["session"]
    else:
        # Fallback: buat session baru dengan cookies
        if isinstance(acc, dict):
            cookies = acc.get("cookies") or {}
            if not cookies:
                all_cookies = load_cookies()
                prem_cookies = load_premium_cookies()
                cookies = all_cookies.get(email) or prem_cookies.get(email) or {}
        else:
            all_cookies = load_cookies()
            prem_cookies = load_premium_cookies()
            cookies = all_cookies.get(email) or prem_cookies.get(email) or {}

        if not cookies:
            _status(
                f"📁 <b>AMBIL FILE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Cookie tidak ditemukan. Set cookie dulu."
                f"</blockquote>")
            return
        session = make_requests_session()
        session.cookies.update(cookies)

    now      = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{now}_ivas_numbers_{email.split('@')[0]}.xlsx"
    os.makedirs("file", exist_ok=True)
    filepath = f"file/{filename}"

    export_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept":     "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer":    f"{BASE}/portal/numbers",
    }

    try:
        # ── Coba endpoint asli /portal/numbers/export terlebih dahulu ──
        # Pakai header browser biasa (BUKAN X-Requested-With) agar server tidak reject
        export_headers_real = {
            "User-Agent": export_headers["User-Agent"],
            "Accept":     "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer":    f"{BASE}/portal/numbers",
        }
        _log("AMBILFILE", f"coba endpoint asli: {EXPORT_URL}", Fore.CYAN)
        r_export = session.get(EXPORT_URL, headers=export_headers_real, timeout=90)

        content_type = r_export.headers.get("Content-Type", "")
        content_disp = r_export.headers.get("Content-Disposition", "")

        got_file = (
            r_export.status_code == 200
            and "/login" not in str(getattr(r_export, "url", ""))
            and (
                "spreadsheet" in content_type
                or "application/octet-stream" in content_type
                or "attachment" in content_disp
                or "csv" in content_type
                or "excel" in content_type
            )
        )

        if got_file:
            # Endpoint asli mengembalikan file langsung
            raw_bytes = r_export.content
            # Tentukan ekstensi dari content-disposition atau content-type
            if "csv" in content_type or ".csv" in content_disp:
                ext = "csv"
                filename = filename.replace(".xlsx", ".csv")
                filepath = filepath.replace(".xlsx", ".csv")
            else:
                ext = "xlsx"

            with open(filepath, "wb") as f_out:
                f_out.write(raw_bytes)
            total_size = len(raw_bytes)

            if status_msg_id:
                delete_msg(chat_id, status_msg_id)

            with open(filepath, "rb") as f:
                requests.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                    data={
                        "chat_id":    chat_id,
                        "caption": (
                            f"📁 <b>FILE IVAS BERHASIL DIAMBIL</b>\n\n"
                            f"<blockquote>"
                            f"📧 Email  : <code>{email}</code>\n"
                            f"📄 File   : <code>{filename}</code>\n"
                            f"📦 Ukuran : <b>{total_size // 1024} KB</b>\n"
                            f"🕐 Waktu  : <code>{now}</code>\n"
                            f"✅ Sumber : Endpoint Asli IVAS"
                            f"</blockquote>"
                        ),
                        "parse_mode": "HTML"
                    },
                    files={"document": (filename, f)}
                )
            os.remove(filepath)
            _log("AMBILFILE", "berhasil via endpoint asli", Fore.GREEN)
            return

        # ── Fallback: DataTable JSON → build Excel manual ──
        _log("AMBILFILE", "endpoint asli tidak mengembalikan file, fallback DataTable", Fore.YELLOW)

        dt_headers = {
            "User-Agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept":           "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
            "Referer":          f"{BASE}/portal/numbers",
        }
        dt_params_base = {
            "columns[0][data]": "number_id",         "columns[0][name]": "id",
            "columns[1][data]": "Number",             "columns[1][name]": "Number",
            "columns[2][data]": "range",              "columns[2][name]": "range",
            "columns[3][data]": "A2P",                "columns[3][name]": "A2P",
            "columns[4][data]": "LimitA2P",           "columns[4][name]": "LimitA2P",
            "columns[5][data]": "limit_cli_a2p",      "columns[5][name]": "limit_cli_a2p",
            "columns[6][data]": "limit_cli_did_a2p",  "columns[6][name]": "limit_cli_did_a2p",
            "columns[7][data]": "action",             "columns[7][name]": "action",
            "order[0][column]": 1, "order[0][dir]": "desc",
            "search[value]": "", "search[regex]": "false",
        }

        r_check = session.get(f"{BASE}/portal/numbers",
                              params={**dt_params_base, "draw": 1, "start": 0, "length": 1},
                              headers=dt_headers, timeout=30)

        if "/login" in str(getattr(r_check, "url", "")) or r_check.status_code != 200:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Session tidak valid. Cookie expired?\n"
                f"📄 HTTP {r_check.status_code}"
                f"</blockquote>")
            return

        try:
            meta = r_check.json()
        except Exception:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Gagal parse respon server. Cookie mungkin expired.\n"
                f"📄 Respon: <code>{r_check.text[:200]}</code>"
                f"</blockquote>")
            return

        total = int(meta.get("recordsTotal", 0))
        if total == 0:
            _status(
                f"📁 <b>AMBIL FILE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"⚠️ Tidak ada nomor aktif untuk di-export."
                f"</blockquote>")
            return

        r_all = session.get(f"{BASE}/portal/numbers",
                            params={**dt_params_base, "draw": 2, "start": 0, "length": total},
                            headers=dt_headers, timeout=120)

        if r_all.status_code != 200:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Server error: HTTP {r_all.status_code}"
                f"</blockquote>")
            return

        try:
            data = r_all.json().get("data", [])
        except Exception:
            _status(
                f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Gagal parse data dari server."
                f"</blockquote>")
            return

        if not data:
            _status(
                f"📁 <b>AMBIL FILE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"⚠️ Data kosong, coba lagi."
                f"</blockquote>")
            return

        # Build Excel dari DataTable data
        wb = Workbook()
        ws = wb.active
        ws.title = "Numbers"

        header_row  = ["No", "Number", "Range", "Rate (A2P)", "Limit by Range", "SID/Range Limit", "SID→DID Limit"]
        header_fill = PatternFill("solid", fgColor="2D6A9F")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(data, 1):
            num_raw = row.get("Number", "")
            try:
                num_str = str(int(num_raw))
            except Exception:
                num_str = str(num_raw)
            ws.append([
                i, num_str,
                re.sub(r"<[^>]+>", "", str(row.get("range", ""))).strip(),
                row.get("A2P", ""),
                row.get("LimitA2P", ""),
                row.get("limit_cli_a2p", ""),
                row.get("limit_cli_did_a2p", ""),
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(filepath)

        if status_msg_id:
            delete_msg(chat_id, status_msg_id)

        with open(filepath, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data={
                    "chat_id":    chat_id,
                    "caption": (
                        f"📁 <b>FILE IVAS BERHASIL DIAMBIL</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email  : <code>{email}</code>\n"
                        f"🔢 Total  : <b>{len(data)}</b> nomor\n"
                        f"📄 File   : <code>{filename}</code>\n"
                        f"🕐 Waktu  : <code>{now}</code>\n"
                        f"✅ Sumber : DataTable API"
                        f"</blockquote>"
                    ),
                    "parse_mode": "HTML"
                },
                files={"document": (filename, f)}
            )
        os.remove(filepath)
        _log("AMBILFILE", f"berhasil via DataTable: {len(data)} nomor", Fore.GREEN)

    except Exception as e:
        _status(
            f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"❌ Error: <code>{str(e)[:200]}</code>"
            f"</blockquote>")
        
def del_account(text):
    try:
        _, email = text.split()
        global accounts
        accounts = [a for a in accounts if a["email"] != email]
        save_accounts()
        tg_active(f"  Akun dihapus: {email}")
    except:
        tg_active("  Format salah /delakun email")

def detect_language(text):
    try:
        if not text or len(text) < 10: return "#Unknown"
        text = re.sub(r"\d+", "", text).strip()
        if len(text) < 5: return "#Unknown"
        lang_code = detect(text)
        return LANG_CODE_MAP.get(lang_code, f"#{lang_code.upper()}")
    except LangDetectException:
        return "#Unknown"
        
def list_accounts(chat_id, user_id):
    try:
        if not accounts:
            send_msg(chat_id, "Belum ada akun")
            return
        msg = "  <b>LIST AKUN</b>\n\n"
        now = time.time()
        for i, acc in enumerate(accounts, 1):
            email = acc.get("email", "Unknown")
            safe_email = email if user_id == OWNER_ID else mask_email(email)
            last_login = acc.get("last_login", 0)
            status = "ACTIVE  " if now - last_login < LOGIN_COOLDOWN else "OFFLINE  "
            msg += f"{i}. {safe_email} | {status}\n"
        send_msg(chat_id, msg)
    except Exception as e:
        send_msg(chat_id, f"  Error list akun: {e}")
        
def add_token_tier(text, chat_id):
    """Owner: /addtoken user_id tier hari"""
    try:
        parts = text.split()
        if len(parts) < 4:
            return send_msg(chat_id,
                "❌ Format:\n<code>/addtoken user_id tier hari</code>\n\n"
                "Tier: <b>starter / pro / elite / ultra</b>\n"
                "Contoh: <code>/addtoken 123456789 pro 30</code>")
        uid, tier, hari = parts[1], parts[2].lower(), int(parts[3])
        if tier not in TOKEN_TIERS or tier == "free":
            return send_msg(chat_id, "❌ Tier tidak valid!\nPilih: starter / pro / elite / ultra")
        expired = time.time() + (hari * 86400)
        premium_users[str(uid)] = {"tier": tier, "expired": expired}
        save_premium(premium_users)
        # Reset token user ke limit tier baru — fix bug token tetap 0 saat ganti role
        users_d = load_users()
        new_limit = TOKEN_TIERS[tier]["tokens_day"]
        if str(uid) not in users_d:
            users_d[str(uid)] = {}
        users_d[str(uid)]["tokens"] = new_limit
        users_d[str(uid)]["last_token_reset"] = get_wib_date()
        save_users(users_d)
        t = TOKEN_TIERS[tier]
        tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']}/hari"
        send_msg(chat_id,
            f"✅ <b>PAKET TOKEN AKTIF</b>\n\n"
            f"<blockquote>"
            f"👤 User ID  : <code>{uid}</code>\n"
            f"🏷️ Paket   : {t['emoji']} <b>{t['label']}</b>\n"
            f"🎫 Token   : {tok_str}\n"
            f"📧 Max Email: {t['max_email']} akun\n"
            f"📅 Durasi  : {hari} hari"
            f"</blockquote>")
        try:
            send_msg(int(uid),
                f"🎉 <b>PAKET TOKEN AKTIF!</b>\n\n"
                f"<blockquote>"
                f"🏷️ Paket    : {t['emoji']} <b>{t['label']}</b>\n"
                f"🎫 Token    : {tok_str}\n"
                f"📧 Max Email: {t['max_email']} akun\n"
                f"📅 Durasi   : {hari} hari\n\n"
                f"Token reset otomatis jam 00:00 WIB.\n"
                f"Ketik /cekprem untuk cek status."
                f"</blockquote>")
        except: pass
    except Exception as e:
        send_msg(chat_id, f"❌ Error: {e}")
        
def add_cookie_premium(text, chat_id, user_id):
    cmd_addcookie(chat_id, user_id)  
        
def del_cookie_premium(text, chat_id, user_id):
    try:
        parts = text.split()
        if len(parts) < 2:
            return send_msg(chat_id, "❌ Format:\n/delcookie email@gmail.com")
        email = parts[1].strip().lower()

        # Cek kepemilikan: email harus milik user ini (atau owner)
        if not is_owner(user_id):
            users_d = load_users()
            owned = users_d.get(str(user_id), {}).get("emails", [])
            if email not in owned:
                return send_msg(chat_id, "❌ Akun tidak ditemukan di akun kamu")

        # Hapus dari premium-cookie.json — dikunci supaya tidak nabrak akun lain
        with _cookie_file_lock:
            premium_cookies = load_premium_cookies()
            if email in premium_cookies:
                del premium_cookies[email]
                save_premium_cookies(premium_cookies)

        # Hapus dari user_accounts (password-based) di users.json
        users_d = load_users()
        uid = str(user_id)
        if uid in users_d:
            before = users_d[uid].get("user_accounts", [])
            users_d[uid]["user_accounts"] = [a for a in before if a.get("email") != email]
            # Hapus juga dari emails list
            if email in users_d[uid].get("emails", []):
                users_d[uid]["emails"].remove(email)
            save_users(users_d)

        send_msg(chat_id, f"✅ Akun dihapus:\n<code>{email}</code>")
    except Exception as e:
        send_msg(chat_id, f"❌ Error: {e}")                    
        
def del_token_tier(text, chat_id):
    """Owner: /deltoken user_id"""
    try:
        _, uid = text.split()
        if uid not in premium_users:
            return send_msg(chat_id, "❌ User tidak memiliki paket aktif")
        del premium_users[uid]
        save_premium(premium_users)
        send_msg(chat_id, f"✅ Paket user <code>{uid}</code> dihapus → kembali ke FREE")
        try:
            send_msg(int(uid), "⚠️ Paket token kamu telah dinonaktifkan oleh owner.")
        except: pass
    except:
        send_msg(chat_id, "❌ Format:\n/deltoken user_id")

def cmd_resettoken(text, chat_id):
    """Owner: /resettoken user_id — reset token user ke limit tier sekarang"""
    try:
        parts = text.split()
        if len(parts) < 2:
            return send_msg(chat_id,
                "❌ Format:\n<code>/resettoken user_id</code>\n\n"
                "Contoh: <code>/resettoken 123456789</code>")
        uid = parts[1].strip()
        users_d = load_users()
        tier = premium_users.get(uid, {}).get("tier", "free")
        new_limit = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])["tokens_day"]
        if uid not in users_d:
            users_d[uid] = {}
        users_d[uid]["tokens"] = new_limit
        users_d[uid]["last_token_reset"] = get_wib_date()
        save_users(users_d)
        t = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
        tok_str = "♾️ Unlimited" if new_limit >= 99999 else str(new_limit)
        send_msg(chat_id,
            f"✅ <b>Token direset!</b>\n\n"
            f"<blockquote>"
            f"👤 User ID : <code>{uid}</code>\n"
            f"🏷️ Tier    : {t['emoji']} {t['label']}\n"
            f"🎫 Token   : {tok_str}/hari"
            f"</blockquote>")
        try:
            send_msg(int(uid),
                f"🔄 <b>Token kamu direset oleh owner!</b>\n\n"
                f"<blockquote>🏷️ Tier  : {t['emoji']} {t['label']}\n"
                f"🎫 Token : {tok_str}/hari</blockquote>")
        except: pass
    except Exception as e:
        send_msg(chat_id, f"❌ Error: {e}")

def is_owner(user_id): return user_id == OWNER_ID

def list_token_tier(chat_id):
    """Owner: /listtoken"""
    if not premium_users:
        return send_msg(chat_id, "Belum ada user dengan paket aktif.")
    now = time.time()
    msg = "🏆 <b>LIST PAKET TOKEN AKTIF</b>\n\n"
    for i, (uid, data) in enumerate(premium_users.items(), 1):
        tier = data.get("tier", "starter")
        t = TOKEN_TIERS.get(tier, TOKEN_TIERS["starter"])
        sisa = max(0, int((data.get("expired", 0) - now) // 86400))
        msg += f"{i}. <code>{uid}</code> | {t['emoji']} {t['label']} | {sisa} hari\n"
    send_msg(chat_id, msg)

def cmd_beli(chat_id, user_id):
    tier = get_user_tier(user_id)
    t_cur = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
    aktif_tag = f"  ✅ aktif" if tier != "free" else ""
    msg = (
        "🛒 <b>PAKET PREMIUM SPIDERMAT BOT</b>\n\n"
        "<blockquote>"
        f"Paket kamu saat ini: {t_cur['emoji']} <b>{t_cur['label']}</b>{aktif_tag}\n\n"
        "⭐ <b>PREMIUM</b> — 100 token/hari, 5 akun IVAS\n"
        "  📅 7 Hari   : <b>Rp 25.000</b>\n"
        "  📆 15 Hari  : <b>Rp 50.000</b>\n"
        "  🗓️ 30 Hari  : <b>Rp 100.000</b>\n\n"
        "Bayar via QRIS — konfirmasi ke owner setelah transfer."
        "</blockquote>\n\n"
        "👇 Pilih durasi:"
    )
    prices_s = PACKAGE_PRICES.get("starter", {})
    rows = [
        [{"text": f"📅 7 Hari — Rp 25.000",  "callback_data": "pkg_buy:starter:7"}],
        [{"text": f"📆 15 Hari — Rp 50.000", "callback_data": "pkg_buy:starter:15"}],
        [{"text": f"🗓️ 30 Hari — Rp 100.000","callback_data": "pkg_buy:starter:30"}],
    ]
    send_inline_keyboard_grid(chat_id, msg, rows)

def handle_pkg_info_cb(chat_id, user_id, tier_key, cb_id, msg_id):
    answer_callback_query(cb_id)
    if tier_key not in TOKEN_TIERS or tier_key == "free":
        return
    t = TOKEN_TIERS[tier_key]
    tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']} token/hari"
    tier_now = get_user_tier(user_id)
    aktif_label = " ✅ <i>(paket kamu sekarang)</i>" if tier_now == tier_key else ""
    prices = PACKAGE_PRICES.get(tier_key, {})

    price_lines = ""
    for days, info in DURATION_INFO.items():
        harga = prices.get(days, 0)
        price_lines += f"  {info['emoji']} {info['label']:9s}: <b>Rp {harga:,}</b>\n".replace(",", ".")

    detail = (
        f"{t['emoji']} <b>PAKET {t['label']}</b>{aktif_label}\n\n"
        f"<blockquote>"
        f"🎫 Token/hari  : <b>{tok_str}</b>\n"
        f"📧 Max Email   : <b>{t['max_email']} akun</b> IVAS\n"
        f"🔄 Reset Token : 00:00 WIB\n\n"
        f"✅ <b>Semua fitur aktif:</b>\n"
        f"  • /addcookie — kelola cookie IVAS\n"
        f"  • /addemail — tambah akun IVAS\n"
        f"  • /addnum — tambah nomor test\n"
        f"  • /delnumall — kembalikan semua nomor\n"
        f"  • /myrange — cek range aktif\n"
        f"  • /ambilfile — export nomor ke Excel\n"
        f"  • /cekivas — traffic WhatsApp per negara (gratis)\n"
        f"  • /cekivasv2 — cek SMS berdasarkan rentang waktu (gratis)\n"
        f"  • /toprcv [negara] — detail range + kode per negara (gratis)\n"
        f"  • /rangeterbaru — range paling baru terima SMS (gratis)\n"
        f"  • SMS notif otomatis ke grup/PM\n\n"
        f"💰 <b>Harga:</b>\n"
        f"{price_lines}"
        f"</blockquote>\n\n"
        f"👇 Pilih durasi untuk lanjut bayar via QRIS:"
    )
    rows = []
    for days, info in DURATION_INFO.items():
        harga = prices.get(days, 0)
        rows.append([{
            "text": f"{info['emoji']} {info['label']} — Rp {harga:,}".replace(",", "."),
            "callback_data": f"pkg_buy:{tier_key}:{days}"
        }])
    rows.append([{"text": "🔙 Lihat Paket Lain", "callback_data": "pkg_back"}])
    delete_msg(chat_id, msg_id)
    send_inline_keyboard_grid(chat_id, detail, rows)


# ================= PAKASIR PAYMENT FUNCTIONS =================

def pakasir_create_qris(order_id, amount):
    try:
        r = requests.post(
            f"{PAKASIR_BASE}/api/transactioncreate/qris",
            json={"project": PAKASIR_PROJECT, "order_id": order_id,
                  "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=15
        )
        data = r.json()
        return data.get("payment")
    except Exception as e:
        print(f"[PAKASIR] create error: {e}")
        return None

def pakasir_check_status(order_id, amount):
    try:
        r = requests.get(
            f"{PAKASIR_BASE}/api/transactiondetail",
            params={"project": PAKASIR_PROJECT, "order_id": order_id,
                    "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=10
        )
        data = r.json()
        return data.get("transaction", {}).get("status", "unknown")
    except Exception as e:
        print(f"[PAKASIR] status error: {e}")
        return "unknown"

def pakasir_cancel(order_id, amount):
    try:
        requests.post(
            f"{PAKASIR_BASE}/api/transactioncancel",
            json={"project": PAKASIR_PROJECT, "order_id": order_id,
                  "amount": amount, "api_key": PAKASIR_API_KEY},
            timeout=10
        )
    except Exception as e:
        print(f"[PAKASIR] cancel error: {e}")

def generate_qr_image(qr_string):
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
                       box_size=10, border=4)
    qr.add_data(qr_string)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf

def send_photo_msg(chat_id, photo_bytes, caption, reply_markup=None):
    data = {"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"}
    if reply_markup:
        data["reply_markup"] = json.dumps(reply_markup)
    r = requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
        data=data,
        files={"photo": ("qris.png", photo_bytes, "image/png")},
        timeout=20
    )
    res = r.json()
    if res.get("ok"):
        return res["result"]["message_id"]
    return None

def payment_checker(user_id, chat_id, order_id, tier, days, amount, qr_msg_id):
    deadline = time.time() + 15 * 60  # 15 menit
    t = TOKEN_TIERS.get(tier, {})
    label = t.get("label", tier.upper())
    while time.time() < deadline:
        time.sleep(5)
        if pending_payments.get(user_id, {}).get("order_id") != order_id:
            return
        status = pakasir_check_status(order_id, amount)
        if status == "completed":
            pending_payments.pop(user_id, None)
            delete_msg(chat_id, qr_msg_id)
            _add_token_tier_direct(user_id, tier, days)
            emoji = t.get("emoji", "✅")
            exp_str = get_tier_expiry_str(user_id)
            send_msg(chat_id,
                f"✅ <b>PEMBAYARAN BERHASIL!</b>\n\n"
                f"Paket {emoji} <b>{label}</b> selama <b>{days} hari</b> telah aktif.\n"
                f"📅 Aktif hingga: <b>{exp_str}</b>\n\n"
                f"Selamat menggunakan SPIDERMAT OTP BOT! 🚀"
            )
            # Kirim laporan pembelian ke channel log
            def _purchase_log():
                uname = BOT_USERNAME
                now_str = datetime.now(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
                udisp_buy = get_user_display(user_id)
                msg_log = (
                    f"💰 <b>PEMBELIAN BERHASIL</b>\n\n"
                    f"<blockquote>"
                    f"👤 User    : {udisp_buy}\n"
                    f"🆔 ID      : <code>{user_id}</code>\n"
                    f"📦 Paket   : {emoji} <b>{label}</b>\n"
                    f"📅 Durasi  : <b>{days} hari</b>\n"
                    f"💵 Nominal : <b>Rp {amount:,}</b>\n".replace(",", ".") +
                    f"🕐 Waktu   : {now_str} WIB\n"
                    f"🔖 Order   : <code>{order_id}</code>"
                    f"</blockquote>"
                )
                markup = {"inline_keyboard": [[{"text": "🤖 Buka Bot", "url": f"https://t.me/{uname}"}]]} if uname else None
                try:
                    payload = {"chat_id": LOG_CHANNEL_ID, "text": msg_log, "parse_mode": "HTML"}
                    if markup:
                        payload["reply_markup"] = markup
                    requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage", json=payload, timeout=10)
                except Exception as le:
                    print(f"PURCHASE LOG ERROR: {le}")
            threading.Thread(target=_purchase_log, daemon=True).start()
            return
        elif status in ("expired", "cancelled"):
            pending_payments.pop(user_id, None)
            delete_msg(chat_id, qr_msg_id)
            send_msg(chat_id,
                f"⏰ <b>Pembayaran kadaluarsa.</b>\n\n"
                f"QR QRIS untuk paket <b>{label}</b> sudah tidak valid.\n"
                f"Ketik /beli untuk membuat tagihan baru."
            )
            return
    pending_payments.pop(user_id, None)
    delete_msg(chat_id, qr_msg_id)
    pakasir_cancel(order_id, amount)
    send_msg(chat_id,
        f"⏰ <b>Waktu bayar habis (15 menit).</b>\n\n"
        f"Tagihan paket <b>{label}</b> dibatalkan otomatis.\n"
        f"Ketik /beli untuk membuat tagihan baru."
    )

def get_tier_expiry_str(user_id):
    premium = load_premium()
    entry = premium.get(str(user_id), {})
    exp = entry.get("expired", 0)
    if not exp or exp == 0:
        return "-"
    try:
        # expired disimpan sebagai float timestamp (time.time())
        dt = datetime.fromtimestamp(float(exp))
        return dt.strftime("%d %b %Y %H:%M WIB")
    except Exception:
        return str(exp)


def _add_token_tier_direct(user_id, tier, days):
    """Helper internal — aktifkan paket tier langsung via parameter (bukan teks /addtoken)."""
    try:
        uid = str(user_id)
        if tier not in TOKEN_TIERS or tier == "free":
            return
        expired = time.time() + (int(days) * 86400)
        premium_users[uid] = {"tier": tier, "expired": expired}
        save_premium(premium_users)
        users_d = load_users()
        new_limit = TOKEN_TIERS[tier]["tokens_day"]
        if uid not in users_d:
            users_d[uid] = {}
        users_d[uid]["tokens"] = new_limit
        users_d[uid]["last_token_reset"] = get_wib_date()
        save_users(users_d)
        _log("TOKENTIER", f"tier {tier} aktif untuk {uid} selama {days} hari", Fore.GREEN)
    except Exception as e:
        _log("TOKENTIER", f"_add_token_tier_direct error: {e}", Fore.RED)

def handle_pkg_buy_cb(chat_id, user_id, data, cb_id, msg_id):
    answer_callback_query(cb_id)
    parts = data.split(":")
    if len(parts) != 2:
        return
    tier_key, days_str = parts
    try:
        days = int(days_str)
    except ValueError:
        return
    if tier_key not in PACKAGE_PRICES or days not in PACKAGE_PRICES[tier_key]:
        send_msg(chat_id, "❌ Paket tidak valid.")
        return
    amount = PACKAGE_PRICES[tier_key][days]
    t = TOKEN_TIERS.get(tier_key, {})
    label = t.get("label", tier_key.upper())
    dur_info = DURATION_INFO[days]
    ts = int(time.time())
    order_id = f"SPIDER{user_id}{ts}"

    delete_msg(chat_id, msg_id)

    # ── Mode QRIS Statis (tanpa Pakasir) ─────────────────────────────────────
    if not PAKASIR_PROJECT or not PAKASIR_API_KEY:
        caption = (
            f"📲 <b>PEMBAYARAN QRIS — {t.get('emoji','')} {label} {dur_info['label']}</b>\n\n"
            f"<blockquote>"
            f"💰 Nominal    : <b>Rp {amount:,}</b>\n".replace(",", ".") +
            f"📋 Order ID   : <code>{order_id}</code>\n\n"
            f"📝 <b>Cara Bayar:</b>\n"
            f"1. Scan QR di atas dengan m-banking / e-wallet\n"
            f"2. Transfer tepat <b>Rp {amount:,}</b> ke atas\n".replace(",", ".") +
            f"3. Screenshot bukti bayar, kirim ke owner\n"
            f"4. Owner akan aktivasi paket kamu manual\n"
            f"</blockquote>\n\n"
            f"📩 <b>Kirim bukti ke owner:</b> <a href='https://{LINK_OWNER}'>Contact Owner</a>"
        )
        try:
            if os.path.exists(QRIS_PATH):
                with open(QRIS_PATH, "rb") as f:
                    requests.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                        data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                        files={"photo": f},
                        timeout=15
                    )
            else:
                send_msg(chat_id, caption)
        except Exception as e:
            send_msg(chat_id, caption)
        return

    # ── Mode Pakasir (pembayaran otomatis) ────────────────────────────────────
    if user_id in pending_payments:
        old = pending_payments[user_id]
        send_msg(chat_id,
            f"⚠️ Kamu masih punya tagihan aktif (Order <code>{old['order_id']}</code>).\n"
            f"Selesaikan dulu atau batalkan dengan tombol di bawah QR sebelumnya."
        )
        return

    loading_msg = send_msg_return_id(chat_id,
        f"⏳ Membuat tagihan QRIS untuk paket <b>{t.get('emoji','')} {label}</b> "
        f"{dur_info['label']}...\nMohon tunggu sebentar."
    )

    payment = pakasir_create_qris(order_id, amount)
    if not payment:
        if loading_msg:
            delete_msg(chat_id, loading_msg)
        send_msg(chat_id, "❌ Gagal membuat tagihan. Coba lagi atau hubungi owner.")
        return

    qr_string   = payment.get("payment_number", "")
    total        = payment.get("total_payment", amount)
    expired_at   = payment.get("expired_at", "")
    try:
        exp_dt = datetime.fromisoformat(expired_at.replace("Z", "+00:00"))
        exp_str = exp_dt.strftime("%H:%M WIB, %d %b %Y")
    except Exception:
        exp_str = expired_at

    if loading_msg:
        delete_msg(chat_id, loading_msg)

    try:
        qr_buf = generate_qr_image(qr_string)
    except Exception as e:
        print(f"[QR] generate error: {e}")
        send_msg(chat_id, "❌ Gagal membuat QR code. Coba lagi.")
        return

    caption = (
        f"📲 <b>TAGIHAN QRIS — {t.get('emoji','')} {label} {dur_info['label']}</b>\n\n"
        f"<blockquote>"
        f"💰 Total Bayar : <b>Rp {total:,}</b>\n".replace(",", ".") +
        f"⏰ Berlaku s/d : <b>{exp_str}</b>\n\n"
        f"📋 Order ID    : <code>{order_id}</code>"
        f"</blockquote>\n\n"
        f"Scan QR di atas menggunakan aplikasi e-wallet / m-banking.\n"
        f"Paket aktif <b>otomatis</b> setelah pembayaran berhasil. ✅"
    )
    markup = {"inline_keyboard": [[
        {"text": "❌ Batalkan Pembayaran", "callback_data": f"cancel_payment:{order_id}:{amount}"}
    ]]}
    qr_msg_id = send_photo_msg(chat_id, qr_buf, caption, markup)
    if not qr_msg_id:
        send_msg(chat_id, "❌ Gagal mengirim QR. Coba lagi.")
        return

    pending_payments[user_id] = {
        "order_id": order_id, "tier": tier_key, "days": days,
        "amount": amount, "chat_id": chat_id, "qr_msg_id": qr_msg_id
    }
    threading.Thread(
        target=payment_checker,
        args=(user_id, chat_id, order_id, tier_key, days, amount, qr_msg_id),
        daemon=True
    ).start()

def send_msg_return_id(chat_id, text):
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=10
        )
        res = r.json()
        if res.get("ok"):
            return res["result"]["message_id"]
    except Exception:
        pass
    return None

def delete_msg(chat_id, msg_id):
    """Hapus pesan bot. Abaikan error (misal sudah dihapus atau terlalu lama)."""
    try:
        _tg_request("deleteMessage", data={"chat_id": chat_id, "message_id": msg_id})
    except Exception:
        pass

def send_msg(chat_id, text):
    _tg_request("sendMessage", data={"chat_id": chat_id, "text": text, "parse_mode": "HTML"})

def send_cek_msg(chat_id, text):
    """
    Kirim pesan hasil /cekivas /cekrange /toprcv /rangeterbaru.
    Otomatis hapus pesan bot sebelumnya di chat yang sama (anti-spam di grup).
    Return msg_id baru (int) atau None.
    """
    # Hapus pesan sebelumnya jika ada
    with _last_cek_lock:
        prev_id = _last_cek_msgs.get(chat_id)
    if prev_id:
        delete_msg(chat_id, prev_id)

    # Kirim pesan baru dan simpan msg_id-nya
    try:
        r = _tg_request("sendMessage", data={
            "chat_id":    chat_id,
            "text":       text,
            "parse_mode": "HTML",
        })
        if r and r.status_code == 200:
            res = r.json()
            if res.get("ok"):
                new_id = res["result"]["message_id"]
                with _last_cek_lock:
                    _last_cek_msgs[chat_id] = new_id
                return new_id
    except Exception:
        pass
    return None

# ================= FORCE JOIN =================
def check_force_join(user_id):
    """Cek apakah user belum join channel/grup wajib. Return list yang belum."""
    not_joined = []
    for ch in FORCE_JOIN_CHANNELS:
        try:
            username = ch["username"].lstrip("@")
            chat_id_arg = username if not username.lstrip("-").isdigit() else int(username)
            r = _tg_request("getChatMember", data={
                "chat_id": f"@{username}" if isinstance(chat_id_arg, str) else chat_id_arg,
                "user_id": user_id
            })
            if not r:
                continue  # Network error → skip, jangan block user
            res = r.json()
            if not res.get("ok"):
                # Channel tidak ditemukan / bot bukan admin → skip, jangan block user
                _log("FORCEJOIN", f"skip '{username}': {res.get('description','?')}", Fore.YELLOW)
                continue
            status = res["result"]["status"]
            if status in ("member", "administrator", "creator"):
                continue  # Sudah join ✅
            if status == "left":
                not_joined.append(ch)  # Belum join → blokir
            # "kicked"/"restricted" → skip (jangan blokir)
        except Exception as e:
            _log("FORCEJOIN", f"error cek '{ch.get('username','?')}': {e}", Fore.YELLOW)
            continue  # Error → skip, jangan block user
    return not_joined

def send_force_join_msg(chat_id, not_joined):
    """Kirim pesan wajib join dengan tombol URL (bukan link teks)."""
    rows = [[{"text": ch["label"], "url": ch["url"]}] for ch in not_joined]
    rows.append([{"text": "✅ Saya Sudah Join", "callback_data": "check_join"}])
    keyboard = {"inline_keyboard": rows}
    text = (
        "🚫 <b>WAJIB JOIN DULU!</b>\n\n"
        "<blockquote>Kamu belum join semua channel/grup yang diwajibkan.\n"
        "Join dulu, lalu klik <b>✅ Saya Sudah Join</b>.</blockquote>"
    )
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
    except Exception as e:
        print(f"FORCE JOIN MSG ERROR: {e}")

# ================= ACTIVITY LOG =================
def send_activity_log(user_id, username_display, feature_name, status="✅ Berhasil"):
    """Kirim laporan aktivitas ke channel log secara async (tidak blocking)."""
    def _send():
        now_str = datetime.now(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")
        msg = (
            f"📋 <b>LAPORAN AKTIVITAS</b>\n\n"
            f"<blockquote>"
            f"👤 User   : {username_display}\n"
            f"🆔 ID     : <code>{user_id}</code>\n"
            f"🔧 Fitur  : <b>{feature_name}</b>\n"
            f"📊 Status : {status}\n"
            f"🕐 Waktu  : {now_str} WIB"
            f"</blockquote>"
        )
        uname = BOT_USERNAME
        markup = {"inline_keyboard": [[{"text": "🤖 Buka Bot", "url": f"https://t.me/{uname}"}]]} if uname else None
        try:
            payload = {"chat_id": LOG_CHANNEL_ID, "text": msg, "parse_mode": "HTML"}
            if markup:
                payload["reply_markup"] = markup
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json=payload, timeout=10
            )
        except Exception as e:
            print(f"ACTIVITY LOG ERROR: {e}")
    threading.Thread(target=_send, daemon=True).start()
    
def cek_premium(chat_id, user_id):
    my_groups = get_user_groups(user_id)
    grup_status = f"{len(my_groups)} grup" if my_groups else "Belum addgrup (PM aktif)"
    tok = token_status_str(user_id)
    users_d = load_users()
    email_count = len(users_d.get(str(user_id), {}).get("emails", []))
    tier = get_user_tier(user_id)
    t = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])

    if user_id == OWNER_ID:
        return send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"👑 Mode     : OWNER\n"
            f"🎫 Token    : {tok}\n"
            f"📧 Email    : {email_count} akun\n"
            f"💬 Grup     : {grup_status}"
            f"</blockquote>")

    prem = premium_users.get(str(user_id))
    if prem and time.time() > prem.get("expired", 0):
        del premium_users[str(user_id)]
        save_premium(premium_users)
        prem = None
        tier = "free"
        t = TOKEN_TIERS["free"]

    user_key = get_or_create_user_key(user_id)
    email_limit = get_tier_email_limit(user_id)

    if tier == "free":
        send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"{t['emoji']} Paket    : <b>{t['label']}</b>\n"
            f"🔑 Key      : <code>{user_key}</code>\n"
            f"🎫 Token    : {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"📧 Email    : {email_count}/{email_limit} akun\n"
            f"💬 Grup     : {grup_status}\n\n"
            f"🛒 Upgrade paket → /beli"
            f"</blockquote>")
    else:
        sisa_hari = max(0, int((prem.get("expired", 0) - time.time()) // 86400))
        tok_str = "♾️ Unlimited" if t["tokens_day"] >= 99999 else f"{t['tokens_day']}/hari"
        send_msg(chat_id,
            f"📊 <b>STATUS AKUN</b>\n\n"
            f"<blockquote>"
            f"{t['emoji']} Paket    : <b>{t['label']}</b>\n"
            f"🔑 Key      : <code>{user_key}</code>\n"
            f"🎫 Token    : {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"⚡ Limit    : {tok_str}\n"
            f"📅 Sisa     : {sisa_hari} hari\n"
            f"📧 Email    : {email_count}/{email_limit} akun\n"
            f"💬 Grup     : {grup_status}"
            f"</blockquote>")                        

# ================= MENU & COMMANDS SYSTEM =================
def handle_start(user_id, chat_id):
    owner  = is_owner(user_id)
    THUMBNAIL_PATH = "./thumbnail.png"
    tok = token_status_str(user_id)

    tier = get_user_tier(user_id)
    t_info = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
    tier_badge = f"{t_info['emoji']} {t_info['label']}"

    if owner:
        caption = (
            "🤖 <b>SPIDERMAT OTP BOT</b>\n"
            "<i>SMS/OTP monitoring — Platform IVAS</i>\n\n"
            "👑 <b>OWNER PANEL</b>\n"
            "<blockquote>"
            "/addtoken — aktivasi paket user\n"
            "/deltoken — hapus paket user\n"
            "/resettoken — reset token user manual\n"
            "/listtoken — list paket aktif\n"
            "/setcookie\n"
            "/addakun\n"
            "/delakun\n"
            "/listakun\n"
            "/statsms"
            "</blockquote>\n\n"
            f"🎫 <b>Token:</b> {tok}\n\n"
            "🛠️ <b>FITUR</b>\n"
            "<blockquote>"
            "/addcookie\n"
            "/addemail email password\n"
            "/listemail\n"
            "/delemail\n"
            "/delcookie email\n"
            "/addnum\n"
            "/delnumall\n"
            "/myrange\n"
            "/ambilfile\n"
            "/cekivas\n"
            "/cekivasv2\n"
            "/toprcv [negara]\n"
            "/rangeterbaru\n"
            "/cekprem"
            "</blockquote>\n\n"
            "💬 <b>GROUP</b>\n"
            "<blockquote>"
            "/addgrup\n"
            "/delgrup\n"
            "/listgrup"
            "</blockquote>"
        )
    else:
        user_key = get_or_create_user_key(user_id)
        email_limit = get_tier_email_limit(user_id)
        caption = (
            "🤖 <b>SPIDERMAT OTP BOT</b>\n"
            "<i>SMS/OTP monitoring — Platform IVAS</i>\n\n"
            f"🔑 <b>Key:</b> <code>{user_key}</code>\n"
            f"🏷️ <b>Paket:</b> {tier_badge}\n"
            f"🎫 <b>Token:</b> {tok}  <i>(reset 00:00 WIB)</i>\n"
            f"📧 <b>Max Email:</b> {email_limit} akun\n\n"
            "🛠️ <b>FITUR</b> <i>(1 fitur = 1 token)</i>\n"
            "<blockquote>"
            "/addcookie\n"
            "/addemail email password\n"
            "/listemail\n"
            "/delemail\n"
            "/delcookie email\n"
            "/addnum\n"
            "/delnumall\n"
            "/myrange\n"
            "/ambilfile\n"
            "/cekivas\n"
            "/cekivasv2\n"
            "/toprcv [negara]\n"
            "/rangeterbaru\n"
            "/cekprem"
            "</blockquote>\n\n"
            "💬 <b>GROUP</b> <i>(gratis)</i>\n"
            "<blockquote>"
            "/addgrup\n"
            "/delgrup\n"
            "/listgrup"
            "</blockquote>\n\n"
            "🛒 <b>Upgrade Paket</b> → /beli\n"
            f"📩 <a href='https://{LINK_OWNER}'>Contact Owner</a>"
        )

    try:
        if os.path.exists(THUMBNAIL_PATH):
            with open(THUMBNAIL_PATH, "rb") as photo:
                r = requests.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                    data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                    files={"photo": photo},
                    timeout=15
                )
            if not r.json().get("ok"):
                send_msg(chat_id, caption)
        else:
            send_msg(chat_id, caption)
    except Exception as e:
        send_msg(chat_id, caption)

def code_to_flag(code):
    try: return ''.join(chr(127397 + ord(c)) for c in code.upper())
    except: return "  "
        
def add_email(text, chat_id, user_id, msg_id):
    try:
        parts = text.split()
        if len(parts) < 3:
            return send_msg(chat_id, "❌ Format:\n/addemail email@gmail.com password")
        email    = parts[1].strip().lower()
        password = parts[2].strip()
        if "@" not in email:
            return send_msg(chat_id, "❌ Email tidak valid!")

        users     = load_users()
        uid       = str(user_id)
        user_data = users.get(uid, {"emails": []})
        current_count = len(user_data.get("emails", []))

        # Limit email berdasarkan tier paket
        email_limit = get_tier_email_limit(user_id)
        if not is_owner(user_id) and current_count >= email_limit:
            tier = get_user_tier(user_id)
            t_d = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])
            if tier == "free":
                return send_msg(chat_id,
                    f"❌ <b>Limit akun FREE: {email_limit}</b>\n\n"
                    f"<blockquote>Upgrade paket untuk tambah lebih banyak akun.\n"
                    f"Ketik /beli untuk lihat paket tersedia.\n"
                    f"📩 <a href='https://{LINK_OWNER}'>Contact Owner</a></blockquote>")
            else:
                return send_msg(chat_id,
                    f"❌ Limit paket {t_d['emoji']} <b>{t_d['label']}</b>: maksimal {email_limit} akun!\n\n"
                    f"<blockquote>Upgrade ke paket lebih tinggi via /beli</blockquote>")
        if email in user_data.get("emails", []):
            return send_msg(chat_id, "❌ Akun sudah ada!")

        # Simpan email ke list & simpan password di user_accounts
        user_data.setdefault("emails", []).append(email)
        user_accs = [a for a in user_data.get("user_accounts", []) if a.get("email") != email]
        user_accs.append({"email": email, "password": password})
        user_data["user_accounts"] = user_accs
        users[uid] = user_data
        save_users(users)

        # Coba login persis seperti add_account owner
        acc = {
            "email": email, "password": password,
            "cookies": {}, "session": make_httpx_client(),
            "last_login": 0, "csrf_token": ""
        }
        send_msg(chat_id, f"⏳ Mencoba login ke <code>{email}</code>...")
        if login(acc):
            acc["last_login"] = time.time()
            send_msg(chat_id, f"✅ <b>Akun aktif &amp; login:</b>\n<code>{email}</code>")
        else:
            # Login gagal — langsung trigger cookie flow tanpa perlu /addcookie lagi
            guide_msg_id = send_msg(chat_id,
                f"⚠️ <b>Login otomatis gagal</b> untuk <code>{email}</code>\n\n"
                f"<blockquote>Password mungkin salah atau akun butuh verifikasi.\n"
                f"Pasang cookie manual di bawah agar akun langsung aktif 👇</blockquote>\n\n"
                + _cookie_guide_text("ADD COOKIE", email)
            )
            pending_addcookie[user_id] = {"email": email, "msg_id": guide_msg_id}
    except Exception as e:
        send_msg(chat_id, f"❌ Error tambah akun: {e}")

def list_email(chat_id, user_id):
    users = load_users()
    if str(user_id) not in users or not users[str(user_id)]["emails"]: return send_msg(chat_id, "  Belum ada email")
    msg = "  <b>LIST EMAIL</b>\n\n"
    for i, em in enumerate(users[str(user_id)]["emails"], 1): msg += f"{i}. {em}\n"
    send_msg(chat_id, msg)        
        
def get_user_emails(user_id):
    """Kembalikan daftar email milik user: owner -> dari accounts, premium -> dari users.json"""
    if is_owner(user_id):
        return [acc["email"] for acc in accounts]
    users = load_users()
    return users.get(str(user_id), {}).get("emails", [])

# ================= DELEMAIL FLOW =================
def command_delemail(chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun yang terdaftar.")
    buttons = [{"text": f"🗑️ {em}", "callback_data": f"de:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:de"})
    send_inline_keyboard(chat_id,
        "🗑️ <b>HAPUS EMAIL</b>\n\n"
        "<blockquote>⚠️ Pilih email yang ingin dihapus.\n"
        "Akun beserta cookie &amp; session akan langsung dihentikan.\n\n"
        "Aksi ini <b>tidak bisa dibatalkan</b>!</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_delemail_select_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⚠️ Konfirmasi dulu!")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>HAPUS EMAIL</b>\n\n❌ Email tidak ditemukan di akun kamu.")
        return
    buttons = [
        {"text": "✅ Ya, Hapus",  "callback_data": f"dec:{email}"},
        {"text": "❌ Batalkan",   "callback_data": "cancel:de"},
    ]
    delete_and_send_inline(chat_id, msg_id,
        f"🗑️ <b>HAPUS EMAIL — KONFIRMASI</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"❗ Cookie, session, dan semua data akun ini akan dihapus permanen.\n"
        f"Apakah kamu yakin?</blockquote>",
        buttons)

def handle_delemail_confirm_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Menghapus akun...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>HAPUS EMAIL</b>\n\n❌ Email tidak ditemukan di akun kamu.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"🗑️ <b>HAPUS EMAIL</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Menghentikan session &amp; menghapus data..."
        f"</blockquote>")

    try:
        uid = str(user_id)

        # 1. Hapus dari users.json (emails + user_accounts)
        users_d = load_users()
        if uid in users_d:
            users_d[uid]["emails"] = [e for e in users_d[uid].get("emails", []) if e != email]
            users_d[uid]["user_accounts"] = [a for a in users_d[uid].get("user_accounts", []) if a.get("email") != email]
            save_users(users_d)

        # 2. Hapus dari premium-cookie.json — dikunci supaya tidak nabrak akun lain
        with _cookie_file_lock:
            prem_cookies = load_premium_cookies()
            if email in prem_cookies:
                del prem_cookies[email]
                save_premium_cookies(prem_cookies)

        # 3. Hapus dari _premium_acc_cache (hentikan session)
        _premium_acc_cache.pop(email, None)

        # 4. Bersihkan cache ranges & session trackers
        _ranges_cache.pop(email, None)
        _recv_csrf_cache.pop(email, None)
        _last_cookie_refresh.pop(email, None)
        _last_cookie_notif.pop(email, None)
        _keepalive_warn_count.pop(email, None)
        _session_fail_time.pop(email, None)
        _session_notified.pop(email, None)
        _session_retry_time.pop(email, None)
        _session_recovered.pop(email, None)

        # 5. Bersihkan pending state user ini
        pending_addnum.pop(user_id, None)
        pending_addcookie.pop(user_id, None)
        pending_setcookie.pop(user_id, None)

        delete_and_send(chat_id, proc_id,
            f"🗑️ <b>HAPUS EMAIL BERHASIL</b>\n\n"
            f"<blockquote>"
            f"📧 Email: <code>{email}</code>\n"
            f"✅ Akun berhasil dihapus\n"
            f"✅ Cookie &amp; session dihentikan\n"
            f"✅ Semua data akun dibersihkan"
            f"</blockquote>")

    except Exception as e:
        _log("DELEMAIL", f"error: {e}", Fore.RED)
        try:
            delete_and_send(chat_id, proc_id,
                f"🗑️ <b>HAPUS EMAIL ERROR</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"❌ Error: <code>{str(e)[:200]}</code>"
                f"</blockquote>")
        except Exception:
            pass

def delete_and_send_inline(chat_id, msg_id, text, buttons):
    """Hapus pesan lama lalu kirim pesan baru dengan inline keyboard."""
    delete_msg(chat_id, msg_id)
    return send_inline_keyboard(chat_id, text, buttons)

# ================= ADDNUM FLOW =================
def command_addnum(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.\nTambah dulu dengan /addemail atau /addakun")
    buttons = [{"text": f"📧 {em}", "callback_data": f"an:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:an"})
    send_inline_keyboard(chat_id,
        "➕ <b>ADD NUMBER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Ketik target nomor atau negara\n"
        "   Contoh: <code>SAUDI ARABIA 15022</code>\n"
        "   Contoh: <code>INDONESIA 500</code>\n"
        "3. Bot akan proses penambahan nomor ke akun\n\n"
        "⚠️ Pastikan cookie sudah aktif sebelum add number</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_addnum_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "✅ Email dipilih!")
    emails = get_user_emails(user_id)
    if email not in emails:
        answer_callback_query(cb_id, "❌ Email tidak ditemukan")
        return
    new_msg_id = delete_and_send_with_cancel(chat_id, msg_id,
        f"➕ <b>ADD NUMBER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"<blockquote>✏️ Ketik range yang ingin ditambahkan:\n\n"
        f"<b>1 Range:</b>\n"
        f"<code>BENIN 851</code>\n\n"
        f"<b>Multi Range (pisah enter/koma):</b>\n"
        f"<code>BENIN 851\nMOZAMBIQUE 4234\nSAUDI ARABIA 15022</code></blockquote>",
        "an"
    )
    pending_addnum[user_id] = {"email": email, "msg_id": new_msg_id}

def _fetch_fresh_csrf_from_page(session, url, current_csrf=""):
    """Ambil fresh CSRF token dari halaman HTML. Fallback ke current_csrf jika gagal."""
    try:
        r = session.get(url, timeout=15)
        if r.status_code != 200:
            return current_csrf
        soup = BeautifulSoup(r.text, "html.parser")
        meta = soup.find("meta", {"name": "csrf-token"})
        if meta and meta.get("content"):
            return meta["content"]
        inp = soup.find("input", {"name": "_token"})
        if inp and inp.get("value"):
            return inp["value"]
        m = re.search(r"['\"]_token['\"]\s*[,:]?\s*['\"]([A-Za-z0-9_\-+/=]{20,})['\"]", r.text)
        if m:
            return m.group(1)
    except Exception:
        pass
    return current_csrf


def _do_addnum_range(acc, session, csrf, target_text, progress_cb=None):
    """Fetch test numbers dan add untuk 1 range. Return dict hasil."""
    test_url = f"{BASE}/portal/numbers/test"
    hdrs = {
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          test_url,
    }

    # Refresh CSRF dari halaman test numbers (bukan dari cache portal umum)
    fresh_csrf = _fetch_fresh_csrf_from_page(session, test_url, csrf)
    if fresh_csrf and fresh_csrf != csrf:
        acc["csrf_token"] = fresh_csrf
        csrf = fresh_csrf

    def _fetch_test(length):
        p = {
            "draw":                   "1",
            "columns[0][data]":       "range",
            "columns[0][name]":       "terminations.range",
            "columns[1][data]":       "test_number",
            "columns[1][name]":       "terminations.test_number",
            "columns[2][data]":       "id",
            "columns[2][name]":       "id",
            "columns[3][data]":       "limit_did_a2p",
            "columns[3][name]":       "limit_did_a2p",
            "columns[4][data]":       "limit_cli_did_a2p",
            "columns[4][name]":       "limit_cli_did_a2p",
            "order[0][column]":       "0",
            "order[0][dir]":          "asc",
            "start":                  "0",
            "length":                 str(length),
            "search[value]":          target_text,
            "search[regex]":          "false",
        }
        r = session.get(test_url, params=p, headers=hdrs, timeout=25)
        if r.status_code != 200:
            raise Exception(f"HTTP {r.status_code}")
        return r.json()

    try:
        # Probe dulu untuk tahu total nomor yang tersedia (recordsFiltered)
        probe       = _fetch_test(1)
        total_avail = int(probe.get("recordsFiltered", probe.get("recordsTotal", 0)))
        # Fetch semua nomor di range — minimal 100, max 2000
        fetch_count = max(100, min(total_avail if total_avail > 0 else 2000, 2000))
        data = _fetch_test(fetch_count)
        rows = data.get("data", [])
    except Exception as e:
        return {"success": 0, "fail": 0, "skipped": False, "total": 0,
                "skip_msg": "", "not_found": False, "error": str(e)}

    fallback_fields = ["range", "test_number", "id", "limit_did_a2p", "limit_cli_did_a2p",
                       "term", "A2P", "created_at", "action"]
    rn_lower = re.sub(r"\s+", " ", target_text.lower().strip())
    items = []
    for row in rows:
        if isinstance(row, list):
            row = dict(zip(fallback_fields, row))
        rng_raw = re.sub(r"<[^>]+>", "", str(row.get("range", ""))).strip()
        rng_norm = re.sub(r"\s+", " ", rng_raw.lower())
        # Cocokkan jika nama range mengandung query atau sebaliknya (toleran whitespace)
        if rng_norm != rn_lower and rn_lower not in rng_norm:
            continue
        tid = str(row.get("id", "") or row.get("DT_RowId", "")).strip()
        if tid and not tid.isdigit():
            m2 = re.search(r"(\d+)", tid)
            tid = m2.group(1) if m2 else ""
        if tid:
            items.append({"tid": tid, "rng": rng_raw})

    if not items:
        return {"success": 0, "fail": 0, "skipped": False, "total": 0,
                "skip_msg": "", "not_found": True, "error": None}

    add_url  = f"{BASE}/portal/numbers/termination/number/add"
    add_hdrs = {
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          test_url,
        "Origin":           BASE,
        "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
    }

    success_count = 0
    fail_count    = 0
    skipped       = False
    skip_msg      = ""
    total_items   = len(items)
    _last_cb_at   = [0]

    for idx, item in enumerate(items):
        tid = item["tid"]
        try:
            resp = session.post(add_url, data={"id": tid, "_token": csrf},
                                headers=add_hdrs, timeout=15)
            try:
                jr      = resp.json()
                message = str(jr.get("message", jr.get("msg", jr.get("error", str(jr)))))
                st      = jr.get("status", jr.get("success", ""))
                ok      = str(st).lower() in ("success", "ok", "true", "1") or st is True or st == 1
                if not ok:
                    ok = any(k in message.lower() for k in
                             ("berhasil", "success", "added", "good job", "successfully", "done add number", "done"))
                if not ok and any(k in message.lower() for k in
                                  ("too many", "maximum", "limit", "penuh")):
                    skipped  = True
                    skip_msg = message
                    break
            except Exception:
                raw = resp.text.lower()
                ok  = any(k in raw for k in ("berhasil", "success", "added", "good job"))
                if any(k in raw for k in ("too many", "maximum", "limit", "penuh")):
                    skipped  = True
                    skip_msg = f"HTTP {resp.status_code}: limit tercapai"
                    break
            if ok:
                success_count += 1
            else:
                fail_count += 1
            time.sleep(0.1)  # jeda minimal — cukup hindari rate-limit server
        except Exception:
            fail_count += 1

        # Kirim progress callback setiap 10 nomor atau di nomor terakhir
        done = idx + 1
        if progress_cb and (done - _last_cb_at[0] >= 10 or done == total_items):
            try:
                progress_cb(done, total_items, success_count, fail_count)
            except Exception:
                pass
            _last_cb_at[0] = done

    return {"success": success_count, "fail": fail_count, "skipped": skipped,
            "total": total_items, "skip_msg": skip_msg, "not_found": False, "error": None}


def process_addnum_target(chat_id, user_id, target_text):
    state = pending_addnum.pop(user_id, None)
    if not state:
        return False
    email  = state["email"]
    msg_id = state["msg_id"]

    # Parse multi-range: pisah per baris atau koma
    raw_ranges = re.split(r"[\n,]+", target_text)
    ranges = [r.strip() for r in raw_ranges if r.strip()]
    if not ranges:
        return False

    preview = ", ".join(f"<code>{r}</code>" for r in ranges[:3])
    if len(ranges) > 3:
        preview += f" +{len(ranges)-3} lainnya"

    proc_id = delete_and_send(chat_id, msg_id,
        f"➕ <b>ADD NUMBER</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n"
        f"🎯 {'Range' if len(ranges) == 1 else f'{len(ranges)} Range'}: {preview}\n\n"
        f"⏳ Mencari nomor di range...</blockquote>")

    def _run():
        try:
            multi = len(ranges) > 1
            acc = get_acc_by_email(email)

            if not acc:
                delete_and_send(chat_id, proc_id,
                    f"➕ <b>ADD NUMBER</b>\n\n"
                    f"❌ Akun <code>{email}</code> tidak ditemukan.\n"
                    f"<blockquote>Pastikan sudah /addemail dan /addcookies untuk akun ini.</blockquote>")
                return

            if not ensure_login(acc):
                delete_and_send(chat_id, proc_id,
                    f"➕ <b>ADD NUMBER</b>\n\n"
                    f"❌ Session akun <code>{email}</code> tidak aktif.\n"
                    f"<blockquote>Cookie expired atau tidak valid.\nGunakan /addcookies untuk memperbarui cookie.</blockquote>")
                return

            session = acc["session"]
            csrf    = acc.get("csrf_token", "")
            results = []

            for i, rng_target in enumerate(ranges):
                done_lines = ""
                for prev in results:
                    if prev.get("error"):
                        st = "❌ Error"
                    elif prev.get("not_found"):
                        st = "❌ Tdk ditemukan"
                    elif prev["skipped"] and prev["success"] == 0:
                        st = "⚠️ Penuh"
                    elif prev["success"] > 0:
                        st = f"✅ {prev['success']} nomor"
                    else:
                        st = "❌ Gagal"
                    done_lines += f"• <code>{prev['range']}</code>: {st}\n"

                if multi:
                    edit_msg(chat_id, proc_id,
                        f"➕ <b>ADD NUMBER</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"⏳ [{i+1}/{len(ranges)}] Proses: <code>{rng_target}</code>...\n"
                        + (f"\n{done_lines.strip()}" if done_lines else "")
                        + f"</blockquote>")
                else:
                    edit_msg(chat_id, proc_id,
                        f"➕ <b>ADD NUMBER</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"🎯 Range: <code>{rng_target}</code>\n\n"
                        f"⏳ Memulai add nomor...</blockquote>")

                def make_progress_cb(rng_name, p_id, is_multi, i_idx, tot_ranges, d_lines):
                    def _cb(done, total, ok, fail):
                        pct = int(done / total * 100) if total else 0
                        bar_filled = int(pct / 10)
                        bar = "▓" * bar_filled + "░" * (10 - bar_filled)
                        if is_multi:
                            edit_msg(chat_id, p_id,
                                f"➕ <b>ADD NUMBER</b>\n\n"
                                f"<blockquote>"
                                f"📧 Email: <code>{email}</code>\n"
                                f"⏳ [{i_idx+1}/{tot_ranges}] <code>{rng_name}</code>\n"
                                f"[{bar}] {pct}%\n"
                                f"✅ {ok} berhasil | ❌ {fail} gagal | 📊 {done}/{total}\n"
                                + (f"\n{d_lines.strip()}" if d_lines else "")
                                + f"</blockquote>")
                        else:
                            edit_msg(chat_id, p_id,
                                f"➕ <b>ADD NUMBER</b>\n\n"
                                f"<blockquote>"
                                f"📧 Email: <code>{email}</code>\n"
                                f"🎯 Range: <code>{rng_name}</code>\n\n"
                                f"[{bar}] {pct}%\n"
                                f"✅ {ok} berhasil | ❌ {fail} gagal | 📊 {done}/{total}</blockquote>")
                    return _cb

                cb = make_progress_cb(rng_target, proc_id, multi, i, len(ranges), done_lines)
                r = _do_addnum_range(acc, session, csrf, rng_target, progress_cb=cb)
                results.append({"range": rng_target, **r})
                if i < len(ranges) - 1:
                    time.sleep(0.5)

            if not multi:
                r = results[0]
                if r.get("error"):
                    result_text = (
                        f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email : <code>{email}</code>\n"
                        f"🎯 Range : <code>{ranges[0]}</code>\n"
                        f"❌ Error : <code>{r['error'][:150]}</code>"
                        f"</blockquote>"
                    )
                elif r.get("not_found"):
                    result_text = (
                        f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email : <code>{email}</code>\n"
                        f"🎯 Range : <code>{ranges[0]}</code>\n"
                        f"❌ Range tidak ditemukan di Test Numbers.\n\n"
                        f"Pastikan nama range benar (case-sensitive).\n"
                        f"Cek /toprcv untuk nama range yang tersedia."
                        f"</blockquote>"
                    )
                elif r["skipped"] and r["success"] == 0:
                    result_text = (
                        f"➕ <b>ADD NUMBER GAGAL</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email : <code>{email}</code>\n"
                        f"🎯 Range : <code>{ranges[0]}</code>\n"
                        f"⚠️ Slot nomor di akun sudah penuh.\n\n"
                        f"Hubungi admin IVAS untuk tambah kuota."
                        f"</blockquote>"
                    )
                elif r["skipped"]:
                    result_text = (
                        f"➕ <b>ADD NUMBER SELESAI (PARSIAL)</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email  : <code>{email}</code>\n"
                        f"🎯 Range  : <code>{ranges[0]}</code>\n"
                        f"✅ Berhasil: <b>{r['success']}</b> nomor\n"
                        f"❌ Gagal  : <b>{r['fail']}</b> nomor\n"
                        f"⚠️ Berhenti: Slot akun sudah penuh"
                        f"</blockquote>"
                    )
                else:
                    result_text = (
                        f"➕ <b>ADD NUMBER {'BERHASIL' if r['success'] > 0 else 'GAGAL'}</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email  : <code>{email}</code>\n"
                        f"🎯 Range  : <code>{ranges[0]}</code>\n"
                        f"✅ Berhasil: <b>{r['success']}</b> / {r['total']} nomor\n"
                        f"❌ Gagal  : <b>{r['fail']}</b> nomor"
                        f"</blockquote>"
                    )
            else:
                total_ok   = sum(1 for r in results if r.get("success", 0) > 0)
                total_fail = sum(1 for r in results if r.get("success", 0) == 0 and not r.get("skipped") and not r.get("error") and not r.get("not_found"))
                lines = ""
                for r in results:
                    if r.get("error"):
                        status = "❌ Error fetch"
                    elif r.get("not_found"):
                        status = "❌ Tidak ditemukan"
                    elif r["skipped"] and r["success"] == 0:
                        status = "⚠️ Penuh"
                    elif r["skipped"]:
                        status = "✅ (lalu penuh)"
                    elif r["success"] > 0:
                        status = "✅"
                    else:
                        status = "❌ Gagal"
                    lines += f"• <code>{r['range']}</code>: {status}\n"

                result_text = (
                    f"➕ <b>ADD NUMBER SELESAI</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"🔢 Total: ✅ <b>{total_ok}</b> berhasil | ❌ <b>{total_fail}</b> gagal\n\n"
                    f"{lines.strip()}"
                    f"</blockquote>"
                )

            if multi:
                edit_msg(chat_id, proc_id, result_text)
            else:
                delete_and_send(chat_id, proc_id, result_text)

        except Exception as e:
            _log("ADDNUM", f"error tak terduga: {e}", Fore.RED)
            try:
                delete_and_send(chat_id, proc_id,
                    f"➕ <b>ADD NUMBER ERROR</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Terjadi error: <code>{str(e)[:200]}</code>"
                    f"</blockquote>")
            except Exception:
                pass

    threading.Thread(target=_run, daemon=True).start()
    return True


# ================= DELNUMALL FLOW =================
def command_delnumall(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"da:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:da"})
    send_inline_keyboard(chat_id,
        "🗑️ <b>DELETE ALL NUMBER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan otomatis return semua nomor yang aktif\n"
        "3. Tunggu konfirmasi selesai\n\n"
        "⚠️ Semua nomor akan dikembalikan ke pool IVAS!</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_delnumall_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "🗑️ <b>DELETE ALL NUMBER</b>\n\n❌ Email tidak ditemukan.")
        return
    proc_id = delete_and_send(chat_id, msg_id,
        f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang menghapus semua nomor..."
        f"</blockquote>")

    def _run_delnumall():
        try:
            acc_target = get_acc_by_email(email)
            if not acc_target:
                delete_and_send(chat_id, proc_id,
                    f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Akun tidak ditemukan. Pastikan sudah /addemail dan /addcookies."
                    f"</blockquote>")
                return

            if not ensure_login(acc_target):
                delete_and_send(chat_id, proc_id,
                    f"🗑️ <b>DELETE ALL NUMBER</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Gagal login/verifikasi session. Perbarui cookie."
                    f"</blockquote>")
                return

            ok, res = return_all_base(acc_target)
            if ok:
                delete_and_send(chat_id, proc_id,
                    f"🗑️ <b>DELETE ALL NUMBER BERHASIL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"✅ Semua nomor berhasil dikembalikan ke pool!"
                    f"</blockquote>")
            else:
                delete_and_send(chat_id, proc_id,
                    f"🗑️ <b>DELETE ALL NUMBER GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ {str(res)[:150]}"
                    f"</blockquote>")
        except Exception as e:
            _log("DELNUMALL", f"error: {e}", Fore.RED)
            try:
                delete_and_send(chat_id, proc_id,
                    f"🗑️ <b>DELETE ALL NUMBER ERROR</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Error: <code>{str(e)[:200]}</code>"
                    f"</blockquote>")
            except Exception:
                pass

    threading.Thread(target=_run_delnumall, daemon=True).start()

# ================= MYRANGE FLOW =================
def command_myrange(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"mr:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:mr"})
    send_inline_keyboard(chat_id,
        "📊 <b>MY RANGE</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan menampilkan semua range di My Numbers\n"
        "3. Termasuk jumlah nomor per range</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_myrange_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "📊 <b>MY RANGE</b>\n\n❌ Email tidak ditemukan.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"📊 <b>MY RANGE</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang mengambil data range..."
        f"</blockquote>")

    def _run_myrange():
        try:
            acc_target = get_acc_by_email(email)
            if not acc_target:
                delete_and_send(chat_id, proc_id,
                    f"📊 <b>MY RANGE</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Akun tidak ditemukan. Pastikan sudah /addemail dan /addcookies."
                    f"</blockquote>")
                return

            if not ensure_login(acc_target):
                delete_and_send(chat_id, proc_id,
                    f"📊 <b>MY RANGE</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Gagal login/verifikasi session. Perbarui cookie."
                    f"</blockquote>")
                return

            data = _fetch_myrange_data(acc_target)
            rows = data.get("data", [])
            total = data.get("recordsTotal", 0)

            if rows and isinstance(rows[0], list):
                rows = [dict(zip(col_data, r)) for r in rows]

            from collections import Counter
            range_count = Counter()
            for row in rows:
                if isinstance(row, dict):
                    rng = re.sub(r"<[^>]+>", "", str(row.get("range", ""))).strip()
                    if rng:
                        range_count[rng] += 1

            if not range_count:
                delete_and_send(chat_id, proc_id,
                    f"📊 <b>MY RANGE</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"ℹ️ Tidak ada nomor di My Numbers."
                    f"</blockquote>")
                return

            lines = ""
            for i, (rng, cnt) in enumerate(sorted(range_count.items()), 1):
                lines += f"{i}. <b>{rng}</b> — {cnt} nomor\n"

            result_text = (
                f"📊 <b>MY RANGE</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔢 Total: <b>{total}</b> nomor | <b>{len(range_count)}</b> range\n\n"
                f"{lines.strip()}"
                f"</blockquote>"
            )
            delete_and_send(chat_id, proc_id, result_text)

        except Exception as ex:
            _log("MYRANGE", f"error: {ex}", Fore.RED)
            try:
                delete_and_send(chat_id, proc_id,
                    f"📊 <b>MY RANGE</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Error: {str(ex)[:150]}"
                    f"</blockquote>")
            except Exception:
                pass

    threading.Thread(target=_run_myrange, daemon=True).start()

# ================= AMBILFILE FLOW =================
def command_testapi(chat_id, user_id, text):
    """
    /testapi [email]
    Test semua endpoint IVAS dari session bot yang sudah authenticated.
    Hasilnya dikirim ke Telegram. OWNER only.
    """
    import traceback

    parts = text.strip().split(None, 1)
    target_email = parts[1].strip() if len(parts) > 1 else None

    if target_email:
        acc = get_acc_by_email(target_email)
        if not acc:
            send_msg(chat_id, f"❌ Email <code>{target_email}</code> tidak ditemukan di akun bot.")
            return
    else:
        accs = get_user_emails(user_id)
        if not accs:
            send_msg(chat_id, "❌ Belum ada akun. Tambah dulu dengan /addemail.")
            return
        acc = get_acc_by_email(accs[0])
        if not acc:
            send_msg(chat_id, "❌ Akun tidak bisa diload.")
            return

    email = acc["email"]
    status_id = send_msg(chat_id,
        f"🔬 <b>TEST API IVAS</b>\n\n"
        f"<blockquote>📧 Akun: <code>{email}</code>\n"
        f"⏳ Memulai test semua endpoint...</blockquote>")

    lines = [f"🔬 <b>TEST API IVAS</b>", f"<blockquote>📧 <code>{email}</code>\n"]

    def _add(label, ok, detail=""):
        icon = "✅" if ok else "❌"
        row  = f"{icon} <b>{label}</b>"
        if detail:
            row += f"\n   <code>{detail[:200]}</code>"
        lines.append(row)

    try:
        # ── Pastikan session aktif ──
        if not ensure_login(acc):
            _add("Login/Session", False, "Cookie expired / gagal login")
            lines.append("</blockquote>")
            delete_and_send(chat_id, status_id, "\n".join(lines))
            return
        _add("Login/Session", True, "Session aktif")

        session = acc["session"]

        # ── TEST 1: GET /portal/numbers (login check + CSRF) ──
        try:
            r = session.get(f"{BASE}/portal/numbers", timeout=20)
            ct = r.headers.get("content-type","")
            login_ok = r.status_code == 200 and "/login" not in str(r.url)
            csrf_found = ""
            if login_ok:
                import re as _re
                from bs4 import BeautifulSoup as _BS
                soup = _BS(r.text, "html.parser")
                m = soup.find("meta", {"name": "csrf-token"})
                if m and m.get("content"):
                    csrf_found = m["content"][:20] + "..."
                else:
                    inp = soup.find("input", {"name": "_token"})
                    if inp and inp.get("value"):
                        csrf_found = inp["value"][:20] + "..."
                    else:
                        csrf_found = "NOT FOUND"
            detail = f"HTTP {r.status_code} | CSRF: {csrf_found}" if login_ok else f"HTTP {r.status_code} → redirect ke login"
            _add("GET /portal/numbers", login_ok, detail)
        except Exception as e:
            _add("GET /portal/numbers", False, str(e)[:100])

        # ── TEST 2: GET /portal/numbers/test (DataTable) ──
        try:
            p = {
                "draw": "1", "start": "0", "length": "3",
                "columns[0][data]": "range", "columns[0][name]": "terminations.range",
                "columns[1][data]": "test_number", "columns[1][name]": "terminations.test_number",
                "columns[2][data]": "id", "columns[2][name]": "id",
                "order[0][column]": "0", "order[0][dir]": "asc",
                "search[value]": "", "search[regex]": "false",
            }
            hdrs = {"Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "Referer": f"{BASE}/portal/numbers/test"}
            r2 = session.get(f"{BASE}/portal/numbers/test", params=p, headers=hdrs, timeout=20)
            if r2.status_code == 200:
                try:
                    j2  = r2.json()
                    tot = j2.get("recordsTotal", "?")
                    fil = j2.get("recordsFiltered", "?")
                    rows = j2.get("data", [])
                    sample = ""
                    if rows:
                        row0 = rows[0]
                        if isinstance(row0, dict):
                            sample = f" | sample range: {str(row0.get('range',''))[:30]}"
                    _add("GET /portal/numbers/test", True,
                         f"HTTP 200 JSON ✓ | total={tot} filtered={fil}{sample}")
                except Exception as je:
                    _add("GET /portal/numbers/test", False,
                         f"HTTP {r2.status_code} tapi gagal parse JSON: {je} | raw: {r2.text[:80]}")
            else:
                _add("GET /portal/numbers/test", False, f"HTTP {r2.status_code} | {r2.text[:80]}")
        except Exception as e:
            _add("GET /portal/numbers/test", False, str(e)[:100])

        # ── TEST 3: GET /portal/numbers/export ──
        try:
            hdrs3 = {"Accept": "text/html,application/xhtml+xml,*/*",
                     "Referer": f"{BASE}/portal/numbers"}
            r3 = session.get(EXPORT_URL, headers=hdrs3, timeout=30)
            ct3 = r3.headers.get("content-type", "")
            cd3 = r3.headers.get("content-disposition", "")
            size3 = len(r3.content)
            redirect_to_login = "/login" in str(r3.url)
            is_file = any(k in ct3 for k in ("spreadsheet","octet-stream","excel","csv")) or "attachment" in cd3
            detail3 = f"HTTP {r3.status_code} | CT: {ct3[:60]} | CD: {cd3[:40]} | size: {size3}B"
            if redirect_to_login:
                detail3 += " | → REDIRECT LOGIN"
            if is_file:
                detail3 += " | ✓ FILE RESPONSE"
            else:
                detail3 += f" | body: {r3.text[:60]}"
            _add("GET /portal/numbers/export", r3.status_code == 200 and not redirect_to_login, detail3)
        except Exception as e:
            _add("GET /portal/numbers/export", False, str(e)[:100])

        # ── TEST 4: CSRF fresh dari /portal/numbers/test ──
        try:
            fresh = _fetch_fresh_csrf_from_page(session, f"{BASE}/portal/numbers/test",
                                                acc.get("csrf_token",""))
            if fresh and len(fresh) > 10:
                _add("CSRF refresh dari test page", True, f"{fresh[:25]}... (len={len(fresh)})")
            else:
                _add("CSRF refresh dari test page", False, f"Dapat: {repr(fresh)}")
        except Exception as e:
            _add("CSRF refresh dari test page", False, str(e)[:100])

        # ── TEST 5: POST /portal/numbers/termination/number/add (dry-run, id=0) ──
        try:
            csrf_use = acc.get("csrf_token","")
            add_hdrs = {"Accept": "application/json, text/javascript, */*; q=0.01",
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": f"{BASE}/portal/numbers/test",
                        "Origin": BASE,
                        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
            r5 = session.post(f"{BASE}/portal/numbers/termination/number/add",
                              data={"id": "0", "_token": csrf_use},
                              headers=add_hdrs, timeout=15)
            try:
                j5 = r5.json()
                detail5 = f"HTTP {r5.status_code} JSON: {str(j5)[:150]}"
            except Exception:
                detail5 = f"HTTP {r5.status_code} text: {r5.text[:150]}"
            # id=0 harusnya error tapi validasi CSRF harus pass
            csrf_ok = r5.status_code not in (401, 419, 403)
            _add("POST /numbers/termination/number/add (dry id=0)", csrf_ok, detail5)
        except Exception as e:
            _add("POST /numbers/termination/number/add", False, str(e)[:100])

        # ── TEST 6: POST /portal/numbers/return/allnumber/bluck (dry-run, cek CSRF only) ──
        try:
            csrf_use2 = acc.get("csrf_token","")
            hdrs6 = {"X-Requested-With": "XMLHttpRequest",
                     "Referer": f"{BASE}/portal/numbers",
                     "Origin": BASE,
                     "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
            # Pakai _method override untuk cek tanpa eksekusi — kirim OPTIONS dulu
            r6 = session.options(RETURN_ALL_URL, headers=hdrs6, timeout=10)
            detail6 = f"OPTIONS HTTP {r6.status_code} | Allow: {r6.headers.get('allow',r6.headers.get('Allow','?'))}"
            _add("OPTIONS /numbers/return/allnumber/bluck", True, detail6)
        except Exception as e:
            _add("OPTIONS /numbers/return/allnumber/bluck", False, str(e)[:100])

    except Exception as e:
        lines.append(f"❌ Fatal error: <code>{traceback.format_exc()[-300:]}</code>")

    lines.append("</blockquote>")
    delete_and_send(chat_id, status_id, "\n".join(lines))


def command_ambilfile(text, chat_id, user_id):
    emails = get_user_emails(user_id)
    if not emails:
        return send_msg(chat_id, "❌ Belum ada email/akun.")
    buttons = [{"text": f"📧 {em}", "callback_data": f"af:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:af"})
    send_inline_keyboard(chat_id,
        "📁 <b>AMBIL FILE</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Bot akan mengambil data nomor dari IVAS\n"
        "3. File Excel (.xlsx) dikirim otomatis ke chat ini\n\n"
        "💡 File berisi semua nomor aktif beserta range/negara</blockquote>\n\n"
        "👇 Pilih email:",
        buttons)

def handle_ambilfile_email_cb(chat_id, user_id, email, cb_id, msg_id):
    answer_callback_query(cb_id, "⏳ Memproses...")
    emails = get_user_emails(user_id)
    if email not in emails:
        delete_and_send(chat_id, msg_id,
            "📁 <b>AMBIL FILE</b>\n\n❌ Email tidak ditemukan.")
        return

    proc_id = delete_and_send(chat_id, msg_id,
        f"📁 <b>AMBIL FILE</b>\n\n"
        f"<blockquote>"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Sedang mengambil &amp; menyusun file export..."
        f"</blockquote>")

    def _run_ambilfile():
        try:
            acc_target = get_acc_by_email(email)
            if not acc_target:
                delete_and_send(chat_id, proc_id,
                    f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Akun tidak ditemukan. Pastikan sudah /addemail dan /addcookies."
                    f"</blockquote>")
                return

            if not ensure_login(acc_target):
                delete_and_send(chat_id, proc_id,
                    f"📁 <b>AMBIL FILE GAGAL</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Cookie expired atau tidak valid. Perbarui cookie dulu."
                    f"</blockquote>")
                return

            export_numbers_ivas(chat_id, acc_target, status_msg_id=proc_id)
        except Exception as e:
            _log("AMBILFILE", f"error: {e}", Fore.RED)
            try:
                delete_and_send(chat_id, proc_id,
                    f"📁 <b>AMBIL FILE ERROR</b>\n\n"
                    f"<blockquote>"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Error: <code>{str(e)[:200]}</code>"
                    f"</blockquote>")
            except Exception:
                pass

    threading.Thread(target=_run_ambilfile, daemon=True).start()

def delete_msg(chat_id, message_id):
    try: requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage", data={"chat_id": chat_id, "message_id": message_id}, timeout=10)
    except: pass    

def detect_country_and_flag(full_num, fallback_country="UNKNOWN"):
    try:
        parsed = phonenumbers.parse("+" + full_num, None)
        region = phonenumbers.region_code_for_number(parsed)
        if region:
            flag = code_to_flag(region)
            country_name = geocoder.description_for_number(parsed, "en")
            if not country_name: country_name = fallback_country
            return country_name.upper(), flag
    except Exception as e: print("FLAG ERROR:", e)
    return fallback_country, "  "
    
def parse_cookie_input(raw_text):
    try:
        data = json.loads(raw_text)
        if isinstance(data, list):
            cookie_dict = {}
            for item in data:
                if isinstance(item, dict) and "name" in item and "value" in item:
                    cookie_dict[item["name"]] = item["value"]
            return cookie_dict if cookie_dict else None
        elif isinstance(data, dict):
            return data
        return None
    except:
        return None

def get_recv_csrf(acc, _retry=0) -> str:
    """
    Ambil CSRF token dari halaman /portal/sms/received.
    iVAS pakai per-page rotating CSRF — semua POST ke getsms API
    WAJIB pakai token dari halaman ini, bukan dari /portal umum.
    Di-cache 15 menit per akun.
    """
    email = acc.get("email", "")
    now   = time.time()
    cached = _recv_csrf_cache.get(email)
    if cached and (now - cached["ts"]) < RECV_CSRF_TTL:
        return cached["csrf"]
    try:
        worker_before = _IVAS_ORIGIN
        r = acc["session"].get(RECV_URL, timeout=15)
        if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
            mark_worker_limited(worker_before)
            return get_recv_csrf(acc, _retry=_retry + 1)
        if "/login" in str(r.url):
            return acc.get("recv_csrf") or acc.get("csrf_token", "")
        soup = BeautifulSoup(r.text, "html.parser")
        csrf = ""
        meta = soup.find("meta", {"name": "csrf-token"})
        if meta:
            csrf = meta.get("content", "")
        if not csrf:
            inp = soup.find("input", {"name": "_token"})
            if inp:
                csrf = inp.get("value", "")
        if not csrf:
            m = re.search(r"['\"]_token['\"]\s*[,:]?\s*['\"]([A-Za-z0-9_\-+/=]{20,})['\"]", r.text)
            if m:
                csrf = m.group(1)
        if csrf:
            acc["recv_csrf"] = csrf
            _recv_csrf_cache[email] = {"csrf": csrf, "ts": now}
            return csrf
    except Exception as e:
        _log("WARN", f"get_csrf [{email}]: {e}", Fore.YELLOW)
    return acc.get("recv_csrf") or acc.get("csrf_token", "")


def verify_cookie_session(acc, _retry=0):
    """
    Verifikasi session cookie.
    1. GET /portal — cek tidak redirect ke /login & ambil csrf_token umum
    2. GET /portal/sms/received — ambil recv_csrf khusus untuk SMS API
    """
    try:
        session = acc["session"]
        worker_before = _IVAS_ORIGIN
        r = session.get(f"{BASE}/portal", timeout=15)
        if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
            mark_worker_limited(worker_before)
            return verify_cookie_session(acc, _retry=_retry + 1)
        if "/login" in str(r.url):
            return False
        soup = BeautifulSoup(r.text, "html.parser")
        token_input = soup.find("input", {"name": "_token"})
        if token_input:
            acc["csrf_token"] = token_input["value"]
        else:
            token_meta = soup.find("meta", {"name": "csrf-token"})
            if token_meta:
                acc["csrf_token"] = token_meta.get("content", "")
        # Ambil recv_csrf — WAJIB untuk POST ke SMS API
        email = acc.get("email", "")
        _recv_csrf_cache.pop(email, None)  # Paksa refresh recv_csrf setelah verify
        get_recv_csrf(acc)
        return True
    except Exception as e:
        print(f"Cookie verify error: {e}")
        return False

def send_inline_keyboard(chat_id, text, buttons):
    keyboard = {"inline_keyboard": [[{"text": b["text"], "callback_data": b["callback_data"]}] for b in buttons]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def send_inline_keyboard_grid(chat_id, text, rows):
    """Kirim pesan dengan inline keyboard grid (2D array). rows = [[btn, btn], [btn], ...]"""
    def make_btn(b):
        if "url" in b:
            return {"text": b["text"], "url": b["url"]}
        return {"text": b["text"], "callback_data": b["callback_data"]}
    keyboard = {"inline_keyboard": [[make_btn(b) for b in row] for row in rows]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def edit_msg(chat_id, message_id, text, remove_keyboard=False):
    try:
        payload = {"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "HTML"}
        if remove_keyboard:
            payload["reply_markup"] = {"inline_keyboard": []}
        requests.post(f"https://api.telegram.org/bot{BOT_TOKEN}/editMessageText", json=payload, timeout=10)
    except:
        pass

def delete_and_send(chat_id, msg_id, text):
    """Hapus pesan lama, kirim pesan baru. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def delete_and_send_keyboard(chat_id, msg_id, text, buttons):
    """Hapus pesan lama, kirim pesan baru dengan inline keyboard. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    return send_inline_keyboard(chat_id, text, buttons)

def send_with_cancel(chat_id, text, cancel_key):
    """Kirim pesan baru dengan tombol ❌ Batalkan. Return message_id."""
    keyboard = {"inline_keyboard": [[{"text": "❌ Batalkan", "callback_data": f"cancel:{cancel_key}"}]]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=10
        )
        return r.json().get("result", {}).get("message_id")
    except:
        return None

def delete_and_send_with_cancel(chat_id, msg_id, text, cancel_key):
    """Hapus pesan lama, kirim pesan baru dengan tombol ❌ Batalkan. Return message_id baru."""
    delete_msg(chat_id, msg_id)
    return send_with_cancel(chat_id, text, cancel_key)

def answer_callback_query(callback_query_id, text=""):
    requests.post(
        f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery",
        data={"callback_query_id": callback_query_id, "text": text}
    )

# ================= SETCOOKIE FLOW (OWNER) =================
def cmd_setcookie(chat_id):
    if not accounts:
        send_msg(chat_id, "❌ Belum ada akun. Tambah dulu dengan /addakun")
        return
    buttons = [{"text": f"📧 {acc['email']}", "callback_data": f"setcookie:{acc['email']}"} for acc in accounts]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:sc"})
    send_inline_keyboard(chat_id,
        "🍪 <b>SET COOKIE — OWNER</b>\n\n"
        "<blockquote>📋 Cara Penggunaan:\n"
        "1. Pilih email akun IVAS di bawah\n"
        "2. Kirim full JSON cookie dari browser\n"
        "3. Bot akan verifikasi session otomatis\n\n"
        "💡 Export cookie: DevTools → Application → Cookies</blockquote>\n\n"
        "👇 Pilih email:",
        buttons
    )

def handle_setcookie_callback(chat_id, user_id, email, callback_query_id, msg_id):
    answer_callback_query(callback_query_id, "✅ Email dipilih!")
    new_msg_id = delete_and_send_with_cancel(chat_id, msg_id,
        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"<blockquote>📤 Sekarang kirim full JSON cookie kamu.\n\n"
        f"Format array (export browser):\n"
        f"<code>[{{\"name\":\"key\",\"value\":\"val\"}}]</code>\n\n"
        f"Atau format dict:\n"
        f"<code>{{\"laravel_session\":\"...\",\"XSRF-TOKEN\":\"...\"}}</code></blockquote>",
        "sc"
    )
    pending_setcookie[user_id] = {"email": email, "msg_id": new_msg_id}

def process_cookie_input(chat_id, user_id, text):
    state = pending_setcookie.pop(user_id, None)
    if not state:
        return False

    email = state["email"]
    msg_id = state["msg_id"]

    cookie_dict = parse_cookie_input(text)
    if not cookie_dict:
        new_id = delete_and_send_with_cancel(chat_id, msg_id,
            f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
            f"📧 Email: <code>{email}</code>\n\n"
            f"❌ <b>Format JSON tidak valid!</b>\n"
            f"<blockquote>Kirim ulang cookie dalam format yang benar.</blockquote>",
            "sc"
        )
        pending_setcookie[user_id] = {"email": email, "msg_id": new_id}
        return True

    proc_id = delete_and_send(chat_id, msg_id,
        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Memverifikasi cookie..."
    )

    found = False
    with accounts_lock:
        for acc in accounts:
            if acc["email"] == email:
                found = True
                if "session" not in acc or acc["session"] is None:
                    acc["session"] = make_httpx_client()
                acc["session"].cookies.clear()
                acc["session"].cookies.update(cookie_dict)

                # Reset session fail flags agar run_bot tidak skip akun ini
                _session_notified[email] = False
                _session_fail_time.pop(email, None)
                _session_retry_time.pop(email, None)
                _session_recovered.pop(email, None)
                acc["last_login"] = 0

                if verify_cookie_session(acc):
                    acc["last_login"] = time.time()
                    # Ambil fresh cookies setelah verifikasi berhasil
                    fresh = extract_session_cookies(acc["session"])
                    cookies_to_save = fresh if fresh else cookie_dict
                    acc["cookies"] = cookies_to_save
                    # WAJIB dikunci — cegah race condition read-modify-write yang bisa
                    # menghapus/menimpa cookie akun lain saat addakun/setcookie berjalan
                    # bersamaan dengan thread polling akun lain (bug fatal multi-akun).
                    with _cookie_file_lock:
                        all_cdata = load_cookies()
                        all_cdata[email] = cookies_to_save
                        save_cookies(all_cdata)
                    # Reset keepalive timer & session fail flags — langsung aktif tanpa restart
                    _last_cookie_refresh[email] = time.time()
                    _session_notified[email] = False
                    _session_fail_time.pop(email, None)
                    _session_retry_time.pop(email, None)
                    _session_recovered.pop(email, None)
                    # Hapus ranges cache — paksa fetch fresh saat poll berikutnya
                    _ranges_cache.pop(email, None)
                    delete_and_send(chat_id, proc_id,
                        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
                        f"✅ <b>Cookie berhasil disimpan &amp; langsung aktif!</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"🔑 Total cookie: <b>{len(cookies_to_save)}</b> key\n"
                        f"✔️ Session aktif &amp; terverifikasi\n"
                        f"🔄 Fresh cookie langsung dipakai (tanpa restart)"
                        f"</blockquote>"
                    )
                    # Langsung cek ulang SEMUA akun lain — jangan sampai ada yang
                    # diam-diam ikut ke-invalidate tanpa ketahuan.
                    _recheck_other_accounts_async(email, chat_id)
                else:
                    # Cookie tidak valid — jangan simpan, beri pesan ramah
                    acc["session"].cookies.clear()
                    if acc.get("cookies"):
                        acc["session"].cookies.update(acc["cookies"])
                    delete_and_send(chat_id, proc_id,
                        f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
                        f"❌ <b>Cookie tidak valid / expired!</b>\n\n"
                        f"<blockquote>"
                        f"📧 Email: <code>{email}</code>\n"
                        f"🔑 Total cookie dikirim: <b>{len(cookie_dict)}</b> key\n\n"
                        f"Cookie ini tidak bisa login ke server IVAS.\n"
                        f"Silakan ambil cookie <b>fresh</b> dari browser dan coba lagi 😊\n\n"
                        f"💡 Tips: Buka DevTools → Application → Cookies → copy semua"
                        f"</blockquote>"
                    )
                return True

    if not found:
        delete_and_send(chat_id, proc_id,
            f"🍪 <b>SET COOKIE — OWNER</b>\n\n"
            f"❌ Email <code>{email}</code> tidak ditemukan di daftar akun."
        )
    return True

# ================= ADDCOOKIE FLOW (TOKEN) =================
def verify_cookie_dict(cookie_dict):
    try:
        session = make_httpx_client(timeout=15)
        session.cookies.update(cookie_dict)
        r = session.get(f"{BASE}/portal", timeout=15)
        return "/login" not in str(r.url)
    except:
        return False

def _cookie_guide_text(tag, email=None):
    """Teks arahan cookie yang konsisten — dipakai di beberapa tempat."""
    email_line = f"📧 Email: <code>{email}</code>\n\n" if email else ""
    return (
        f"🍪 <b>{tag}</b>\n\n"
        f"{email_line}"
        f"<blockquote>"
        f"📋 <b>Cara Export Cookie dari Browser:</b>\n\n"
        f"1️⃣ Login ke <b>IVAS</b> via browser\n"
        f"2️⃣ Buka <b>DevTools</b> (F12 / klik kanan → Inspect)\n"
        f"3️⃣ Tab <b>Application</b> → <b>Cookies</b> → pilih domain IVAS\n"
        f"4️⃣ Copy semua → paste sebagai JSON\n\n"
        f"📌 <b>Format yang diterima:</b>\n"
        f"• Array: <code>[{{\"name\":\"key\",\"value\":\"val\"}}]</code>\n"
        f"• Dict: <code>{{\"laravel_session\":\"...\",\"XSRF-TOKEN\":\"...\"}}</code>\n\n"
        f"💡 Bisa pakai ekstensi <b>EditThisCookie</b> / <b>Cookie Editor</b> untuk export otomatis"
        f"</blockquote>\n\n"
        f"📤 <b>Kirim JSON cookie kamu sekarang:</b>"
    )

def cmd_addcookie(chat_id, user_id):
    users = load_users()
    emails = users.get(str(user_id), {}).get("emails", [])
    if not emails:
        send_msg(chat_id, "❌ Belum ada email. Tambah dulu dengan /addemail")
        return

    # Kalau hanya 1 email — langsung ke step JSON, skip pemilihan email
    if len(emails) == 1:
        email = emails[0]
        msg_id = send_msg(chat_id, _cookie_guide_text("ADD COOKIE", email))
        pending_addcookie[user_id] = {"email": email, "msg_id": msg_id}
        return

    # Banyak email — tampilkan pilihan dulu
    buttons = [{"text": f"📧 {em}", "callback_data": f"addcookie:{em}"} for em in emails]
    buttons.append({"text": "❌ Batalkan", "callback_data": "cancel:ac"})
    send_inline_keyboard(chat_id,
        "🍪 <b>ADD COOKIE</b>\n\n"
        "<blockquote>Kamu punya beberapa akun.\n"
        "Pilih email yang ingin diperbarui cookie-nya:</blockquote>\n\n"
        "👇 Pilih email:",
        buttons
    )

def handle_addcookie_callback(chat_id, user_id, email, callback_query_id, msg_id):
    answer_callback_query(callback_query_id, "✅ Email dipilih!")
    users = load_users()
    emails = users.get(str(user_id), {}).get("emails", [])
    if email not in emails:
        answer_callback_query(callback_query_id, "❌ Email tidak ditemukan")
        return
    new_msg_id = delete_and_send_with_cancel(
        chat_id, msg_id, _cookie_guide_text("ADD COOKIE", email), "ac"
    )
    pending_addcookie[user_id] = {"email": email, "msg_id": new_msg_id}

def process_addcookie_input(chat_id, user_id, text):
    state = pending_addcookie.pop(user_id, None)
    if not state:
        return False

    email = state["email"]
    msg_id = state["msg_id"]

    cookie_dict = parse_cookie_input(text)
    if not cookie_dict:
        err_txt = (
            f"🍪 <b>ADD COOKIE</b>\n\n"
            f"📧 Email: <code>{email}</code>\n\n"
            f"❌ <b>Format JSON tidak valid!</b>\n\n"
            f"<blockquote>Pastikan kamu kirim JSON yang valid.\n"
            f"Contoh format array:\n"
            f"<code>[{{\"name\":\"laravel_session\",\"value\":\"xxx\"}}]</code>\n\n"
            f"Atau format dict:\n"
            f"<code>{{\"laravel_session\":\"xxx\",\"XSRF-TOKEN\":\"yyy\"}}</code>\n\n"
            f"📤 Kirim ulang JSON cookie kamu:</blockquote>"
        )
        new_id = delete_and_send_with_cancel(chat_id, msg_id, err_txt, "ac")
        pending_addcookie[user_id] = {"email": email, "msg_id": new_id}
        return True

    proc_id = delete_and_send(chat_id, msg_id,
        f"🍪 <b>ADD COOKIE</b>\n\n"
        f"📧 Email: <code>{email}</code>\n\n"
        f"⏳ Memverifikasi cookie ke server IVAS..."
    )

    try:
        # Verifikasi dulu pakai session sementara — jangan simpan sebelum verified
        tmp_session = make_httpx_client(timeout=15)
        tmp_session.cookies.update(cookie_dict)
        tmp_acc = {"session": tmp_session, "email": email, "cookies": cookie_dict, "csrf_token": ""}
        valid = verify_cookie_session(tmp_acc)

        if valid:
            # Ambil fresh cookies dari session setelah verifikasi berhasil
            fresh = extract_session_cookies(tmp_session)
            cookies_to_save = fresh if fresh else cookie_dict

            # WAJIB dikunci — cegah race condition read-modify-write yang bisa
            # menghapus/menimpa cookie akun premium lain (bug fatal multi-akun).
            with _cookie_file_lock:
                prem_cookies = load_premium_cookies()
                prem_cookies[email] = cookies_to_save
                save_premium_cookies(prem_cookies)

            # Langsung update _premium_acc_cache tanpa tunggu sync 30 detik
            if email in _premium_acc_cache:
                cached = _premium_acc_cache[email]
                cached["cookies"] = cookies_to_save
                cached["csrf_token"] = tmp_acc.get("csrf_token", "")
                cached["last_login"] = time.time()
                cached["session"].cookies.clear()
                cached["session"].cookies.update(cookies_to_save)
            else:
                new_acc = {
                    "email": email, "password": None,
                    "cookies": cookies_to_save,
                    "session": make_httpx_client(),
                    "last_login": time.time(),
                    "csrf_token": tmp_acc.get("csrf_token", ""),
                }
                new_acc["session"].cookies.update(cookies_to_save)
                _premium_acc_cache[email] = new_acc

            # Reset keepalive timer agar diprioritaskan saat ping berikutnya
            _last_cookie_refresh[email] = time.time()
            # Reset session fail flags
            _session_notified[email] = False
            _session_fail_time.pop(email, None)
            _session_retry_time.pop(email, None)
            _session_recovered.pop(email, None)
            # Hapus ranges cache — paksa fetch fresh saat poll berikutnya
            _ranges_cache.pop(email, None)
            # Paksa run_bot sync segera agar thread worker baru langsung spawn
            global _force_bot_sync
            _force_bot_sync = True

            delete_and_send(chat_id, proc_id,
                f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
                f"✅ <b>Cookie berhasil disimpan &amp; langsung aktif!</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔑 Total cookie: <b>{len(cookies_to_save)}</b> key\n"
                f"✔️ Session aktif &amp; terverifikasi\n"
                f"🔄 Fresh cookie langsung dipakai (tanpa restart)"
                f"</blockquote>"
            )
        else:
            # Cookie tidak valid — jangan simpan, beri pesan ramah
            delete_and_send(chat_id, proc_id,
                f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
                f"❌ <b>Cookie tidak valid / expired!</b>\n\n"
                f"<blockquote>"
                f"📧 Email: <code>{email}</code>\n"
                f"🔑 Total cookie dikirim: <b>{len(cookie_dict)}</b> key\n\n"
                f"Cookie ini tidak bisa login ke server IVAS.\n"
                f"Silakan ambil cookie <b>fresh</b> dari browser dan coba lagi 😊\n\n"
                f"💡 Tips: Buka DevTools → Application → Cookies → copy semua"
                f"</blockquote>"
            )
    except Exception as e:
        delete_and_send(chat_id, proc_id,
            f"🍪 <b>ADD COOKIE — TOKEN</b>\n\n"
            f"❌ Terjadi error saat verifikasi: <code>{e}</code>"
        )
    return True

# ===== LOCK PER-AKUN =====
# Mencegah 2 thread (mis. thread polling akun A + thread recheck yang dipicu
# saat setcookie akun B) barengan login/refresh cookie utk akun yang SAMA.
# Tanpa lock ini, 2 request login bisa saling timpa cookies.clear()/update()
# di httpx.Client yang sama → session jadi rusak/keliru walau akunnya valid.
_acc_op_locks       = {}
_acc_op_locks_guard = threading.Lock()

def _get_acc_lock(email):
    if not email:
        return threading.Lock()
    with _acc_op_locks_guard:
        lock = _acc_op_locks.get(email)
        if lock is None:
            lock = threading.Lock()
            _acc_op_locks[email] = lock
        return lock

def ensure_login(acc):
    """Wrapper terkunci per-akun — body asli di _ensure_login_inner()."""
    email = acc.get("email", "")
    with _get_acc_lock(email):
        return _ensure_login_inner(acc)

def _ensure_login_inner(acc):
    now = time.time()
    email = acc.get("email", "")
    if now - acc.get("last_login", 0) < LOGIN_COOLDOWN:
        return True

    if acc.get("cookies"):
        if "session" not in acc or acc["session"] is None:
            acc["session"] = make_httpx_client()
        acc["session"].cookies.clear()
        acc["session"].cookies.update(acc["cookies"])
        if verify_cookie_session(acc):
            acc["last_login"] = now
            fresh = extract_session_cookies(acc["session"])
            if fresh and fresh != acc.get("cookies"):
                acc["cookies"] = fresh
                save_fresh_cookies_auto(email, fresh)
            if _session_notified.get(email):
                _session_notified[email] = False
                _session_fail_time.pop(email, None)
                _session_retry_time.pop(email, None)
                if not _session_recovered.get(email):
                    _session_recovered[email] = True
                    requests.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                        data={"chat_id": OWNER_ID, "text": f"✅ <b>SESSION PULIH</b>\n\n📧 Email: <code>{email}</code>\nSession berhasil aktif kembali secara otomatis.", "parse_mode": "HTML"},
                        timeout=10
                    )
            return True
        _log("COOKIE", f"expired [{email}] — coba login ulang", Fore.YELLOW)

    if login(acc):
        acc["last_login"] = now
        if _session_notified.get(email):
            _session_notified[email] = False
            _session_fail_time.pop(email, None)
            _session_retry_time.pop(email, None)
            if not _session_recovered.get(email):
                _session_recovered[email] = True
                requests.post(
                    f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                    data={"chat_id": OWNER_ID, "text": f"✅ <b>SESSION PULIH</b>\n\n📧 Email: <code>{email}</code>\nLogin password berhasil, session aktif kembali.", "parse_mode": "HTML"},
                    timeout=10
                )
        return True

    if not _session_notified.get(email):
        _session_notified[email] = True
        _session_recovered[email] = False
        _session_fail_time[email] = now
        _log("SESSION", f"gagal [{email}] — notif dikirim", Fore.RED)
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={
                "chat_id": OWNER_ID,
                "text": (
                    f"⚠️ <b>SESSION EXPIRED</b>\n\n"
                    f"📧 Email: <code>{email}</code>\n"
                    f"❌ Cookie expired & login password gagal.\n\n"
                    f"Bot akan otomatis retry setiap 10 menit.\n"
                    f"Perbarui cookie dengan /setcookie atau /addcookie."
                ),
                "parse_mode": "HTML"
            },
            timeout=10
        )

    _session_retry_time[email] = now
    return False
 
def _recheck_other_accounts(exclude_email, chat_id=None):
    """
    Dipanggil setelah SATU akun berhasil login/setcookie.
    Cek ulang SEMUA akun lain SAAT ITU JUGA (bukan nunggu siklus retry 5-10 menit),
    supaya kalau ada akun lain yang ikut ke-invalidate — misal karena cookie-nya
    diambil dari browser yang sama lalu di-logout untuk ambil cookie akun baru —
    langsung ketahuan dan langsung dicoba dipulihkan (verify cookie / re-login password).
    """
    try:
        with accounts_lock:
            others = [a for a in accounts if a.get("email") != exclude_email]
        casualties = []
        for a in others:
            em = a.get("email", "")
            if not em:
                continue
            was_ok = not _session_notified.get(em)
            a["last_login"] = 0        # paksa verifikasi ulang sekarang juga
            still_ok = ensure_login(a)  # sudah terkunci per-akun di dalam
            if was_ok and not still_ok:
                casualties.append(em)
        if casualties and chat_id:
            send_msg(chat_id,
                "⚠️ <b>AKUN LAIN IKUT TERDAMPAK!</b>\n\n"
                "<blockquote>" +
                "\n".join(f"📧 <code>{em}</code>" for em in casualties) +
                "\n\nKemungkinan cookie akun di atas diambil dari browser yang sama "
                "lalu ke-logout saat kamu ambil cookie akun baru.\n\n"
                "Kirim cookie <b>fresh</b> lagi lewat /setcookie untuk akun di atas 🙏"
                "</blockquote>"
            )
        elif casualties:
            for em in casualties:
                _log("SESSION", f"⚠️ {em} ikut ke-invalidate setelah update akun lain", Fore.RED)
    except Exception as e:
        _log("WARN", f"recheck other accounts: {e}", Fore.YELLOW)

def _recheck_other_accounts_async(exclude_email, chat_id=None):
    threading.Thread(
        target=_recheck_other_accounts, args=(exclude_email, chat_id),
        daemon=True, name="recheck-others"
    ).start()

def cek_ivas(chat_id=None, user_id=None):
    """Alias — sekarang memanggil cek_traffic_range (traffic WhatsApp)."""
    cek_traffic_range(chat_id if chat_id else OWNER_ID, user_id if user_id else OWNER_ID)


# ================= CEK TRAFFIC RANGE WHATSAPP =================
IVASMS_BASE = "https://www.ivasms.com"

def _fetch_whatsapp_sms_datatable(cookies: dict, base_url: str,
                                   limit: int = 100, search: str = "") -> tuple:
    """
    Ambil data SMS WhatsApp dari DataTable JSON API private IVAS server.
    Endpoint: GET {base_url}/portal/sms/test/sms?app=WhatsApp&draw=1&...
    Return (list_of_rows, total_filtered_int).
    Setiap row: {"range": str, "number": str, "message": str, "senttime": str}
    """
    results = []
    total_filtered = 0
    try:
        tmp = make_httpx_client(timeout=15)
        tmp.cookies.update(cookies)
        params = {
            "app":              "WhatsApp",
            "draw":             1,
            "start":            0,
            "length":           limit,
            "search[value]":    search,
            "search[regex]":    "false",
            "order[0][column]": 4,
            "order[0][dir]":    "desc",
            "columns[0][data]": "range",
            "columns[0][name]": "range",
            "columns[1][data]": "termination.test_number",
            "columns[1][name]": "termination.test_number",
            "columns[2][data]": "originator",
            "columns[2][name]": "originator",
            "columns[3][data]": "messagedata",
            "columns[3][name]": "messagedata",
            "columns[4][data]": "senttime",
            "columns[4][name]": "senttime",
        }
        r = tmp.get(
            f"{base_url}/portal/sms/test/sms",
            params=params,
            headers={
                "User-Agent":        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                "Accept":            "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With":  "XMLHttpRequest",
                "Referer":           f"{base_url}/portal/sms/test/sms?app=WhatsApp",
            },
            timeout=15
        )
        if r.status_code != 200 or "/login" in str(r.url):
            return results, total_filtered
        d = r.json()
        total_filtered = int(d.get("recordsFiltered", 0))
        for row in d.get("data", []):
            range_name = row.get("range", "")
            message    = row.get("messagedata", "")
            senttime   = row.get("senttime", "")
            try:
                tn      = row.get("termination", {})
                raw_num = tn.get("test_number", "") if isinstance(tn, dict) else ""
                html_num = BeautifulSoup(raw_num, "html.parser").get_text(strip=True)
            except Exception:
                html_num = ""
            if range_name:
                results.append({
                    "range":    range_name,
                    "number":   html_num,
                    "message":  message,
                    "senttime": senttime,
                })
    except Exception as e:
        _log("CEKRANGE", f"datatable fetch error ({base_url}): {e}", Fore.YELLOW)
    return results, total_filtered


def _strip_range_code(full_range: str) -> str:
    """'INDONESIA 228067' → 'INDONESIA' | 'IVORY COAST 4297' → 'IVORY COAST'"""
    parts = full_range.strip().split()
    if parts and parts[-1].isdigit():
        return " ".join(parts[:-1])
    return full_range.strip()


def _get_cookies_from_users(user_id) -> list:
    """Kembalikan list cookies dict yang tersedia untuk user ini (snapshot aman)."""
    cookies_list = []
    try:
        if is_owner(user_id):
            with accounts_lock:
                for acc in accounts:
                    c = acc.get("cookies", {})
                    if c:
                        cookies_list.append(c)
        else:
            for em in get_user_emails(user_id):
                a = get_acc_by_email(em)
                if a:
                    c = a.get("cookies", {})
                    if c:
                        cookies_list.append(c)
    except Exception:
        pass
    return cookies_list


def _get_owner_cookies() -> list:
    """Ambil cookies dari semua akun OWNER — dipakai untuk cekivas/cekivasv2/toprcv/rangeterbaru."""
    cookies_list = []
    try:
        with accounts_lock:
            for acc in accounts:
                c = acc.get("cookies", {})
                if c:
                    cookies_list.append(c)
    except Exception:
        pass
    return cookies_list


def cek_traffic_range(chat_id, user_id):
    """
    /cekivas — Traffic WhatsApp saat ini (60 menit terakhir) dari akun IVAS owner.
    Auto-delete pesan sebelumnya di chat yang sama.
    """
    try:
        from collections import Counter
        send_cek_msg(chat_id, "⏳ Mengambil traffic sekarang...")

        cookies_list = _get_owner_cookies()

        all_rows = []
        for ck in cookies_list:
            try:
                rows, _ = _fetch_whatsapp_sms_datatable(ck, BASE, limit=500)
                if rows:
                    all_rows = rows
                    break
            except Exception as e:
                _log("CEKIVAS", f"fetch error: {e}", Fore.YELLOW)

        # Smart: cari window dengan traffic RATE tertinggi (SMS per menit)
        _windows = [
            (timedelta(minutes=1),  1,  "1 menit terakhir"),
            (timedelta(minutes=5),  5,  "5 menit terakhir"),
            (timedelta(minutes=15), 15, "15 menit terakhir"),
            (timedelta(minutes=30), 30, "30 menit terakhir"),
        ]
        now_dt   = datetime.now()
        best_rows  = []
        best_label = ""
        best_rate  = -1.0

        for delta, mins, lbl in _windows:
            cutoff = now_dt - delta
            tmp = []
            for r in all_rows:
                try:
                    st = datetime.strptime(r["senttime"][:19], "%Y-%m-%d %H:%M:%S")
                    if st >= cutoff:
                        tmp.append(r)
                except Exception:
                    pass
            rate = len(tmp) / mins  # SMS per menit
            if rate > best_rate:
                best_rate  = rate
                best_rows  = tmp
                best_label = lbl

        sms_rows  = best_rows
        win_label = best_label

        fallback = not sms_rows and bool(all_rows)
        if fallback:
            sms_rows  = all_rows[:50]
            win_label = "50 SMS terbaru"

        now_str = now_dt.strftime("%d/%m/%Y %H:%M:%S WIB")

        if sms_rows:
            country_counts = Counter(_strip_range_code(r["range"]) for r in sms_rows)
            total_sms = len(sms_rows)
            total_neg = len(country_counts)
            top20     = country_counts.most_common(20)

            body = [f"{total_neg} negara — {total_sms} SMS ({win_label})\n"]
            for i, (negara, cnt) in enumerate(top20, 1):
                body.append(f"{i:2}. {negara}  {cnt} SMS")
            body.append(f"\nKetik /toprcv [negara] untuk detail range.")
            lines = [
                "📡 <b>TRAFFIC WHATSAPP SEKARANG</b>",
                "<blockquote>" + "\n".join(body) + "</blockquote>",
                f"🕐 {now_str}",
            ]
        else:
            if not cookies_list:
                isi = "Tidak ada akun IVAS aktif.\nHubungi owner."
            else:
                isi = "Belum ada SMS WhatsApp masuk saat ini.\nCoba lagi beberapa saat."
            lines = [
                "📡 <b>TRAFFIC WHATSAPP SEKARANG</b>",
                f"<blockquote>{isi}</blockquote>",
                f"🕐 {now_str}",
            ]

        send_cek_msg(chat_id, "\n".join(lines))

    except Exception as e:
        _log("CEKIVAS", f"fatal: {e}", Fore.RED)
        try:
            send_cek_msg(chat_id, f"❌ Gagal ambil data: <code>{str(e)[:200]}</code>")
        except Exception:
            pass


_COUNTRY_ALIASES = {
    "TAJIK":         "TAJIKISTAN",
    "TAJIKIS":       "TAJIKISTAN",
    "UZBEK":         "UZBEKISTAN",
    "KAZAKH":        "KAZAKHSTAN",
    "KAZ":           "KAZAKHSTAN",
    "KYRGYZ":        "KYRGYZSTAN",
    "TURKMEN":       "TURKMENISTAN",
    "AFGHAN":        "AFGHANISTAN",
    "AFGHAN":        "AFGHANISTAN",
    "INDO":          "INDONESIA",
    "INDON":         "INDONESIA",
    "INA":           "INDONESIA",
    "IDN":           "INDONESIA",
    "MALAY":         "MALAYSIA",
    "MYS":           "MALAYSIA",
    "THAI":          "THAILAND",
    "THA":           "THAILAND",
    "VIET":          "VIETNAM",
    "VNM":           "VIETNAM",
    "PHIL":          "PHILIPPINES",
    "PHILI":         "PHILIPPINES",
    "PHL":           "PHILIPPINES",
    "PINOY":         "PHILIPPINES",
    "CAMBO":         "CAMBODIA",
    "KHM":           "CAMBODIA",
    "MYAN":          "MYANMAR",
    "BURMA":         "MYANMAR",
    "LAOS":          "LAO",
    "TIMOR":         "TIMOR LESTE",
    "BRUNEI":        "BRUNEI DARUSSALAM",
    "SINGA":         "SINGAPORE",
    "SGP":           "SINGAPORE",
    "PAKIS":         "PAKISTAN",
    "PAK":           "PAKISTAN",
    "BANGLA":        "BANGLADESH",
    "BGD":           "BANGLADESH",
    "SRILANKA":      "SRI LANKA",
    "LANKA":         "SRI LANKA",
    "LKA":           "SRI LANKA",
    "NEPAL":         "NEPAL",
    "NPL":           "NEPAL",
    "BHUTAN":        "BHUTAN",
    "MONGOL":        "MONGOLIA",
    "MNG":           "MONGOLIA",
    "SAUDI":         "SAUDI ARABIA",
    "KSA":           "SAUDI ARABIA",
    "ARAB":          "SAUDI ARABIA",
    "UAE":           "UNITED ARAB EMIRATES",
    "EMIRAT":        "UNITED ARAB EMIRATES",
    "DUBAI":         "UNITED ARAB EMIRATES",
    "QATAR":         "QATAR",
    "KUWAI":         "KUWAIT",
    "KWT":           "KUWAIT",
    "BAHRAIN":       "BAHRAIN",
    "OMAN":          "OMAN",
    "YEMEN":         "YEMEN",
    "IRAQ":          "IRAQ",
    "IRAN":          "IRAN",
    "SYRIA":         "SYRIA",
    "JORDAN":        "JORDAN",
    "LIBAN":         "LEBANON",
    "LEBANO":        "LEBANON",
    "ISRAEL":        "ISRAEL",
    "PALEST":        "PALESTINE",
    "EGYPT":         "EGYPT",
    "MESIR":         "EGYPT",
    "NIGERIA":       "NIGERIA",
    "NGA":           "NIGERIA",
    "GHANA":         "GHANA",
    "KENYA":         "KENYA",
    "ETHIOPIA":      "ETHIOPIA",
    "ETHIOP":        "ETHIOPIA",
    "TANZANI":       "TANZANIA",
    "UGANDA":        "UGANDA",
    "Rwanda":        "RWANDA",
    "CAMEROON":      "CAMEROON",
    "CAMERO":        "CAMEROON",
    "IVORY":         "IVORY COAST",
    "BENIN":         "BENIN",
    "TOGO":          "TOGO",
    "GHANA":         "GHANA",
    "SENEGAL":       "SENEGAL",
    "MALI":          "MALI",
    "GUINEA":        "GUINEA",
    "SIERRA":        "SIERRA LEONE",
    "LIBERIA":       "LIBERIA",
    "MOZAMB":        "MOZAMBIQUE",
    "ZAMBIA":        "ZAMBIA",
    "ZIMBAB":        "ZIMBABWE",
    "ANGOLA":        "ANGOLA",
    "CONGO":         "CONGO",
    "GABON":         "GABON",
    "MAURITIUS":     "MAURITIUS",
    "MADAGA":        "MADAGASCAR",
    "SUDAN":         "SUDAN",
    "SOMALIA":       "SOMALIA",
    "ERITREA":       "ERITREA",
    "DJIBOUTI":      "DJIBOUTI",
    "RUSSIA":        "RUSSIA",
    "RUSIA":         "RUSSIA",
    "RUS":           "RUSSIA",
    "UKRAINE":       "UKRAINE",
    "UKRAINA":       "UKRAINE",
    "UKRAIN":        "UKRAINE",
    "BELARUS":       "BELARUS",
    "GEORGIA":       "GEORGIA",
    "ARMENIA":       "ARMENIA",
    "AZERBAI":       "AZERBAIJAN",
    "MOLDOVA":       "MOLDOVA",
    "POLANDIA":      "POLAND",
    "POLAND":        "POLAND",
    "POL":           "POLAND",
    "CZECH":         "CZECH REPUBLIC",
    "SLOVAK":        "SLOVAKIA",
    "HUNGARI":       "HUNGARY",
    "ROMANI":        "ROMANIA",
    "BULGARI":       "BULGARIA",
    "SERBIA":        "SERBIA",
    "KROASI":        "CROATIA",
    "CROATIA":       "CROATIA",
    "ALBANIA":       "ALBANIA",
    "TURKI":         "TURKEY",
    "TURKEY":        "TURKEY",
    "TUR":           "TURKEY",
    "GREECE":        "GREECE",
    "YUNANI":        "GREECE",
    "PORTUGAL":      "PORTUGAL",
    "PORTUGAL":      "PORTUGAL",
    "SPAIN":         "SPAIN",
    "SPANYOL":       "SPAIN",
    "FRANCE":        "FRANCE",
    "PRANCIS":       "FRANCE",
    "GERMANY":       "GERMANY",
    "JERMAN":        "GERMANY",
    "ITALY":         "ITALY",
    "ITALIA":        "ITALY",
    "UK":            "UNITED KINGDOM",
    "INGGRIS":       "UNITED KINGDOM",
    "BRITAIN":       "UNITED KINGDOM",
    "ENGLAND":       "UNITED KINGDOM",
    "SWEDEN":        "SWEDEN",
    "SWEDIA":        "SWEDEN",
    "NORWAY":        "NORWAY",
    "NORWAY":        "NORWAY",
    "DENMARK":       "DENMARK",
    "FINLAND":       "FINLAND",
    "NETHERL":       "NETHERLANDS",
    "BELANDA":       "NETHERLANDS",
    "BELGIA":        "BELGIUM",
    "BELGIUM":       "BELGIUM",
    "SWISS":         "SWITZERLAND",
    "AUSTRIA":       "AUSTRIA",
    "CHINA":         "CHINA",
    "TIONGKOK":      "CHINA",
    "CHN":           "CHINA",
    "JAPAN":         "JAPAN",
    "JEPANG":        "JAPAN",
    "JPN":           "JAPAN",
    "KOREA":         "SOUTH KOREA",
    "SKOREA":        "SOUTH KOREA",
    "KORSEL":        "SOUTH KOREA",
    "TAIWAN":        "TAIWAN",
    "HONGKONG":      "HONG KONG",
    "HKONG":         "HONG KONG",
    "USA":           "UNITED STATES",
    "AMERI":         "UNITED STATES",
    "KANADA":        "CANADA",
    "CANADA":        "CANADA",
    "BRAZIL":        "BRAZIL",
    "BRASIL":        "BRAZIL",
    "MEKSIKO":       "MEXICO",
    "MEXICO":        "MEXICO",
    "ARGENTIN":      "ARGENTINA",
    "COLOMBIA":      "COLOMBIA",
    "PERU":          "PERU",
    "CHILE":         "CHILE",
    "VENEZU":        "VENEZUELA",
    "ECUADOR":       "ECUADOR",
    "BOLIVIA":       "BOLIVIA",
    "PARAGUAY":      "PARAGUAY",
    "URUGUAY":       "URUGUAY",
    "AUSTRA":        "AUSTRALIA",
    "AUS":           "AUSTRALIA",
    "NEWZEAL":       "NEW ZEALAND",
    "NZ":            "NEW ZEALAND",
    "PAPUA":         "PAPUA NEW GUINEA",
    "PNG":           "PAPUA NEW GUINEA",
}


def _smart_country_expand(q: str) -> list:
    """
    Smart expand nama negara pendek/alias ke kemungkinan nama penuh.
    Return list nama yang akan dicoba, prioritas dari yang paling spesifik.
    """
    q_up = q.strip().upper()
    candidates = [q_up]

    # Cek alias dict exact match
    if q_up in _COUNTRY_ALIASES:
        full = _COUNTRY_ALIASES[q_up]
        if full not in candidates:
            candidates.append(full)

    # Cek alias dict prefix match (kalau q adalah prefix dari alias key)
    for alias, full_name in _COUNTRY_ALIASES.items():
        if alias.startswith(q_up) or q_up.startswith(alias):
            if full_name not in candidates:
                candidates.append(full_name)

    # Cek partial match — alias yang mengandung q
    for alias, full_name in _COUNTRY_ALIASES.items():
        if q_up in alias or q_up in full_name:
            if full_name not in candidates:
                candidates.append(full_name)

    return candidates


def top_rcv(chat_id, user_id, query: str):
    """
    /toprcv [nama negara] — Detail range + kode aktif per negara, pakai akun owner.
    Smart search: otomatis expand nama pendek/alias ke nama negara penuh.
    Range codes ditampilkan dalam <code> tag agar bisa langsung diklik & disalin.
    """
    try:
        from collections import Counter
        q = query.strip().upper()
        if not q:
            send_cek_msg(chat_id,
                "❓ <b>Cara pakai /toprcv</b>\n\n"
                "<blockquote>/toprcv INDONESIA\n/toprcv TAJIK\n/toprcv KAZAKH\n/toprcv SAUDI</blockquote>\n\n"
                "💡 <i>Bot otomatis cari negara walau nama pendek/singkat!</i>\n"
                "Ketik /cekivas dulu untuk lihat negara yang aktif."
            )
            return

        # Smart expand: cari semua kemungkinan nama negara dari query pendek
        search_candidates = _smart_country_expand(q)
        display_q = search_candidates[0] if search_candidates else q

        send_cek_msg(chat_id, f"⏳ Mencari range <b>{display_q}</b>...")

        cookies_list = _get_owner_cookies()
        if not cookies_list:
            send_cek_msg(chat_id, "❌ Tidak ada akun IVAS aktif.")
            return

        sms_rows   = []
        total_filt = 0
        matched_q  = display_q

        # Coba setiap kandidat nama negara sampai ada hasil
        for candidate in search_candidates:
            if sms_rows:
                break
            for ck in cookies_list:
                try:
                    rows, total_f = _fetch_whatsapp_sms_datatable(
                        ck, BASE, limit=500, search=candidate
                    )
                    if rows:
                        sms_rows   = rows
                        total_filt = total_f
                        matched_q  = candidate
                        break
                except Exception as e:
                    _log("TOPRCV", f"fetch error ({candidate}): {e}", Fore.YELLOW)

        # Fallback: jika masih kosong, coba langsung IVASMS_BASE
        if not sms_rows:
            for candidate in search_candidates[:2]:
                if sms_rows:
                    break
                for ck in cookies_list:
                    try:
                        rows, total_f = _fetch_whatsapp_sms_datatable(
                            ck, IVASMS_BASE, limit=500, search=candidate
                        )
                        if rows:
                            sms_rows   = rows
                            total_filt = total_f
                            matched_q  = candidate
                            break
                    except Exception:
                        pass

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M WIB")

        if sms_rows:
            range_counts = Counter(r["range"] for r in sms_rows)
            top_ranges   = range_counts.most_common(50)
            shown        = len(top_ranges)

            found_label = f" (ditemukan sebagai: {matched_q})" if matched_q != q else ""
            body = [f"🔍 Query: <b>{q}</b>{found_label}\n{total_filt:,} SMS — {shown} range\n"]
            for i, (full_rng, cnt) in enumerate(top_ranges, 1):
                # Wrap range code dalam <code> agar bisa diklik/disalin langsung
                body.append(f"{i:2}. <code>{full_rng}</code>  — {cnt} SMS")
            body.append("\n👆 Klik kode range di atas untuk menyalin otomatis!")
            lines = [
                f"📊 <b>TOP RANGE: {matched_q}</b>",
                "<blockquote>" + "\n".join(body) + "</blockquote>",
                f"🕐 {now_str}",
            ]
        else:
            tried = ", ".join(search_candidates[:3])
            lines = [
                f"📊 <b>TOP RANGE: {q}</b>",
                f"<blockquote>Tidak ada SMS WhatsApp untuk <b>{q}</b> saat ini.\n\n"
                f"Sudah dicoba: {tried}\n\n"
                f"Ketik /cekivas untuk lihat negara yang aktif sekarang.</blockquote>",
                f"🕐 {now_str}",
            ]

        send_cek_msg(chat_id, "\n".join(lines))

    except Exception as e:
        _log("TOPRCV", f"fatal: {e}", Fore.RED)
        try:
            send_cek_msg(chat_id, f"❌ Gagal ambil data: <code>{str(e)[:200]}</code>")
        except Exception:
            pass


def range_terbaru(chat_id, user_id):
    """
    /rangeterbaru — Range paling baru terima SMS WhatsApp, pakai akun owner.
    Auto-delete pesan sebelumnya.
    """
    try:
        send_cek_msg(chat_id, "⏳ Mengambil range terbaru...")

        cookies_list = _get_owner_cookies()
        if not cookies_list:
            send_cek_msg(chat_id, "❌ Tidak ada akun IVAS aktif.")
            return

        sms_rows = []
        for ck in cookies_list:
            try:
                rows, _ = _fetch_whatsapp_sms_datatable(ck, BASE, limit=100)
                if rows:
                    sms_rows = rows
                    break
            except Exception as e:
                _log("RANGETERBARU", f"fetch error: {e}", Fore.YELLOW)

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M WIB")

        if sms_rows:
            seen, seen_s = [], set()
            for r in sms_rows:
                if r["range"] not in seen_s:
                    seen_s.add(r["range"])
                    seen.append(r)

            body = [f"{len(seen)} range aktif menerima SMS\n"]
            for i, r in enumerate(seen[:30], 1):
                waktu = r["senttime"][11:16] if len(r["senttime"]) >= 16 else r["senttime"]
                body.append(f"{i:2}. {r['range']}  {waktu}")
            body.append("\nSalin kode range → ketik /toprcv [negara] untuk detail.")
            lines = [
                "🆕 <b>RANGE TERBARU WHATSAPP</b>",
                "<blockquote>" + "\n".join(body) + "</blockquote>",
                f"🕐 {now_str}",
            ]
        else:
            lines = [
                "🆕 <b>RANGE TERBARU WHATSAPP</b>",
                "<blockquote>Belum ada SMS WhatsApp terbaru saat ini.</blockquote>",
                f"🕐 {now_str}",
            ]

        send_cek_msg(chat_id, "\n".join(lines))

    except Exception as e:
        _log("RANGETERBARU", f"fatal: {e}", Fore.RED)
        try:
            send_cek_msg(chat_id, f"❌ Gagal ambil data: <code>{str(e)[:200]}</code>")
        except Exception:
            pass


# ── CEKIVASV2 — filter SMS berdasarkan rentang waktu terakhir ──────────────

_CEKV2_LABELS = {
    10:   "10 Detik",
    60:   "1 Menit",
    600:  "10 Menit",
    1800: "30 Menit",
}

def cek_ivas_v2(chat_id, user_id):
    """
    /cekivasv2 — Tampilkan pilihan rentang waktu, lalu filter SMS sesuai pilihan.
    """
    # hapus pesan cek sebelumnya dulu (anti-spam)
    with _last_cek_lock:
        prev_id = _last_cek_msgs.get(chat_id)
    if prev_id:
        delete_msg(chat_id, prev_id)

    msg_text = (
        "📡 <b>CEK SMS WHATSAPP</b>\n\n"
        "<blockquote>Pilih rentang waktu untuk cek SMS yang masuk:</blockquote>"
    )
    keyboard = {"inline_keyboard": [
        [
            {"text": "⏱ 10 Detik", "callback_data": "cekv2:10"},
            {"text": "⏱ 1 Menit",  "callback_data": "cekv2:60"},
        ],
        [
            {"text": "⏱ 10 Menit", "callback_data": "cekv2:600"},
            {"text": "⏱ 30 Menit", "callback_data": "cekv2:1800"},
        ],
    ]}
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            json={
                "chat_id":      chat_id,
                "text":         msg_text,
                "parse_mode":   "HTML",
                "reply_markup": keyboard,
            },
            timeout=10,
        )
        res = r.json()
        if res.get("ok"):
            with _last_cek_lock:
                _last_cek_msgs[chat_id] = res["result"]["message_id"]
    except Exception as e:
        _log("CEKIVASV2", f"send keyboard error: {e}", Fore.YELLOW)


def cek_ivas_v2_exec(chat_id, user_id, seconds: int, cq_id: str, orig_msg_id: int):
    """
    Eksekusi filter SMS WhatsApp dalam rentang N detik terakhir.
    Tampilkan semua range yang menerima SMS dalam waktu itu.
    """
    try:
        label = _CEKV2_LABELS.get(seconds, f"{seconds} detik")
        answer_callback_query(cq_id, f"⏳ Mengambil SMS {label} terakhir...")
        delete_msg(chat_id, orig_msg_id)

        send_cek_msg(chat_id, f"⏳ Mencari range aktif {label} terakhir...")

        cookies_list = _get_owner_cookies()
        if not cookies_list:
            send_cek_msg(chat_id, "❌ Tidak ada akun IVAS aktif.")
            return

        sms_rows = []
        for ck in cookies_list:
            try:
                rows, _ = _fetch_whatsapp_sms_datatable(ck, BASE, limit=500)
                if rows:
                    sms_rows = rows
                    break
            except Exception as e:
                _log("CEKIVASV2", f"fetch error: {e}", Fore.YELLOW)

        cutoff = datetime.now() - timedelta(seconds=seconds)
        filtered = []
        for r in sms_rows:
            try:
                st = datetime.strptime(r["senttime"][:19], "%Y-%m-%d %H:%M:%S")
                if st >= cutoff:
                    filtered.append((st, r))
            except Exception:
                pass

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S WIB")

        if filtered:
            # kelompokkan per range, simpan waktu terbaru + jumlah SMS
            range_data = {}
            for st, r in filtered:
                rng = r["range"]
                if rng not in range_data:
                    range_data[rng] = {"last_time": st, "count": 0}
                range_data[rng]["count"] += 1
                if st > range_data[rng]["last_time"]:
                    range_data[rng]["last_time"] = st

            # urutkan berdasarkan waktu terbaru
            sorted_ranges = sorted(range_data.items(), key=lambda x: x[1]["last_time"], reverse=True)

            total_sms   = len(filtered)
            total_range = len(sorted_ranges)

            body = [f"{total_range} range — {total_sms} SMS masuk\n"]
            for i, (rng, info) in enumerate(sorted_ranges[:50], 1):
                waktu = info["last_time"].strftime("%H:%M:%S")
                cnt   = info["count"]
                body.append(f"{i:2}. {rng}  {cnt} SMS  {waktu}")
            if total_range > 50:
                body.append(f"\n... dan {total_range - 50} range lainnya")
            lines = [
                f"📡 <b>RANGE AKTIF — {label.upper()} TERAKHIR</b>",
                "<blockquote>" + "\n".join(body) + "</blockquote>",
                f"🕐 {now_str}",
            ]
        else:
            lines = [
                f"📡 <b>RANGE AKTIF — {label.upper()} TERAKHIR</b>",
                f"<blockquote>Tidak ada SMS masuk dalam {label} terakhir.\nCoba pilih rentang waktu yang lebih panjang.</blockquote>",
                f"🕐 {now_str}",
            ]

        send_cek_msg(chat_id, "\n".join(lines))

    except Exception as e:
        _log("CEKIVASV2", f"fatal: {e}", Fore.RED)
        try:
            send_cek_msg(chat_id, f"❌ Gagal ambil data: <code>{str(e)[:200]}</code>")
        except Exception:
            pass

# ================= UTILS =================
def extract_otp(text):
    m = re.search(r"\b(\d{3}[- ]?\d{3})\b", text)
    if not m: return None
    otp = m.group(0).replace(" ", "")  
    if len(otp) not in (6, 7): return None
    if len(otp) == 6: otp = otp[:3] + "-" + otp[3:]
    return otp    
        
def return_all_base(acc):
    try:
        session = acc["session"]
        # Refresh CSRF dari halaman numbers sebelum POST
        numbers_page = f"{BASE}/portal/numbers"
        fresh_csrf = _fetch_fresh_csrf_from_page(session, numbers_page, acc.get("csrf_token", ""))
        if fresh_csrf:
            acc["csrf_token"] = fresh_csrf
        url = RETURN_ALL_URL
        headers = {
            "X-Requested-With": "XMLHttpRequest",
            "Referer":          numbers_page,
            "Origin":           BASE,
            "Content-Type":     "application/x-www-form-urlencoded; charset=UTF-8",
        }
        r = session.post(url, headers=headers, data={"_token": fresh_csrf}, timeout=30)
        if r.status_code not in (200, 201, 204):
            return False, f"HTTP {r.status_code}: {r.text[:200]}"
        # Cek response body untuk konfirmasi sukses
        try:
            jr  = r.json()
            # Response asli IVAS: {"NumberDoneRemove":"all numbers","count":N,"message":"Successfully returned N numbers"}
            st    = jr.get("status", jr.get("success", jr.get("code", "")))
            msg   = str(jr.get("message", jr.get("msg", jr.get("NumberDoneRemove", ""))))
            count = jr.get("count", -1)
            ok = (
                count >= 0                          # ada field count → endpoint bener
                or jr.get("NumberDoneRemove") is not None
                or str(st).lower() in ("success", "ok", "true", "1", "200")
                or st is True or st == 1 or st == 200
                or any(k in msg.lower() for k in ("berhasil", "success", "returned", "done", "all numbers"))
            )
            if not ok:
                return False, msg or r.text[:200]
            detail = msg
            if count >= 0:
                detail = f"{msg} (total {count} nomor)"
            return True, detail or "Berhasil"
        except Exception:
            raw = r.text.lower()
            if any(k in raw for k in ("error", "gagal", "failed", "invalid", "unauthorized")):
                return False, r.text[:200]
            return True, r.text[:100]
    except Exception as e:
        return False, str(e)
        
def parse_range(rng):
    country = re.sub(r"\s*\(.*?\)", "", rng)
    country = re.sub(r"\d+", "", country)
    country = re.sub(r"\s+", " ", country).strip().upper()
    code_match = re.search(r"\((\d+)\)", rng)
    code = code_match.group(1) if code_match else ""
    return country, code

def extract_service_short(text):
    m = re.search(r"(WhatsApp|Telegram|Google|Facebook|Instagram|Shopee|Tokopedia|Grab|Gojek|TikTok)", text, re.I)
    if m: return SERVICE_SHORT.get(m.group(1).upper(), "#OT")
    return "#OT"

def mask_email(email):
    try:
        name, domain = email.split("@")
        if len(name) <= 2: return name + "*" + "@" + domain
        return f"{name[0]}{'*' * (len(name)-2)}{name[-1]}@{domain}"
    except:
        return email

def stats_sms(chat_id=None):
    total_sms = sms_stats["total_sms"]
    total_otp = sms_stats["total_otp"]
    total_number = len(sms_stats["total_number"])
    msg = f"  <b>STATISTIK SMS OTP</b>\n\n  Total SMS Masuk : {total_sms}\n  Total OTP       : {total_otp}\n  Total Nomor     : {total_number}\n  Total Akun Aktif: {len(accounts)}\n"
    if chat_id:
        send_msg(chat_id, msg)
    else:
        tg_active(msg)                        

def _csrf_from_html(html_text):
    """
    Ekstrak CSRF token dari HTML dengan 3 metode fallback.
    Sama seperti get_recv_csrf() supaya konsisten di seluruh kode.
    """
    soup = BeautifulSoup(html_text, "html.parser")
    # 1. <input type="hidden" name="_token" value="...">
    inp = soup.find("input", {"name": "_token"})
    if inp and inp.get("value"):
        return inp["value"]
    # 2. <meta name="csrf-token" content="...">
    meta = soup.find("meta", {"name": "csrf-token"})
    if meta and meta.get("content"):
        return meta["content"]
    # 3. JS inline: "_token":"xxxx" atau _token: "xxxx"
    import re as _re
    m = _re.search(r"""["']_token["']\s*[,:]?\s*["']([A-Za-z0-9_\-+/=]{20,})["']""", html_text)
    if m:
        return m.group(1)
    return None


def login(acc, _retry=0):
    session = acc["session"]
    email   = acc["email"]
    password = acc.get("password") or ""

    if not password:
        _log("LOGIN", f"Tidak ada password untuk [{email}]", Fore.YELLOW)
        return False

    # ── Jika sudah di /portal, session masih aktif — skip re-login ──
    # Kirim tanpa X-Requested-With (bukan AJAX) agar Laravel balik HTML biasa.
    # httpx tidak support None sebagai nilai header — build dict eksplisit.
    try:
        _portal_hdr = {k: v for k, v in dict(session.headers).items()
                       if k.lower() != "x-requested-with"}
        r_chk = session.get(f"{_IVAS_ORIGIN}/portal", timeout=15,
                            headers=_portal_hdr)
        if r_chk.status_code == 200 and "/login" not in str(r_chk.url):
            _log("LOGIN", f"Sudah login [{email}] — skip", Fore.GREEN)
            tok = _csrf_from_html(r_chk.text)
            if tok:
                acc["csrf_token"] = tok
            return True
    except Exception:
        pass

    worker_before = _IVAS_ORIGIN

    # ── GET halaman login — WAJIB tanpa X-Requested-With ──
    # Jika header XHR dikirim ke GET /login, Laravel menganggapnya AJAX
    # dan mengembalikan JSON / response berbeda tanpa form HTML + CSRF token.
    # httpx tidak support None sebagai nilai header — build dict eksplisit.
    try:
        _login_hdr = {k: v for k, v in dict(session.headers).items()
                      if k.lower() != "x-requested-with"}
        _login_hdr["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
        r = session.get(LOGIN_URL, timeout=20, headers=_login_hdr)
    except Exception as e:
        _log("LOGIN", f"GET /login error [{email}]: {e}", Fore.RED)
        return False

    if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        return login(acc, _retry=_retry + 1)

    # IVAS 429 — exponential backoff, bukan rotasi worker
    if r.status_code == 429:
        wait = min(20 * (2 ** _retry), 120)
        _log("LOGIN", f"IVAS 429 [{email}] — tunggu {wait}s lalu retry", Fore.YELLOW)
        time.sleep(wait)
        if _retry < 4:
            return login(acc, _retry=_retry + 1)
        return False

    # Sudah redirect ke /portal → session masih hidup
    if "/portal" in str(r.url):
        _log("LOGIN", f"Redirect ke portal [{email}] — session aktif", Fore.GREEN)
        tok = _csrf_from_html(r.text)
        if tok:
            acc["csrf_token"] = tok
        return True

    token = _csrf_from_html(r.text)
    if not token:
        _log("LOGIN", f"CSRF TIDAK DITEMUKAN [{email}] (status={r.status_code}, url={r.url})", Fore.RED)
        _log("LOGIN", f"HTML[:200]={r.text[:200]!r}", Fore.RED)
        if _retry < 2:
            time.sleep(10)
            return login(acc, _retry=_retry + 1)
        return False

    acc["csrf_token"] = token
    session.headers.update({
        "X-CSRF-TOKEN":     token,
        "X-Requested-With": "XMLHttpRequest",
    })

    # POST login
    try:
        r2 = session.post(LOGIN_URL, data={
            "_token":   token,
            "email":    email,
            "password": password,
        }, timeout=20)
    except Exception as e:
        _log("LOGIN", f"POST /login error [{email}]: {e}", Fore.RED)
        return False

    if is_worker_blocked(resp=r2) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        return login(acc, _retry=_retry + 1)

    if r2.status_code == 429:
        wait = min(20 * (2 ** _retry), 120)
        _log("LOGIN", f"IVAS 429 POST [{email}] — tunggu {wait}s lalu retry", Fore.YELLOW)
        time.sleep(wait)
        if _retry < 4:
            return login(acc, _retry=_retry + 1)
        return False

    _log("LOGIN", f"Response URL: {r2.url}", Fore.CYAN)

    if "/portal" in str(r2.url) or "Dashboard" in r2.text or "portal" in r2.text.lower():
        _log("LOGIN", f"BERHASIL [{email}]", Fore.GREEN)
        # ── FRESH COOKIE AUTO-CAPTURE ──
        fresh = extract_session_cookies(session)
        if fresh:
            acc["cookies"] = fresh
            acc["last_login"] = time.time()
            with _cookie_file_lock:
                all_cookies = load_cookies()
                all_cookies[email] = fresh
                save_cookies(all_cookies)
            _log("LOGIN", f"Fresh cookie saved: {email} ({len(fresh)} keys)", Fore.GREEN)
        return True
    else:
        _log("LOGIN", f"GAGAL [{email}] — password salah atau perlu verifikasi", Fore.RED)
        return False

def _is_login_page(r) -> bool:
    """Cek apakah response adalah redirect ke halaman login (session expired)."""
    try:
        if "/login" in str(r.url):
            return True
        if r.status_code in (401, 403, 419):
            return True
    except Exception:
        pass
    return False


def _invalidate_session(acc, reason="SESSION_EXPIRED"):
    """
    Force re-verify session pada iterasi berikutnya.
    Hapus ranges cache & recv_csrf cache agar semua di-fetch ulang setelah re-login.
    """
    email = acc.get("email", "")
    acc["last_login"] = 0
    if email:
        _ranges_cache.pop(email, None)
        _recv_csrf_cache.pop(email, None)
    raise Exception(reason)


# _RECV_POST_HEADERS sudah dibuat otomatis oleh _apply_worker_globals() di atas
# dan ikut ter-refresh setiap kali worker berpindah — jangan didefinisikan ulang statis di sini.


def _fetch_myrange_data(acc, _retry=0):
    """Ambil data /portal/numbers (dipakai /myrange) — ikut retry+rotasi worker
    kalau worker aktif kena block/limit/404/401, sama seperti get_ranges/get_numbers/get_sms."""
    my_url = f"{BASE}/portal/numbers"
    col_data = ["Number", "range", "A2P", "LimitA2P", "limit_did_a2p", "limit_cli_a2p", "number_id", "action"]
    col_name = ["Number", "range", "A2P",  "LimitA2P", "limit_did_a2p", "limit_cli_a2p", "number_id", "action"]
    col_qs = "".join(
        f"&columns[{i}][data]={d}&columns[{i}][name]={n}"
        for i, (d, n) in enumerate(zip(col_data, col_name))
    )
    qs = (
        f"draw=1{col_qs}"
        "&order[0][column]=0&order[0][dir]=asc"
        "&start=0&length=2000"
        "&search[value]=&search[regex]=false"
    )
    hdrs = {
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          my_url,
        "X-CSRF-TOKEN":     acc.get("csrf_token", ""),
    }
    session = acc["session"]
    worker_before = _IVAS_ORIGIN
    resp = session.get(f"{my_url}?{qs}", headers=hdrs, timeout=20)

    if (resp.status_code == 401 or resp.status_code == 419) and _retry < len(WORKER_POOL) - 1:
        # Bukan block worker — sesi/CSRF basi di worker ini. Coba re-login sebelum menyerah.
        acc["last_login"] = 0
        if ensure_login(acc):
            hdrs["X-CSRF-TOKEN"] = acc.get("csrf_token", "")
            resp = session.get(f"{my_url}?{qs}", headers=hdrs, timeout=20)

    if is_worker_blocked(resp=resp) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        acc["last_login"] = 0
        ensure_login(acc)
        return _fetch_myrange_data(acc, _retry=_retry + 1)

    if resp.status_code != 200:
        raise Exception(f"HTTP {resp.status_code}")
    return resp.json()


def get_ranges(acc, _retry=0):
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    worker_before = _IVAS_ORIGIN
    r = acc["session"].post(GET_RANGE_URL,
        data={"_token": csrf, "from": today, "to": today},
        headers=_RECV_POST_HEADERS
    )
    if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        return get_ranges(acc, _retry=_retry + 1)
    # IVAS 429 — tunggu lalu retry (bukan rotasi worker)
    if r.status_code == 429:
        wait = min(30 * (2 ** _retry), 180)
        email = acc.get("email", "")
        _log("MYRANGE", f"IVAS 429 [{email}] get_ranges — tunggu {wait}s", Fore.YELLOW)
        time.sleep(wait)
        if _retry < 3:
            return get_ranges(acc, _retry=_retry + 1)
        return []
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_ranges ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")
    ranges = []
    for div in soup.find_all("div", onclick=True):
        if "toggleRange" in div["onclick"]:
            try: ranges.append(div["onclick"].split("'")[1])
            except: pass
    return list(set(ranges))

def get_ranges_cached(acc):
    """Cache ranges 5 menit. Auto-invalidate saat session expired."""
    email = acc.get("email", "")
    now   = time.time()
    entry = _ranges_cache.get(email)
    if entry:
        ts, cached_ranges = entry
        if now - ts < RANGES_CACHE_TTL:
            return cached_ranges
    ranges = get_ranges(acc)
    _ranges_cache[email] = (now, ranges)
    return ranges

def get_numbers(acc, rng, _retry=0):
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    worker_before = _IVAS_ORIGIN
    r = acc["session"].post(GET_NUMBER_URL,
        data={"_token": csrf, "start": today, "end": today, "range": rng},
        headers=_RECV_POST_HEADERS
    )
    if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        return get_numbers(acc, rng, _retry=_retry + 1)
    # IVAS 429 — JANGAN block thread (ngeblok = OTP delay!)
    # Skip cycle ini saja, polling loop akan retry otomatis dalam 1-3 detik
    if r.status_code == 429:
        email = acc.get("email", "")
        _log("MYRANGE", f"IVAS 429 [{email}] get_numbers — skip cycle, retry otomatis", Fore.YELLOW)
        return []
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_numbers ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")
    numbers = []
    for div in soup.find_all("div", onclick=True):
        try:
            val = div["onclick"].split("'")[1]
            if val and val != rng: numbers.append(val)
        except: pass
    return list(set(numbers))

def get_sms(acc, rng, number, _retry=0):  
    today = datetime.now().strftime("%Y-%m-%d")
    csrf  = get_recv_csrf(acc)
    worker_before = _IVAS_ORIGIN
    r = acc["session"].post(GET_SMS_URL,
        data={"_token": csrf, "start": today, "end": today, "Number": number, "Range": rng},
        headers=_RECV_POST_HEADERS
    )
    if is_worker_blocked(resp=r) and _retry < len(WORKER_POOL) - 1:
        mark_worker_limited(worker_before)
        return get_sms(acc, rng, number, _retry=_retry + 1)
    # IVAS 429 — JANGAN block thread (ngeblok = OTP delay!)
    # Cukup return [] sekarang, polling loop akan retry dalam 1-3 detik sendiri
    if r.status_code == 429:
        email = acc.get("email", "")
        _log("MYRANGE", f"IVAS 429 [{email}] get_sms — skip cycle ini, retry otomatis", Fore.YELLOW)
        return []
    if _is_login_page(r):
        _invalidate_session(acc, f"SESSION_EXPIRED: get_sms ({r.url})")
    soup = BeautifulSoup(r.text, "html.parser")  
    sms_texts = []  
    try:  
        texts = list(soup.stripped_strings)  
        for t in texts:  
            t = t.strip()  
            if t.startswith("<#>"): t = t.replace("<#>", "").strip()  
            if re.fullmatch(r"[A-Za-z0-9]{10,}", t): continue  
            t_low = t.lower()  
            if any(x in t_low for x in ["sender", "revenue", "time"]): continue  
            if re.search(r"\b\d{2}:\d{2}:\d{2}\b", t): continue  
            if "$" in t: continue  
            if t and "No SMS Found" not in t: sms_texts.append(t)  
    except Exception as e: _log("ERROR", f"parse sms: {e}", Fore.RED)
    return list(dict.fromkeys(sms_texts))  
    
def format_phone_number(number):
    number = str(number).replace("+", "").replace(" ", "")
    if len(number) >= 10:
        return f"{number[:4]}****{number[-4:]}"
    return number    
    
def normalize_number(num, country_code):
    num = str(num).strip().replace(" ", "").replace("-", "").replace("+", "")
    if num.startswith(country_code): return num
    if num.startswith("0"): return country_code + num[1:]
    return num

def tg_active(msg):
    _tg_request("sendMessage", data={"chat_id": OWNER_ID, "text": msg, "parse_mode": "HTML"})
            
# ================= TELEGRAM LISTENER =================
_TG_POLL_CLIENT = httpx.Client(
    follow_redirects=True,
    timeout=35,
    headers={"User-Agent": "Mozilla/5.0"},
)

def listen_command():
    global last_update_id
    _backoff = 0
    while True:
        try:
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/getUpdates"
            r = _TG_POLL_CLIENT.get(url, params={"offset": last_update_id + 1, "timeout": 25})
            data = r.json()

            for upd in data.get("result", []):
                last_update_id = upd["update_id"]

                # ====== HANDLE CALLBACK QUERY (inline button click) ======
                if "callback_query" in upd:
                    try:
                        cq = upd["callback_query"]
                        cq_id = cq["id"]
                        cq_data = cq.get("data", "")
                        cq_user_id = cq["from"]["id"]
                        cq_chat_id = cq["message"]["chat"]["id"]
                        cq_msg_id = cq["message"]["message_id"]

                        if cq_data == "check_join":
                            not_joined = check_force_join(cq_user_id)
                            if not_joined:
                                answer_callback_query(cq_id, "⚠️ Kamu belum join semua channel/grup!")
                                send_force_join_msg(cq_chat_id, not_joined)
                            else:
                                answer_callback_query(cq_id, "✅ Sudah join semua! Silakan gunakan bot.")
                                delete_msg(cq_chat_id, cq_msg_id)
                                handle_start(cq_user_id, cq_chat_id)
                        elif cq_data.startswith("setcookie:"):
                            if is_owner(cq_user_id):
                                handle_setcookie_callback(cq_chat_id, cq_user_id, cq_data[len("setcookie:"):], cq_id, cq_msg_id)
                            else:
                                answer_callback_query(cq_id, "❌ Khusus OWNER")
                        elif cq_data.startswith("pkg_info:"):
                            handle_pkg_info_cb(cq_chat_id, cq_user_id, cq_data[9:], cq_id, cq_msg_id)
                        elif cq_data.startswith("pkg_buy:"):
                            threading.Thread(target=handle_pkg_buy_cb, args=(cq_chat_id, cq_user_id, cq_data[8:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data == "pkg_back":
                            answer_callback_query(cq_id)
                            delete_msg(cq_chat_id, cq_msg_id)
                            cmd_beli(cq_chat_id, cq_user_id)
                        elif cq_data.startswith("addcookie:"):
                            handle_addcookie_callback(cq_chat_id, cq_user_id, cq_data[len("addcookie:"):], cq_id, cq_msg_id)
                        elif cq_data.startswith("an:"):
                            handle_addnum_email_cb(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id)
                        elif cq_data.startswith("da:"):
                            threading.Thread(target=handle_delnumall_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("af:"):
                            threading.Thread(target=handle_ambilfile_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("mr:"):
                            threading.Thread(target=handle_myrange_email_cb, args=(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("de:"):
                            handle_delemail_select_cb(cq_chat_id, cq_user_id, cq_data[3:], cq_id, cq_msg_id)
                        elif cq_data.startswith("dec:"):
                            threading.Thread(target=handle_delemail_confirm_cb, args=(cq_chat_id, cq_user_id, cq_data[4:], cq_id, cq_msg_id), daemon=True).start()
                        elif cq_data.startswith("dok:"):
                            if is_owner(cq_user_id):
                                handle_delakun_select_cb(cq_chat_id, cq_user_id, cq_data[4:], cq_id, cq_msg_id)
                            else:
                                answer_callback_query(cq_id, "❌ Khusus OWNER")
                        elif cq_data.startswith("dokc:"):
                            if is_owner(cq_user_id):
                                threading.Thread(target=handle_delakun_confirm_cb, args=(cq_chat_id, cq_user_id, cq_data[5:], cq_id, cq_msg_id), daemon=True).start()
                            else:
                                answer_callback_query(cq_id, "❌ Khusus OWNER")
                        elif cq_data.startswith("cancel_payment:"):
                            parts = cq_data[len("cancel_payment:"):].split(":")
                            if len(parts) == 2:
                                cp_order_id, cp_amount_str = parts
                                pending = pending_payments.get(cq_user_id, {})
                                if pending.get("order_id") == cp_order_id:
                                    pending_payments.pop(cq_user_id, None)
                                    answer_callback_query(cq_id, "❌ Pembayaran dibatalkan")
                                    delete_msg(cq_chat_id, cq_msg_id)
                                    threading.Thread(target=pakasir_cancel, args=(cp_order_id, int(cp_amount_str)), daemon=True).start()
                                    send_msg(cq_chat_id,
                                        "❌ <b>Pembayaran dibatalkan.</b>\n\n"
                                        "Tagihan QRIS sudah dibatalkan.\n"
                                        "Ketik /beli untuk membuat tagihan baru."
                                    )
                                else:
                                    answer_callback_query(cq_id, "Tidak ada tagihan aktif")
                            else:
                                answer_callback_query(cq_id)
                        elif cq_data.startswith("cancel:"):
                            answer_callback_query(cq_id, "❌ Dibatalkan")
                            key = cq_data[7:]
                            if key == "sc":
                                pending_setcookie.pop(cq_user_id, None)
                            elif key == "ac":
                                pending_addcookie.pop(cq_user_id, None)
                            elif key == "an":
                                pending_addnum.pop(cq_user_id, None)
                            delete_msg(cq_chat_id, cq_msg_id)
                            send_msg(cq_chat_id, "❌ <b>Aksi dibatalkan.</b>")
                        elif cq_data == "cancel:dok":
                            answer_callback_query(cq_id, "❌ Dibatalkan")
                            delete_msg(cq_chat_id, cq_msg_id)
                            send_msg(cq_chat_id, "❌ <b>Aksi dibatalkan.</b>")
                        elif cq_data.startswith("cekv2:"):
                            try:
                                secs = int(cq_data[6:])
                            except ValueError:
                                answer_callback_query(cq_id, "❌ Data tidak valid")
                                secs = None
                            if secs:
                                threading.Thread(
                                    target=cek_ivas_v2_exec,
                                    args=(cq_chat_id, cq_user_id, secs, cq_id, cq_msg_id),
                                    daemon=True,
                                ).start()
                        else:
                            answer_callback_query(cq_id)
                    except Exception as ex:
                        print(f"Error callback_query: {ex}")
                    continue

                if "message" not in upd: continue
                try:
                    msg = upd["message"]
                    text = msg.get("text", "") or ""
                    user_id = msg["from"]["id"]
                    chat_id = msg["chat"]["id"]
                    msg_id = msg["message_id"]

                    store_username(user_id, msg["from"])

                    owner = is_owner(user_id)
                    is_group = msg["chat"]["type"] in ["group", "supergroup"]

                    # ====== CEK WAJIB JOIN (hanya user non-owner di private chat) ======
                    if not owner and not is_group and text.startswith("/") and text != "/start":
                        not_joined = check_force_join(user_id)
                        if not_joined:
                            send_force_join_msg(chat_id, not_joined)
                            continue

                    # ====== CEK PENDING SETCOOKIE (owner input cookie JSON) ======
                    # Hanya proses di private chat — grup tidak boleh trigger pending state
                    if not is_group and owner and user_id in pending_setcookie and text and not text.startswith("/"):
                        if process_cookie_input(chat_id, user_id, text):
                            continue

                    # ====== CEK PENDING ADDCOOKIE (semua user bisa, input cookie JSON) ======
                    if not is_group and user_id in pending_addcookie and text and not text.startswith("/"):
                        if process_addcookie_input(chat_id, user_id, text):
                            continue

                    # ====== CEK PENDING ADDNUM (semua user bisa, asal sedang dalam pending state) ======
                    if not is_group and user_id in pending_addnum and text and not text.startswith("/"):
                        if process_addnum_target(chat_id, user_id, text):
                            continue

                    # ROUTING COMMAND TEXT
                    udisp = get_user_display(user_id)

                    # ====== GRUP: HANYA COMMAND TERTENTU YANG DIIZINKAN ======
                    _GROUP_ALLOWED = (
                        "/cekivas", "/cekrange", "/cekivasv2", "/toprcv", "/rangeterbaru",
                        "/addgrup", "/delgrup", "/listgrup", "/start",
                    )
                    if is_group and text.startswith("/"):
                        if not any(text.startswith(c) for c in _GROUP_ALLOWED):
                            send_msg(chat_id,
                                "❌ <b>Command ini tidak bisa dipakai di grup!</b>\n\n"
                                "<blockquote>Command yang tersedia di grup:\n"
                                "/cekivas — traffic WhatsApp\n"
                                "/cekivasv2 — filter waktu\n"
                                "/toprcv [negara] — range per negara\n"
                                "/rangeterbaru — range aktif\n"
                                "/addgrup — daftarkan grup ini\n"
                                "/delgrup — hapus grup ini\n"
                                "/listgrup — lihat daftar grup\n\n"
                                "💡 Semua command lain hanya di private chat.</blockquote>"
                            )
                            continue

                    if text == "/start":
                        if not owner:
                            not_joined = check_force_join(user_id)
                            if not_joined:
                                send_force_join_msg(chat_id, not_joined)
                                continue
                        threading.Thread(target=handle_start, args=(user_id, chat_id), daemon=True).start()
                    elif text.startswith("/cekivasv2"):
                        threading.Thread(target=cek_ivas_v2, args=(chat_id, user_id), daemon=True).start()
                        send_activity_log(user_id, udisp, "/cekivasv2")
                    elif text.startswith("/cekivas") or text.startswith("/cekrange"):
                        cmd_name = "/cekivas" if text.startswith("/cekivas") else "/cekrange"
                        threading.Thread(target=cek_traffic_range, args=(chat_id, user_id), daemon=True).start()
                        send_activity_log(user_id, udisp, cmd_name)
                    elif text.startswith("/toprcv"):
                        q = text[len("/toprcv"):].strip()
                        threading.Thread(target=top_rcv, args=(chat_id, user_id, q), daemon=True).start()
                        send_activity_log(user_id, udisp, "/toprcv")
                    elif text.startswith("/rangeterbaru"):
                        threading.Thread(target=range_terbaru, args=(chat_id, user_id), daemon=True).start()
                        send_activity_log(user_id, udisp, "/rangeterbaru")
                    elif text.startswith("/cekprem"): cek_premium(chat_id, user_id)
                    
                    elif text.startswith("/listakun"): 
                        if owner: list_accounts(chat_id, user_id)
                        else: send_msg(chat_id, "  Khusus OWNER")
                    elif text.startswith("/addcookie"):
                        if use_token(user_id):
                            add_cookie_premium(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/addcookie")
                        else: no_token_msg(chat_id)
                    elif text.startswith("/delcookie"):
                        if use_token(user_id):
                            del_cookie_premium(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/delcookie")
                        else: no_token_msg(chat_id)
                    
                    elif text.startswith("/addemail"):
                        if use_token(user_id):
                            add_email(text, chat_id, user_id, msg_id)
                            send_activity_log(user_id, udisp, "/addemail")
                        else: no_token_msg(chat_id)
                    elif text.startswith("/listemail"): list_email(chat_id, user_id)
                    elif text.startswith("/delemail"):
                        if use_token(user_id):
                            command_delemail(chat_id, user_id)
                            send_activity_log(user_id, udisp, "/delemail")
                        else: no_token_msg(chat_id)
                    
                    elif text.startswith("/addgrup"):
                        if is_group:
                            gid = str(chat_id)
                            if gid in get_user_groups(user_id):
                                send_msg(chat_id,
                                    "⚠️ <b>Grup sudah terdaftar!</b>\n\n"
                                    f"<blockquote>🆔 ID: <code>{gid}</code>\n"
                                    f"Grup ini sudah ada di akun kamu.</blockquote>")
                                send_activity_log(user_id, udisp, "/addgrup", "⚠️ Grup sudah ada")
                            else:
                                add_user_group(user_id, gid)
                                send_msg(chat_id,
                                    f"✅ <b>Grup berhasil ditambahkan!</b>\n\n"
                                    f"<blockquote>🆔 ID: <code>{gid}</code>\n"
                                    f"🔑 Key: <code>{get_or_create_user_key(user_id)}</code>\n\n"
                                    f"Notif OTP akan dikirim ke grup ini.</blockquote>")
                                send_activity_log(user_id, udisp, "/addgrup")
                        else:
                            send_msg(chat_id,
                                "❌ <b>Harus dijalankan di dalam grup!</b>\n\n"
                                "<blockquote>Tambahkan bot ke grup kamu dulu,\n"
                                "lalu ketik /addgrup di dalam grup tersebut.</blockquote>")

                    elif text.startswith("/delgrup"):
                        if is_group:
                            gid = str(chat_id)
                            if remove_user_group(user_id, gid):
                                send_msg(chat_id,
                                    f"✅ <b>Grup berhasil dihapus!</b>\n\n"
                                    f"<blockquote>🆔 ID: <code>{gid}</code>\n"
                                    f"Grup ini tidak akan menerima notif OTP lagi.</blockquote>")
                                send_activity_log(user_id, udisp, "/delgrup")
                            else:
                                send_msg(chat_id,
                                    "❌ <b>Grup tidak ditemukan!</b>\n\n"
                                    "<blockquote>Grup ini belum terdaftar di akun kamu.\n"
                                    "Ketik /listgrup untuk lihat daftar grup.</blockquote>")
                        else:
                            send_msg(chat_id,
                                "❌ <b>Harus dijalankan di dalam grup!</b>\n\n"
                                "<blockquote>Ketik /delgrup langsung di grup yang ingin dihapus.</blockquote>")

                    elif text.startswith("/listgrup"):
                        my_groups = get_user_groups(user_id)
                        if not my_groups:
                            send_msg(chat_id,
                                "📋 <b>LIST GRUP KAMU</b>\n\n"
                                "<blockquote>Belum ada grup terdaftar.\n"
                                "Tambah bot ke grup, lalu ketik /addgrup di sana.</blockquote>")
                        else:
                            msg_out = "📋 <b>LIST GRUP KAMU</b>\n\n<blockquote>"
                            for i, g in enumerate(my_groups, 1):
                                msg_out += f"{i}. <code>{g}</code>\n"
                            msg_out += f"\nTotal: {len(my_groups)} grup</blockquote>"
                            send_msg(chat_id, msg_out)

                    elif text.startswith("/addnum"):
                        if use_token(user_id):
                            command_addnum(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/addnum")
                        else: no_token_msg(chat_id)
                        
                    elif text.startswith("/ambilfile"):
                        if use_token(user_id):
                            command_ambilfile(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/ambilfile")
                        else: no_token_msg(chat_id)

                    elif text.startswith("/delnumall"):
                        if use_token(user_id):
                            command_delnumall(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/delnumall")
                        else: no_token_msg(chat_id)

                    elif text.startswith("/testapi"):
                        if owner:
                            threading.Thread(target=command_testapi, args=(chat_id, user_id, text), daemon=True).start()
                        else:
                            send_msg(chat_id, "❌ Khusus OWNER")

                    elif text.startswith("/myrange"):
                        if use_token(user_id):
                            command_myrange(text, chat_id, user_id)
                            send_activity_log(user_id, udisp, "/myrange")
                        else: no_token_msg(chat_id)
                    
                    elif text.startswith("/beli"):
                        cmd_beli(chat_id, user_id)
                        send_activity_log(user_id, udisp, "/beli", "📦 Buka menu pembelian")
                    elif text.startswith("/addtoken"): 
                        if owner: add_token_tier(text, chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/deltoken"): 
                        if owner: del_token_tier(text, chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/resettoken"):
                        if owner: cmd_resettoken(text, chat_id)
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/listtoken"): 
                        if owner: list_token_tier(chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/addakun"): 
                        if owner: add_account(text, chat_id, user_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/delakun"): 
                        if owner: command_delakun(chat_id, user_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/setcookie"): 
                        if owner: cmd_setcookie(chat_id)
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                    elif text.startswith("/statsms"): 
                        if owner: stats_sms(chat_id) 
                        else: send_msg(chat_id, "❌ Khusus OWNER")
                except Exception as ex: 
                    print(f"Error handling message: {ex}")
        except Exception as e:
            _backoff = min(_backoff + 2, 20)
            print(f"Loop listener error: {e} — retry in {_backoff}s")
            time.sleep(_backoff)

            
# ================= POLL ENGINE (per-account) =================

def poll_one_account(acc):
    """
    Satu iterasi polling SMS untuk satu akun.
    Return True jika ada SMS/OTP baru ditemukan, False jika tidak.
    """
    email = acc.get("email", "")
    if not email:
        return False

    found_sms = False  # flag — diset True jika ada OTP dikirim

    # Cek session retry
    if _session_notified.get(email):
        if time.time() - _session_retry_time.get(email, 0) < SESSION_RETRY_INTERVAL:
            return False
        print(f"  AUTO-RETRY SESSION: {email}")
        acc["last_login"] = 0

    if not ensure_login(acc):
        return False

    owner_uid  = _bot_state["email_to_uid"].get(email, OWNER_ID)
    total      = _bot_state["total_accounts"]

    # ── Tentukan tujuan kirim: hanya ke pemilik akun (user/grup mereka sendiri) ──
    acct_groups  = get_user_groups(owner_uid)
    send_targets = acct_groups if acct_groups else [str(owner_uid)]

    try:
        ranges = get_ranges_cached(acc)
    except Exception as e:
        err = str(e)
        if "SESSION_EXPIRED" not in err:
            _log("WARN", f"get_ranges [{email}]: {err}", Fore.YELLOW)
        return False

    def _process_number(rng, num, fallback_country, code):
        """Ambil SMS untuk satu nomor + langsung kirim ke Telegram kalau ada OTP baru.
        Dijalankan paralel per-nomor supaya banyak OTP TIDAK bikin delay."""
        local_found = False
        full_num = normalize_number(num, code)
        if not full_num.isdigit():
            return False

        try:
            sms_list = get_sms(acc, rng, num)
        except Exception as e:
            if "SESSION_EXPIRED" not in str(e):
                _log("WARN", f"get_sms [{email}]: {e}", Fore.YELLOW)
            return False

        for sms in sms_list:
            clean_sms = re.sub(r"\s+", " ", sms.replace("<#>", "")).strip()
            sms_uid = hashlib.md5(f"{num}-{clean_sms}".encode()).hexdigest()

            with _sent_cache_lock:
                if sms_uid in sent_cache:
                    continue

            matches = re.findall(r"\b\d{3}[- ]?\d{3}\b", sms)
            if not matches:
                continue

            otp = matches[0].replace(" ", "-")
            service_name = extract_service_short(sms)
            country, flag = detect_country_and_flag(full_num, fallback_country)

            # Ekstrak region code (#UA, #ID, dll.) dan dial code (+380, +62, dll.)
            try:
                _parsed    = phonenumbers.parse("+" + full_num, None)
                region_code = phonenumbers.region_code_for_number(_parsed) or "??"
                dial_code   = str(_parsed.country_code)
                last4       = full_num[-4:] if len(full_num) >= 4 else full_num
            except Exception:
                region_code = fallback_country[:2] if fallback_country else "??"
                dial_code   = code.lstrip("+") if code else ""
                last4       = full_num[-4:] if len(full_num) >= 4 else full_num

            # Format Garage OTP: FLAG #CODE 💬+DIALCODE 🔢 LAST4 #SERVICE
            masked_num = format_phone_number(full_num)
            # Pilih emoji sesuai service
            _svc_emoji_map = {
                "#WS": "💬", "#TG": "✈️", "#G": "🔍", "#FB": "👥",
                "#IG": "📸", "#TT": "🎵", "#GR": "🚗", "#GJ": "🛵",
                "#SP": "🛒", "#TP": "🛍️",
            }
            _svc_icon = _svc_emoji_map.get(service_name, "💬")
            msg = (
                f"{flag} <b>#{region_code}</b>  {_svc_icon} <code>+{dial_code}</code>  🔢 <code>{last4}</code>  <b>{service_name}</b>"
            )

            # Tombol Garage-style: GET OTP (copy) + NUMBER + CHANNEL
            _GROUP_LINK = "https://t.me/matttttcha"

            for gid in send_targets:
                try:
                    _kb = {"inline_keyboard": [
                        [{"text": f"⚡ {otp} ⚡", "copy_text": {"text": otp}}],
                        [
                            {"text": "📱 NUMBER ↗", "url": _GROUP_LINK},
                            {"text": "📢 CHANNEL ↗", "url": _GROUP_LINK},
                        ],
                    ]}
                    import requests as _req
                    _r = _req.post(
                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                        json={"chat_id": gid, "text": msg, "parse_mode": "HTML",
                              "reply_markup": _kb},
                        timeout=10,
                    )
                    if not _r.json().get("ok"):
                        _log("SEND-ERR", f"→ {gid}: {_r.json().get('description','?')}", Fore.RED)
                except Exception as _se:
                    _log("SEND-ERR", f"→ {gid}: {_se}", Fore.RED)

            with _sent_cache_lock:
                sent_cache.add(sms_uid)
            save_sent_cache_debounced()

            sms_stats["total_sms"] += 1
            sms_stats["total_otp"] += 1
            if len(sms_stats["total_number"]) < 10000:
                sms_stats["total_number"].add(full_num)

            user_display = get_user_display(owner_uid)
            _log("OTP", f"{user_display} ({mask_email(email)}) | {masked_num} | {otp}", Fore.GREEN)
            local_found = True

        return local_found

    for rng in ranges:
        fallback_country, code = parse_range(rng)
        try:
            numbers = get_numbers(acc, rng)
        except Exception as e:
            err = str(e)
            if "SESSION_EXPIRED" not in err:
                _log("WARN", f"get_numbers [{email}]: {err}", Fore.YELLOW)
            continue

        if not numbers:
            continue

        # Fetch SMS untuk SEMUA nomor SECARA PARALEL — bukan satu-satu.
        # Ini kunci fix delay: makin banyak nomor/OTP di akun, makin banyak
        # request jalan bersamaan (bukan berantai), jadi tidak ada tambahan delay.
        max_workers = min(30, len(numbers))
        with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix=f"sms-{email[:10]}") as pool:
            futures = {
                pool.submit(_process_number, rng, num, fallback_country, code): num
                for num in numbers
            }
            for fut in as_completed(futures):
                try:
                    if fut.result():
                        found_sms = True
                except Exception as e:
                    _log("WARN", f"process_number [{email}]: {e}", Fore.YELLOW)

    return found_sms


# ================= AUTO COOKIE REFRESHER =================
def _notify_cookie_expired(email, uid):
    """Kirim notif ke pemilik akun bahwa cookie expired — max 1x per COOKIE_NOTIF_COOLDOWN."""
    now = time.time()
    if now - _last_cookie_notif.get(email, 0) < COOKIE_NOTIF_COOLDOWN:
        return
    _last_cookie_notif[email] = now

    msg = (
        f"⚠️ <b>COOKIE EXPIRED — AUTO REFRESH GAGAL</b>\n\n"
        f"📧 Email: <code>{email}</code>\n"
        f"❌ Cookie sudah expired dan tidak bisa auto-login.\n\n"
        f"<blockquote>Silakan perbarui cookie dengan:\n"
        f"• Owner  : /setcookie\n"
        f"• User   : /addcookie\n\n"
        f"💡 Ambil cookie fresh dari browser:\n"
        f"DevTools → Application → Cookies → copy semua</blockquote>"
    )
    # Kirim notif HANYA ke pemilik akun (bukan bocor ke owner kalau akun milik user)
    target = uid if uid else OWNER_ID
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": target, "text": msg, "parse_mode": "HTML"},
            timeout=10
        )
    except Exception as e:
        _log("NOTIF-ERR", f"[{email}]: {e}", Fore.RED)


def auto_cookie_refresher():
    """
    Background keepalive — ping server tiap 10 menit per akun.

    Cara kerja (tanpa password):
    1. Hit endpoint portal dengan session + cookie yang ada
    2. Server Laravel otomatis extend session lifetime
    3. Extract fresh cookies dari response → simpan ke file
    4. Session tidak pernah sempat expired selama bot hidup

    Jika session sudah benar-benar expired (bot restart lama / downtime):
    → Notif owner + user untuk update cookie manual, TIDAK coba password login
    """
    _log("KEEPALIVE", "background aktif — ping tiap 10 menit", Fore.CYAN)
    time.sleep(90)  # Tunggu bot fully ready + semua akun ter-load dulu

    while True:
        try:
            now = time.time()

            # Kumpulkan semua akun (owner + premium)
            with accounts_lock:
                owner_accs = list(accounts)
            prem_accs = list(_premium_acc_cache.values())
            all_accs = owner_accs + prem_accs

            # Filter hanya akun yang sudah waktunya di-ping
            due = [
                a for a in all_accs
                if a.get("email") and
                   a.get("cookies") and
                   now - _last_cookie_refresh.get(a["email"], 0) >= COOKIE_KEEPALIVE_INTERVAL
            ]

            if due:
                pass  # Silent ping — log hanya muncul jika ada error/warning

            for acc in due:
                email = acc["email"]
                try:
                    # FIX RACE CONDITION: Gunakan sesi TERPISAH (throwaway) untuk keepalive
                    # agar TIDAK mengganggu acc["session"] yang sedang dipakai account_worker thread.
                    # httpx.Client TIDAK thread-safe untuk concurrent use dari 2 thread sekaligus.
                    stored = acc.get("cookies", {})
                    ka_session = make_httpx_client(timeout=15)
                    if stored:
                        ka_session.cookies.update(stored)

                    # Ping ke portal — server Laravel extend session pada setiap request
                    r = ka_session.get(f"{BASE}/portal", timeout=15)

                    if r.status_code == 200 and "/login" not in str(r.url):
                        # Session masih hidup — ambil fresh cookies dari throwaway session
                        soup = BeautifulSoup(r.text, "html.parser")
                        t = soup.find("input", {"name": "_token"})
                        if t:
                            acc["csrf_token"] = t["value"]

                        fresh = extract_session_cookies(ka_session)
                        if fresh:
                            # Update stored cookies (acc["session"] akan pakai ini saat ensure_login berikutnya)
                            acc["cookies"] = fresh
                            acc["last_login"] = now
                            save_fresh_cookies_auto(email, fresh)
                            # Invalidate recv_csrf cache — akan di-refresh saat poll berikutnya
                            _recv_csrf_cache.pop(email, None)
                            # Reset flag session gagal jika sebelumnya sempat error
                            if _session_notified.get(email):
                                _session_notified[email] = False
                                _session_fail_time.pop(email, None)
                                _session_retry_time.pop(email, None)
                                if not _session_recovered.get(email):
                                    _session_recovered[email] = True
                                    _uid_recover = _bot_state.get("email_to_uid", {}).get(email, OWNER_ID)
                                    recover_target = _uid_recover if _uid_recover else OWNER_ID
                                    try:
                                        requests.post(
                                            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                            data={
                                                "chat_id": recover_target,
                                                "text": (
                                                    f"✅ <b>SESSION PULIH</b>\n\n"
                                                    f"📧 Email: <code>{email}</code>\n"
                                                    f"Session aktif kembali secara otomatis."
                                                ),
                                                "parse_mode": "HTML"
                                            },
                                            timeout=10
                                        )
                                    except Exception:
                                        pass
                            _keepalive_warn_count[email] = 0
                            _log("KA-OK", f"{email} — {len(fresh)} cookie di-extend", Fore.GREEN)
                    else:
                        # Hitung berapa kali keepalive gagal berturut-turut
                        fail_n = _keepalive_warn_count.get(email, 0) + 1
                        _keepalive_warn_count[email] = fail_n
                        uid = _bot_state.get("email_to_uid", {}).get(email, OWNER_ID)

                        if fail_n == 1:
                            # Kegagalan PERTAMA → kirim warning awal (sebelum konfirmasi expired)
                            _log("KA-WARN", f"({fail_n}x) {email}", Fore.YELLOW)
                            last_notif = _last_cookie_notif.get(email + "_warn", 0)
                            if now - last_notif > COOKIE_NOTIF_COOLDOWN:
                                _last_cookie_notif[email + "_warn"] = now
                                warn_msg = (
                                    f"⚠️ <b>SESSION WARNING</b>\n\n"
                                    f"📧 Email: <code>{email}</code>\n"
                                    f"Session tidak merespons. Kemungkinan cookie akan segera expired.\n\n"
                                    f"<blockquote>Bot sedang otomatis retry...\n"
                                    f"Jika berlanjut, segera perbarui cookie dengan:\n"
                                    f"• Owner: /setcookie\n"
                                    f"• User: /addcookie</blockquote>"
                                )
                                # Kirim HANYA ke pemilik akun (tidak bocor ke owner jika akun milik user)
                                warn_target = uid if uid else OWNER_ID
                                try:
                                    requests.post(
                                        f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                        data={"chat_id": warn_target, "text": warn_msg, "parse_mode": "HTML"},
                                        timeout=10
                                    )
                                except Exception:
                                    pass
                        else:
                            # Kegagalan ke-2+ → session benar-benar expired
                            _log("KA-DEAD", f"({fail_n}x) {email} — session mati", Fore.RED)

                            # ── AUTO RE-LOGIN jika ada password ──
                            pw = acc.get("password", "")
                            if pw and not _session_notified.get(email):
                                _log("AUTO-LOGIN", f"Coba re-login [{email}]...", Fore.CYAN)
                                try:
                                    acc["session"] = make_httpx_client(timeout=20)
                                    acc["last_login"] = 0
                                    if login(acc):
                                        fresh = extract_session_cookies(acc["session"])
                                        if fresh:
                                            acc["cookies"] = fresh
                                            acc["last_login"] = now
                                            save_fresh_cookies_auto(email, fresh)
                                        _keepalive_warn_count[email] = 0
                                        _session_notified.pop(email, None)
                                        _session_fail_time.pop(email, None)
                                        _session_retry_time.pop(email, None)
                                        _session_recovered.pop(email, None)
                                        _recv_csrf_cache.pop(email, None)
                                        _log("AUTO-LOGIN", f"BERHASIL [{email}]", Fore.GREEN)
                                        notif_target = uid if uid else OWNER_ID
                                        try:
                                            requests.post(
                                                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                                data={"chat_id": notif_target,
                                                      "text": (f"✅ <b>AUTO RE-LOGIN BERHASIL</b>\n\n"
                                                               f"📧 Email: <code>{email}</code>\n"
                                                               f"🔑 Session diperbarui otomatis oleh bot"),
                                                      "parse_mode": "HTML"},
                                                timeout=10
                                            )
                                        except Exception:
                                            pass
                                        _last_cookie_refresh[email] = now
                                        continue  # skip notif expired — sudah pulih
                                    else:
                                        _log("AUTO-LOGIN", f"Gagal [{email}]", Fore.RED)
                                except Exception as _re_err:
                                    _log("AUTO-LOGIN", f"Error [{email}]: {_re_err}", Fore.RED)

                            # Tidak ada password / auto-login gagal → notif user
                            _notify_cookie_expired(email, uid)
                            if not _session_notified.get(email):
                                _session_notified[email] = True
                                _session_recovered[email] = False
                                _session_fail_time[email] = now

                    _last_cookie_refresh[email] = now

                except Exception as e:
                    _log("KA-ERR", f"[{email}]: {e}", Fore.RED)

                time.sleep(2)  # Jeda kecil antar akun — jangan hammering server

            time.sleep(60)  # Loop setiap 1 menit untuk cek akun mana yang due

        except Exception as e:
            _log("KA-ERR", f"loop crash: {e}", Fore.RED)
            time.sleep(60)


# ================= AUTO BACKUP =================
def _collect_backup_files():
    """
    Kumpulkan semua file project secara rekursif, skip folder/file sistem.
    Return list of (absolute_path, archive_name).
    """
    root = os.path.abspath(".")
    collected = []
    for dirpath, dirnames, filenames in os.walk(root):
        # Pruning — hapus dir yang harus diskip dari traversal
        dirnames[:] = [
            d for d in dirnames
            if d not in BACKUP_SKIP_DIRS and not d.startswith(".")
            or d in {"file", "voice"}          # folder project yang harus masuk
        ]
        # Tambahan: pastikan folder tersembunyi non-project tetap diskip
        dirnames[:] = [
            d for d in dirnames
            if not (d.startswith(".") and d not in {"file", "voice"})
            and d not in BACKUP_SKIP_DIRS
        ]

        for fname in filenames:
            # Skip berdasarkan ekstensi
            _, ext = os.path.splitext(fname)
            if ext.lower() in BACKUP_SKIP_EXTS:
                continue
            # Skip file tertentu
            if fname in BACKUP_SKIP_FILES:
                continue
            # Skip file .zip di root (hasil backup lama)
            abs_path = os.path.join(dirpath, fname)
            rel_path = os.path.relpath(abs_path, root)
            collected.append((abs_path, rel_path))

    return collected


def _send_backup_telegram():
    """Scan seluruh project, buat ZIP, dan kirim ke owner Telegram."""
    now_str  = datetime.now().strftime("%d-%m-%Y_%H%M")
    zip_name = f"BACKUP_SPIDERMAT_BOT_{now_str}.zip"
    # Simpan zip di luar root agar tidak ikut ter-scan
    zip_path = f"/tmp/{zip_name}"
    try:
        files_to_backup = _collect_backup_files()
        total_files = len(files_to_backup)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for abs_path, arc_name in files_to_backup:
                try:
                    zf.write(abs_path, arc_name)
                except Exception:
                    pass  # skip file yang tidak bisa dibaca

        size_kb = round(os.path.getsize(zip_path) / 1024, 1)
        size_mb = round(size_kb / 1024, 2)
        size_str = f"{size_mb} MB" if size_kb > 1024 else f"{size_kb} KB"

        with open(zip_path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument",
                data={
                    "chat_id": OWNER_ID,
                    "caption": (
                        f"📦 <b>AUTO BACKUP — {datetime.now().strftime('%d %b %Y %H:%M')}</b>\n\n"
                        f"🗂️ File: <code>{zip_name}</code>\n"
                        f"📁 Total: <b>{total_files} file</b> (termasuk semua folder)\n"
                        f"📏 Ukuran: <b>{size_str}</b>"
                    ),
                    "parse_mode": "HTML",
                },
                files={"document": (zip_name, f, "application/zip")},
                timeout=120,
            )
        _log("BACKUP", f"terkirim — {zip_name} ({size_str}, {total_files} file)", Fore.GREEN)
    except Exception as e:
        _log("BACKUP", f"error: {e}", Fore.RED)
        try:
            requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                data={
                    "chat_id": OWNER_ID,
                    "text": f"❌ <b>Auto Backup Gagal</b>\n\n<code>{e}</code>",
                    "parse_mode": "HTML",
                },
                timeout=10,
            )
        except Exception:
            pass
    finally:
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
        except Exception:
            pass


def run_auto_backup():
    """Background thread: backup langsung saat startup, lalu tiap jam 00:00."""
    _log("BACKUP", "background aktif — startup + tiap 00:00", Fore.CYAN)
    time.sleep(20)  # Tunggu bot & akun selesai init

    # ── Backup pertama: langsung saat startup ─────────────────────────────────
    _log("BACKUP", "kirim backup startup...", Fore.CYAN)
    _send_backup_telegram()

    # ── Loop: backup berikutnya setiap tengah malam ───────────────────────────
    while True:
        try:
            now = datetime.now()
            next_midnight = (now + timedelta(days=1)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            wait_sec = (next_midnight - now).total_seconds()

            jam   = int(wait_sec // 3600)
            menit = int((wait_sec % 3600) // 60)
            _log("BACKUP", f"berikutnya dalam {jam}j {menit}m ({next_midnight.strftime('%d %b %Y 00:00')})", Fore.CYAN)

            time.sleep(wait_sec)

            _log("BACKUP", "tengah malam — dimulai...", Fore.CYAN)
            _send_backup_telegram()

            time.sleep(65)  # Jeda agar tidak trigger 2x di menit yang sama

        except Exception as e:
            _log("BACKUP", f"loop crash: {e}", Fore.RED)
            time.sleep(3600)


# ================= EXPIRY NOTIFIER (paket premium + cookie) =================
def _send_expiry_notif(uid, tier, sisa_detik, level):
    """Kirim notifikasi paket akan expired ke user dan owner."""
    t_info     = TOKEN_TIERS.get(tier, {})
    tier_label = t_info.get("label", tier.upper())
    tier_emoji = t_info.get("emoji", "🏷️")

    jam   = int(sisa_detik // 3600)
    menit = int((sisa_detik % 3600) // 60)
    sisa_str = f"{jam} jam {menit} menit" if jam > 0 else f"{menit} menit"

    if level == "24h":
        icon, judul = "⚠️", "PAKET AKAN EXPIRED — 24 JAM LAGI"
    elif level == "3h":
        icon, judul = "🚨", "PAKET AKAN EXPIRED — 3 JAM LAGI"
    else:
        icon, judul = "🔴", "PAKET AKAN EXPIRED — 1 JAM LAGI"

    msg = (
        f"{icon} <b>{judul}</b>\n\n"
        f"🏷️ Paket: {tier_emoji} <b>{tier_label}</b>\n"
        f"⏳ Sisa waktu: <b>{sisa_str}</b>\n\n"
        f"<blockquote>Segera perpanjang agar monitoring tidak terhenti!\n"
        f"Ketik /beli untuk pilihan paket.</blockquote>"
    )
    msg_owner = (
        f"{icon} <b>INFO PAKET USER</b>\n\n"
        f"👤 User ID: <code>{uid}</code>\n"
        f"🏷️ Paket: {tier_emoji} <b>{tier_label}</b>\n"
        f"⏳ Expired dalam: <b>{sisa_str}</b>"
    )
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": uid, "text": msg, "parse_mode": "HTML"},
            timeout=10,
        )
    except Exception:
        pass
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
            data={"chat_id": OWNER_ID, "text": msg_owner, "parse_mode": "HTML"},
            timeout=10,
        )
    except Exception:
        pass


def run_token_reset():
    """Background thread: reset token semua user tepat jam 00:00 WIB setiap hari."""
    _log("TOKEN-RST", "background aktif — reset tiap 00:00 WIB", Fore.CYAN)
    time.sleep(30)

    while True:
        try:
            tz_wib = timezone(timedelta(hours=7))
            now_wib = datetime.now(tz_wib)
            next_midnight = (now_wib + timedelta(days=1)).replace(
                hour=0, minute=0, second=0, microsecond=0)
            wait_sec = (next_midnight - now_wib).total_seconds()

            jam   = int(wait_sec // 3600)
            menit = int((wait_sec % 3600) // 60)
            _log("TOKEN-RST", f"reset berikutnya dalam {jam}j {menit}m ({next_midnight.strftime('%d %b %Y 00:00 WIB')})", Fore.CYAN)

            time.sleep(wait_sec)

            # ── Jalankan reset semua user ──────────────────────────────────
            _log("TOKEN-RST", "00:00 WIB — reset token semua user...", Fore.CYAN)
            today = get_wib_date()
            users_d   = load_users()
            prem_d    = load_premium()
            now_ts    = time.time()
            reset_count = 0
            for uid_str, udata in users_d.items():
                try:
                    uid_int   = int(uid_str)
                    if uid_int == OWNER_ID:
                        continue
                    prem      = prem_d.get(uid_str, {})
                    expired   = prem.get("expired", 0)
                    tier      = prem.get("tier", "free") if expired > now_ts else "free"
                    new_limit = TOKEN_TIERS.get(tier, TOKEN_TIERS["free"])["tokens_day"]
                    udata["tokens"]           = new_limit
                    udata["last_token_reset"] = today
                    reset_count += 1
                except Exception:
                    continue
            save_users(users_d)
            _log("TOKEN-RST", f"selesai — {reset_count} user direset", Fore.GREEN)

            # Notif ke owner
            try:
                send_msg(OWNER_ID,
                    f"🔄 <b>AUTO RESET TOKEN — 00:00 WIB</b>\n\n"
                    f"<blockquote>✅ {reset_count} user token direset ke limit harian masing-masing.</blockquote>")
            except Exception:
                pass

            time.sleep(65)  # jeda agar tidak trigger 2x di menit yang sama

        except Exception as e:
            _log("TOKEN-RST", f"loop crash: {e}", Fore.RED)
            time.sleep(3600)

def run_expiry_notifier():
    """
    Background thread: cek expired paket premium setiap 30 menit.
    Kirim notif pada:
      - 24 jam sebelum expired
      -  3 jam sebelum expired
      -  1 jam sebelum expired
    Setelah expired, reset agar notif bisa terkirim lagi untuk perpanjangan berikutnya.
    """
    _log("EXPIRY", "background aktif — cek tiap 30 menit", Fore.CYAN)
    time.sleep(60)  # Tunggu bot ready

    while True:
        try:
            now = time.time()
            # Reload dari disk agar selalu data terbaru
            current_premium = load_premium()

            for uid_str, data in current_premium.items():
                expired_at = data.get("expired", 0)
                tier       = data.get("tier", "free")
                if tier == "free" or expired_at <= 0:
                    continue

                sisa = expired_at - now

                if sisa <= 0:
                    # Sudah expired — bersihkan state agar notif bisa jalan lagi setelah perpanjang
                    _notif_expiry_sent.pop(uid_str, None)
                    continue

                sent = _notif_expiry_sent.setdefault(uid_str, set())

                # 24 jam sebelum
                if sisa <= 86400 and "24h" not in sent:
                    sent.add("24h")
                    _log("EXPIRY", f"24j — uid={uid_str} tier={tier}", Fore.YELLOW)
                    _send_expiry_notif(int(uid_str), tier, sisa, "24h")

                # 3 jam sebelum
                if sisa <= 10800 and "3h" not in sent:
                    sent.add("3h")
                    _log("EXPIRY", f"3j — uid={uid_str} tier={tier}", Fore.YELLOW)
                    _send_expiry_notif(int(uid_str), tier, sisa, "3h")

                # 1 jam sebelum
                if sisa <= 3600 and "1h" not in sent:
                    sent.add("1h")
                    _log("EXPIRY", f"1j — uid={uid_str} tier={tier}", Fore.RED)
                    _send_expiry_notif(int(uid_str), tier, sisa, "1h")

        except Exception as e:
            _log("EXPIRY", f"loop crash: {e}", Fore.RED)

        time.sleep(1800)  # Cek ulang setiap 30 menit


def account_worker(acc):
    """
    Thread mandiri per akun — polling loop dengan adaptive sleep.
    - SMS ditemukan    → langsung poll TANPA delay (0s) secepat kilat
    - Tidak ada SMS    → mundur bertahap hingga 3s (hemat CPU & koneksi)
    - Error / session  → backoff lebih panjang hingga 10s
    """
    email      = acc.get("email", "")
    sleep_time = 1.0   # interval awal
    while True:
        try:
            found = poll_one_account(acc)
            if found:
                sleep_time = 0.0   # ada SMS — LANGSUNG poll lagi, zero delay!
            else:
                sleep_time = min(sleep_time + 0.3, 3.0)  # max 3s (dari 5s)
        except Exception as e:
            _log("WORKER", f"error [{email}]: {e}", Fore.RED)
            sleep_time = min(sleep_time * 2, 10.0)   # error → backoff
        if sleep_time > 0:
            time.sleep(sleep_time)


# ================= BOT MANAGER (state sync + thread manager) =================
def run_bot():
    global _premium_acc_cache, _cache_dirty, _last_cache_save, _force_bot_sync
    _account_threads = {}   # email -> Thread
    _last_sync       = 0.0

    _log("BOT-MGR", "per-account threading aktif", Fore.CYAN)

    while True:
        try:
            now = time.time()

            # ---- Sync state setiap 30 detik ATAU saat _force_bot_sync=True ----
            if now - _last_sync >= 30 or _force_bot_sync:

                # Rebuild email → user_id mapping
                new_email_to_uid = {}
                with accounts_lock:
                    for acc in accounts:
                        new_email_to_uid[acc["email"]] = OWNER_ID

                users_data = load_users()
                owner_emails = set(new_email_to_uid.keys())
                for uid_str, udata in users_data.items():
                    try:
                        uid_int = int(uid_str)
                    except Exception:
                        continue
                    for em in udata.get("emails", []):
                        if em not in owner_emails:
                            new_email_to_uid[em] = uid_int

                # Sync premium account sessions (cookie-based + password-based user accounts)
                prem_cookies = load_premium_cookies()
                active_prem_emails = set()
                for uid_str, udata in users_data.items():
                    # --- Cookie-based (legacy / manual cookie) ---
                    for em in udata.get("emails", []):
                        if em in owner_emails or em not in prem_cookies:
                            continue
                        active_prem_emails.add(em)
                        if em not in _premium_acc_cache:
                            prem_acc = {
                                "email": em, "password": None,
                                "cookies": prem_cookies[em],
                                "session": make_httpx_client(),
                                "last_login": 0, "csrf_token": "",
                            }
                            prem_acc["session"].cookies.update(prem_cookies[em])
                            _premium_acc_cache[em] = prem_acc
                        else:
                            cached = _premium_acc_cache[em]
                            if cached.get("cookies") != prem_cookies[em]:
                                cached["cookies"] = prem_cookies[em]
                                cached["session"].cookies.clear()
                                cached["session"].cookies.update(prem_cookies[em])
                                cached["last_login"] = 0

                    # --- Password-based (via /addemail email password) ---
                    for ua in udata.get("user_accounts", []):
                        em  = ua.get("email", "")
                        pwd = ua.get("password", "")
                        if not em or not pwd or em in owner_emails:
                            continue
                        active_prem_emails.add(em)
                        if em not in _premium_acc_cache:
                            prem_acc = {
                                "email": em, "password": pwd,
                                "cookies": {}, "session": make_httpx_client(),
                                "last_login": 0, "csrf_token": "",
                            }
                            _premium_acc_cache[em] = prem_acc
                        else:
                            # Update password jika berubah
                            _premium_acc_cache[em]["password"] = pwd

                for em in list(_premium_acc_cache.keys()):
                    if em not in active_prem_emails:
                        del _premium_acc_cache[em]

                # Update shared state (atomic dict replace)
                with accounts_lock:
                    all_accs = list(accounts) + list(_premium_acc_cache.values())
                _bot_state["email_to_uid"]   = new_email_to_uid
                _bot_state["total_accounts"] = len(all_accs)

                # Spawn thread baru untuk akun yang belum punya thread / thread mati
                active_emails = set()
                for acc in all_accs:
                    em = acc.get("email", "")
                    if not em:
                        continue
                    active_emails.add(em)
                    t = _account_threads.get(em)
                    if t is None or not t.is_alive():
                        nt = threading.Thread(
                            target=account_worker, args=(acc,),
                            daemon=True, name=f"poll-{em[:25]}"
                        )
                        nt.start()
                        _account_threads[em] = nt
                        _log("THREAD+", f"{em}", Fore.GREEN)

                # Hapus thread untuk akun yang sudah dihapus
                for em in [e for e in list(_account_threads) if e not in active_emails]:
                    del _account_threads[em]

                _last_sync       = now
                _force_bot_sync  = False  # Reset flag setelah sync selesai

            else:
                # Health-check ringan setiap siklus (2 detik) — respawn thread mati
                for em, t in list(_account_threads.items()):
                    if not t.is_alive():
                        _log("THREAD~", f"respawn: {em}", Fore.YELLOW)
                        with accounts_lock:
                            all_now = list(accounts) + list(_premium_acc_cache.values())
                        for acc in all_now:
                            if acc.get("email") == em:
                                nt = threading.Thread(
                                    target=account_worker, args=(acc,),
                                    daemon=True, name=f"poll-{em[:25]}"
                                )
                                nt.start()
                                _account_threads[em] = nt
                                break

            # Flush cache kalau dirty tapi belum sempat tersimpan
            if _cache_dirty and time.time() - _last_cache_save >= 5:
                with _sent_cache_lock:
                    save_sent_cache()
                _last_cache_save = time.time()
                _cache_dirty = False

            # Coba balik ke worker utama kalau sudah lewat cooldown rate-limit
            maybe_recover_primary_worker()

            time.sleep(2)

        except Exception as e:
            _log("BOT-MGR", f"error: {e}", Fore.RED)
            time.sleep(2)

            
# ================= KEEP-ALIVE SERVER =================
import json as _json
from http.server import HTTPServer, BaseHTTPRequestHandler

_bot_start_time = time.time()

class KeepAliveHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?")[0].rstrip("/")

        if path in ("", "/", "/health"):
            # Endpoint ringan untuk UptimeRobot / Railway health check
            body = b"OK"
            self.send_response(200)
            self.send_header("Content-Type", "text/plain")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        elif path == "/status":
            # Endpoint JSON lengkap — untuk monitoring manual
            now = time.time()
            uptime_sec = int(now - _bot_start_time)
            uptime_str = f"{uptime_sec // 3600}h {(uptime_sec % 3600) // 60}m {uptime_sec % 60}s"

            with accounts_lock:
                owner_list = [
                    {
                        "email": a.get("email", ""),
                        "active": now - a.get("last_login", 0) < LOGIN_COOLDOWN,
                        "last_keepalive": int(now - _last_cookie_refresh.get(a.get("email",""), 0))
                    }
                    for a in accounts
                ]
            prem_list = [
                {
                    "email": e,
                    "active": now - a.get("last_login", 0) < LOGIN_COOLDOWN,
                    "last_keepalive": int(now - _last_cookie_refresh.get(e, 0))
                }
                for e, a in _premium_acc_cache.items()
            ]

            data = {
                "status": "running",
                "uptime": uptime_str,
                "uptime_seconds": uptime_sec,
                "owner_accounts": owner_list,
                "premium_accounts": prem_list,
                "total_accounts": len(owner_list) + len(prem_list),
                "keepalive_interval_sec": COOKIE_KEEPALIVE_INTERVAL,
            }
            body = _json.dumps(data, indent=2).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not found")

    def log_message(self, format, *args):
        pass  # Nonaktifkan log HTTP agar console bersih

def run_keepalive():
    port = int(os.environ.get("PORT", 5000))
    HTTPServer.allow_reuse_address = True
    server = HTTPServer(("0.0.0.0", port), KeepAliveHandler)
    _log("SERVER", f"port {port} | /health /status", Fore.CYAN)
    server.serve_forever()

# ================= GRACEFUL SHUTDOWN (Railway SIGTERM) =================
def _graceful_shutdown(signum, frame):
    print("")
    _log("SHUTDOWN", "menyimpan state...", Fore.YELLOW)
    try:
        with _sent_cache_lock:
            save_sent_cache()
    except Exception:
        pass
    _log("SHUTDOWN", "state tersimpan. Bye!", Fore.YELLOW)
    sys.exit(0)

signal.signal(signal.SIGTERM, _graceful_shutdown)
signal.signal(signal.SIGINT,  _graceful_shutdown)

# ================= START BOT =================
# Ambil username bot via getMe untuk link tombol laporan
def _init_bot_username():
    global BOT_USERNAME
    try:
        r = requests.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getMe", timeout=10)
        d = r.json()
        if d.get("ok"):
            BOT_USERNAME = d["result"].get("username", "")
    except Exception as e:
        _log("BOT", f"getMe error: {e}", Fore.YELLOW)
_init_bot_username()
db_init()

_print_banner(BOT_USERNAME)

_log("BASE", f"IVAS → {BASE}", Fore.CYAN)
_log("WORKER", f"Pool ({len(WORKER_POOL)}): " + " | ".join(WORKER_POOL), Fore.CYAN)

threading.Thread(target=run_keepalive,        daemon=True).start()
threading.Thread(target=listen_command,       daemon=True).start()
threading.Thread(target=auto_cookie_refresher,daemon=True).start()
threading.Thread(target=run_auto_backup,      daemon=True).start()
threading.Thread(target=run_expiry_notifier,  daemon=True).start()
threading.Thread(target=run_token_reset,      daemon=True).start()
run_bot()
